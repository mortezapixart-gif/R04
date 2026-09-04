# -*- coding: utf-8 -*-
"""
core/excel_export.py
---------------------
موتور خروجی Excel «داشبوردی» گزارش نهایی پرواز.

هدف طراحی: فایل اکسل نباید یک جدول ساده از اعداد باشد؛ کاربر (دانش‌آموز/مربی
کانون) باید در همان نگاه اول بفهمد

    • راکت به کجا رسید، چقدر سریع بود و چه چیزی خوب/ضعیف کار کرد؛
    • دما و رطوبت و فشار و چگالی هوا «با بالا رفتن از ارتفاع» چه تغییری
      کردند (شیت «پروفایل ارتفاع»)؛
    • سرعت و شتاب در هر مرحله از پرواز (رانش ← سیر صعودی آزاد ← سقوط آزاد ←
      نزول با چتر ← فرود) چطور بوده است (شیت «تحلیل مراحل پرواز»).

پنج ویژگی کلی خروجی:
    ۱) راست‌به‌چپ: همهٔ شیت‌ها rightToLeft -- سربرگ جدول‌ها و محور افقی
       نمودارها آینه‌ای می‌شوند.
    ۲) اعداد فارسی: number_format با پیشوند [DBNum1]؛ مقدار سلول همچنان عددِ
       واقعی می‌ماند (قابل فیلتر/فرمول/Pivot) ولی با ارقام فارسی نمایش داده
       می‌شود. فقط در متن‌های آزاد (عنوان/یادداشت) رقم‌ها دستی فارسی شده‌اند.
    ۳) جدول‌های گرافیکی: TableStyle اکسل (نوار‌های متناوب) + نوار داده
       (Data Bar) برای مقایسهٔ چشمی + مقیاس رنگی (Color Scale) + کارت‌های KPI
       رنگی با حاشیهٔ تأکیدی.
    ۴) نمودارهای *بومی* اکسل (Line/Area/Bar/Doughnut) -- نه تصویر؛ کاربر
       می‌تواند روی نقاط کلیک کند، سری اضافه کند یا نمودار را تغییر دهد.
    ۵) ستون‌های اختیاری (رطوبت/اشعه): اگر فریمور آن‌ها را لاگ کند، نمودار و
       ستونِ مربوطه خودکار به گزارش اضافه می‌شود؛ اگر نباشد فقط یک یادداشت در
       شیت «راهنمای خواندن» می‌آید و گزارش سالم تولید می‌شود.

سازگاری: امضای export_excel(path) مثل نسخهٔ قبلی است؛ بقیهٔ پارامترها
اختیاری‌اند (برای تست خودکار و تم روشنِ چاپی). هیچ خطای داخلی، تولید گزارش را
متوقف نمی‌کند -- هر بخش در try خودش است و نتیجه در شیت راهنما یادداشت می‌شود.
"""
from __future__ import annotations

import datetime
import math
import os
import re
import zipfile
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.text import RichTextProperties
from openpyxl.drawing.text import CharacterProperties, Font as DrawingFont, Paragraph, ParagraphProperties
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

# رنگ‌ها از core/palette.py -- تا رنگ هر پارامتر در داشبورد برنامه، گزارش PDF
# و اکسل دقیقاً یکی باشد (نگاه کنید به سیاست رنگی در همان فایل).
from core.report_text import protect_latin_quantities
from core.palette import (
    ALTITUDE, VELOCITY, ACCEL_TOTAL, PRESSURE, TEMPERATURE, TEMPERATURE_AHT,
    TEMPERATURE_MPU,
    HUMIDITY, RADIATION, COLOR_OK, COLOR_WARN, COLOR_ERROR, COLOR_INFO,
    AXIS_X, AXIS_Y, AXIS_Z, COLOR_MISSING,
)

G0 = 9.80665
AIR_GAS_CONSTANT = 287.05     # J/(kg.K) -- چگالی هوا از فشار و دمای واقعیِ همان لحظه
MAX_CHART_POINTS = 1400       # سقف نقاط سری‌های نمودار (روان‌ماندن فایل و نمودار)
SERIES_SHEET = "دادهٔ نمودارها (فنی)"   # شیتِ منبعِ دادهٔ نمودارها (نامِ قبلی: سری‌های نمودار)
SER_HDR = 3                   # سطرِ سربرگِ این شیت (سطرهای ۱ و ۲ توضیحِ کارکردِ شیت است)

# فونتِ کل فایل = همان خانواده‌ای که خودِ برنامه (ui/style.py -> APP_FONT_FAMILY) و
# گزارش PDF (core/hud_report.py) استفاده می‌کنند؛ فایل‌ها هم در assets/Shabnam*.ttf
# هستند. اکسل برخلاف Qt نمی‌تواند فونتِ همراهِ فایل را بارگذاری کند، پس اگر Shabnam روی
# سیستمِ *بیننده* نصب نباشد، اکسل خودش یک فونتِ فارسیِ جایگزین می‌گذارد (فقط نمایش فرق
# می‌کند؛ هیچ داده‌ای خراب نمی‌شود). رقم‌ها به هر حال با [DBNum1] فارسی می‌مانند.
FA_FONT_NAME = "Shabnam"
MIN_FONT_PT = 11.0              # درخواستِ کاربر: هیچ نوشته‌ای ریزتر از ۱۱ نباشد
_MIN_SZ = int(MIN_FONT_PT * 100)     # همان عدد در one-hundredth point (متنِ نمودار)
# مدلِ عرضِ حروفِ پایین (_display_width) برایِ متنِ ~۹pt کوک شده بود؛ با کفِ ۱۱pt
# همه‌چیز ~۱۰٪ جا می‌خواهد. یک ضریبِ یک‌جا، همهٔ حساب‌ها (عرض ستون، تعدادِ خطِ wrap،
# قدِ سطر، شیبِ راهنمایِ نمودار) را هم‌زمان درست نگه می‌دارد.
_W_SCALE = 11.0 / 9.8
FA_DIGITS = str.maketrans("0123456789", "۰۱۲۳۴۵۶۷۸۹")
FA_COMMA = str.maketrans(",", "٬")
_LRM = "\u200e"      # Left-to-Right Mark -- برای اعدادِ منفیِ داخلِ متنِ فارسی (ببین fa_num)

# ---- واژنامهٔ مشترک با گزارش PDF ----
EVENT_NAMES = {"launch": "پرتاب", "burnout": "پایان رانش", "apogee": "اوج پرواز",
               "parachute": "باز شدن چتر", "landing": "برخورد با زمین"}
EVENT_ORDER = ("launch", "burnout", "apogee", "parachute", "landing")
EVENT_TIP = {
    "launch": "شتاب کل از ۲g عبور می‌کند",
    "burnout": "شتاب به زیر ۱g می‌افتد (موتور تمام شد)",
    "apogee": "بیشینهٔ ارتفاع ثبت‌شده",
    "parachute": "کند شدن ناگهانی سرعت سقوط",
    "landing": "ارتفاع و سرعت نزدیک صفر",
}
EVENT_COLOR = {"launch": COLOR_ERROR, "burnout": COLOR_WARN, "apogee": ALTITUDE,
               "parachute": COLOR_OK, "landing": COLOR_MISSING}

# سه ماژولی که رویِ بورد دما ثبت می‌کنند. نام‌ها «ماژول‌محور» است (نه «محیط/تراشه») تا
# دانش‌آموز بداند عدد از کجا آمده؛ ترتیبِ ستون‌ها هم همین ترتیب است (درخواستِ کاربر):
# اولِ ماژولِ دما و رطوبت، بعدِ ماژولِ فشار و دما، بعدِ ماژولِ ژیروسکوپ.
# واحدِ دما عمداً در هیچ‌کدام از این برچسب‌ها نیست: چهار ستونِ دما پشتِ هم، «°C»
# چهار بار تکرار می‌کرد و سربرگ‌ها شلوغ شده بود (درخواستِ کاربر). واحدِ دما در هر
# شیت *یک بار* نوشته می‌شود (زیرِ نوارِ بخش، و رویِ محورِ عمودیِ نمودار).
T_HUM_MOD = "دمای ماژول دما و رطوبت"      # AHT21B -- هوای داخلِ محفظه
T_BARO_MOD = "دمای ماژول فشار و دما"      # BMP280 -- مرجعِ دمای هوا
T_GYRO_MOD = "دمای ماژول ژیروسکوپ"        # MPU6050 -- گرمایِ خودِ تراشه (Die)
T_AIR = "برآیندِ دمای هوا"                # میانگینِ ماژول‌های محیطی، بدونِ ژیروسکوپ
_UNSET = ("", "انتخاب نشده", "نصب نشده", "متصل نیست", "unknown", "--", "None")


def _sensor_selection() -> Dict[str, str]:
    """مدلِ ماژول‌ها، آن‌طور که کاربر در صفحهٔ «انتخاب سنسور» انتخاب کرده.

    گزارشِ Excel باید بگوید هر ستونِ دما مالِ کدام ماژول است (درخواستِ کاربر)؛ پس
    نامِ انتخابی (مثلاً BME280 یا ICM-20948) مستقیم داخلِ برچسبِ ستون و داخلِ
    راهنمایِ نمودار می‌نشیند. اگر چیزی انتخاب نشده باشد، نامِ پیش‌فرضِ برد نوشته
    می‌شود تا برچسب خالی نماند.
    """
    out = {"baro": "BMP280", "gyro": "MPU6050", "hum": "AHT21B"}
    try:
        from core.data_manager import data_manager
        models = {str(k).upper().strip(): str(v).strip()
                  for k, v in (getattr(data_manager, "sensor_models", {}) or {}).items()}
    except Exception:                    # بدونِ data_manager (تست‌ها/کنسول) هم کار کند
        models = {}

    def clean(key: str, default: str) -> str:
        v = models.get(key, "")
        return v if v and v not in _UNSET else default

    out["baro"] = clean("BMP280", out["baro"])
    out["gyro"] = clean("MPU6050", out["gyro"])
    # ماژول دما و رطوبت هنوز در لیستِ انتخاب نیست؛ اگر روزی اضافه شد، همان نام می‌آید
    out["hum"] = clean("AHT21B", clean("HUMIDITY", clean("ENV", out["hum"])))
    return out


def _temp_labels(mods: Dict[str, str]) -> Dict[str, str]:
    """برچسبِ کاملِ ستون + نامِ کوتاهِ راهنما برایِ هر سه ماژول و برآیند."""
    # بدونِ «(°C)» -- واحدِ دما یک‌جا در توضیحِ هر شیت گفته می‌شود (شلوغیِ تکرار)
    lab = {
        "hum": f"دمای ماژول دما و رطوبت ({mods['hum']})",
        "baro": f"دمای ماژول فشار و دما ({mods['baro']})",
        "gyro": f"دمای ماژول ژیروسکوپ ({mods['gyro']})",
        "air": T_AIR,
    }
    # راهنمایِ نمودار جا بدهد: نامِ کوتاه = خودِ اسمِ ماژول (تازه برچسبِ بلندِ ستون
    # در اکسل دو/سه خط می‌شد و راهنما از زیرِ نمودار بیرون می‌زد).
    short = {lab["hum"]: f"دمای {mods['hum']}", lab["baro"]: f"دمای {mods['baro']}",
             lab["gyro"]: f"دمای {mods['gyro']}", lab["air"]: "برآیندِ دمای هوا"}
    return {"labels": lab, "short": short, "mods": mods}

# ستون‌های اختیاری: به‌محض این‌که فریمور این‌ها را در CSV بنویسد، نمودار/ستونِ
# مربوطه خودکار به گزارش اضافه می‌شود (نام‌های رایج ماژول‌های AHT21B و GUVA-S12SD).
OPTIONAL_SERIES: List[Tuple[str, Tuple[str, ...], str, str]] = [
    # دمای ماژول دما و رطوبت در OPTIONAL_SERIES نیست: ستون‌هایِ دما را _Context
    # می‌سازد تا نامِ ماژول، از «انتخابِ ماژول‌های سنسور» در برنامه خوانده شود.
    ("رطوبت (٪)", ("Humidity", "humidity", "RH", "AHT21B_Humidity"), "درصد", HUMIDITY),
    ("شاخص اشعهٔ فرابنفش UV", ("UV_Index", "UVI", "UV"), "شاخص", RADIATION),
    ("تابش خورشید (W/m²)", ("Irradiance", "Radiation", "Solar_Wm2", "Irradiance_Wm2"),
     "وات بر متر²", RADIATION),
]

PHASE_COLORS = {
    "روی سکو": COLOR_MISSING,
    "مرحلهٔ رانش": COLOR_ERROR,
    "سیر صعودی آزاد": COLOR_WARN,
    "سقوط آزاد": COLOR_INFO,
    "نزول با چتر": COLOR_OK,
    "فرود": COLOR_MISSING,
}
PHASE_ORDER = ["مرحلهٔ رانش", "سیر صعودی آزاد", "سقوط آزاد", "نزول با چتر", "فرود"]
PHASE_TIP = {
    "روی سکو": "راکت ساکن روی سکو؛ شتاب‌سنج فقط گرانش (۱g) را نشان می‌دهد.",
    "مرحلهٔ رانش": "موتور روشن است؛ سرعت با شتاب زیاد بالا می‌رود و Max-Q معمولاً این‌جاست.",
    "سیر صعودی آزاد": "سوخت تمام شده ولی راکت با اینرسی بالا می‌رود؛ سرعت به‌تدریج کم می‌شود.",
    "سقوط آزاد": "سرعت سقوط رو به افزایش است تا لحظهٔ باز شدن چتر.",
    "نزول با چتر": "پس از ضربهٔ باز شدن، سرعت به مقدار پایدار نزول می‌رسد.",
    "فرود": "راکت روی زمین است؛ داده‌ها باید نزدیک صفر باشند.",
}
PERSIAN_HEADERS = {
    "Time": "زمان (ثانیه)", "Altitude": "ارتفاع (متر)", "Pressure": "فشار (hPa)",
    "Temperature": T_BARO_MOD, "Temperature_MPU": T_GYRO_MOD,
    "AHT21B_Temperature": T_HUM_MOD, "Temperature_AHT21B": T_HUM_MOD,
    # «Temperature_AHT» دقیقاً نام ستونی است که شبیه‌ساز حالت آموزشی می‌نویسد
    "Temperature_AHT": T_HUM_MOD,
    "AHT_Temperature": T_HUM_MOD, "Humidity_Temperature": T_HUM_MOD,
    "AHT_Temp": T_HUM_MOD, "Temp_AHT": T_HUM_MOD,
    "AccelX": "شتاب X", "AccelY": "شتاب Y", "AccelZ": "شتاب Z",
    "GyroX": "چرخش X", "GyroY": "چرخش Y", "GyroZ": "چرخش Z",
    "Latitude": "عرض جغرافیایی", "Longitude": "طول جغرافیایی",
    "GPS_Altitude": "ارتفاع GPS (متر)", "Voltage": "ولتاژ (V)",
    "Humidity": "رطوبت (٪)", "UV_Index": "شاخص UV", "UVI": "شاخص UV",
    "Irradiance": "تابش (W/m²)", "Radiation": "تابش (W/m²)",
}


# =========================================================================
# فارسی‌سازی اعداد
# =========================================================================
def fa(text: Any) -> str:
    """تبدیل ارقام لاتین یک متن به ارقام فارسی (برای عنوان‌ها و یادداشت‌ها)."""
    return str(text).translate(FA_DIGITS)


def fa_num(value: Any, decimals: int = 1, dash: str = "--") -> str:
    """عدد -> رشته با ارقام و جداکنندهٔ هزارگان فارسی (برای متن‌های آزاد).

    عددِ منفی با دو علامتِ LRM (\u200e) بسته می‌شود: الگوریتمِ دوسوییِ (bidi) اکسل
    در پاراگرافِ راست‌به‌چپ، منفیِ ابتدای عدد را به انتهایش می‌برد و «۶٫۵۰-»
    می‌سازد (همان «به‌هم‌ریختگی» که کاربر دیده بود). LRM آن را به‌عنوان یک
    بخشِ چپ‌به‌راست می‌چسباند؛ ضمناً این رشته دیگر «عددِ ذخیره‌شده به‌صورت متن»
    حساب نمی‌شود، پس مثلثِ زردِ اخطار هم نمی‌گیرد.
    """
    try:
        x = float(value)
    except (TypeError, ValueError):
        return dash if value in (None, "") else str(value)
    if not np.isfinite(x):
        return dash
    txt = fa(f"{x:,.{max(decimals, 0)}f}").translate(FA_COMMA)
    return f"{_LRM}{txt}{_LRM}" if txt[:1] in "-+\u2212" else txt


def numfmt(decimals: int = 1) -> str:
    """قالب عددی اکسل با ارقام فارسی؛ مقدار عدد می‌ماند و فقط نمایش فارسی است."""
    tail = "" if decimals <= 0 else "." + "0" * decimals
    return f"[DBNum1]#,##0{tail}"


def _h(color: str) -> str:
    """کوتاه‌کنندهٔ هگز (نام خصوصیِ همان _hex)."""
    return (color or "").lstrip("#").upper()


def _hex(color: str) -> str:
    return _h(color)


# =========================================================================
# پالت رنگ -- الگو: داشبورد چاپیِ روشن (تصویر مرجع 001)
#   زمینهٔ سفید، کارت‌های سفید با کادر طوسیِ مفت، نوارهای بخش نارنجی با متن
#   سفید، متن اصلی سرمه‌ای و رنگ‌های وضعیت قرمز/سبز/کهربایی.
# =========================================================================
REF = dict(
    orange="E97132", navy="1F3864", ink="1A2430", grey="D9D9D9", canvas="FFFFFF",
    card="FFFFFF", zebra="F7F9FC", red="C00000", green="1E8E4E", amber="C8791A",
    slate="5B6B7C", pale="E9EEF4", arc="E4E9F0", canvas2="EFF3F8",
)

# روی زمینهٔ سفید، رنگ‌های نئونی پالتِ برنامه (که برای پس‌زمینهٔ تیره ساخته
# شده‌اند) کم‌رنگ و ناخوانا می‌شوند؛ پس در تم روشن هر رنگ به هم‌خانوادهٔ
# پررنگِ خودش تبدیل می‌شود (هویت رنگی هر پارامتر حفظ می‌شود: ارتفاع همان
# فیروزه‌ای است، فقط تیره‌تر).
LIGHT_RECOLOR = {
    _h(ALTITUDE): "0E8C7E", _h(VELOCITY): "6C3FC9", _h(ACCEL_TOTAL): "D2691E",
    _h(PRESSURE): "B0306A", _h(TEMPERATURE): "D84315", _h(TEMPERATURE_MPU): "A87900",
    _h(TEMPERATURE_AHT): "4E7A1E",
    _h(HUMIDITY): "77578C", _h(RADIATION): "C8791A",
    _h(COLOR_ERROR): "C0392B", _h(COLOR_WARN): "B77B12", _h(COLOR_OK): "1E8E4E",
    _h(COLOR_INFO): "1F6FB2", _h(COLOR_MISSING): "7B8794",
    _h(AXIS_X): "C0392B", _h(AXIS_Y): "1E8E4E", _h(AXIS_Z): "1F6FB2",
}


def _tint(color: str, toward_white: float = 0.72) -> str:
    """رنگ -> نسخهٔ بازِ همان رنگ (برای پُرکردن زیر نمودار سطحی و هایلایت‌ها)."""
    c = _hex(color)
    try:
        r, g, b = int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)
    except (ValueError, IndexError):
        return c
    f = max(0.0, min(1.0, toward_white))
    mix = lambda x: int(round(x + (255 - x) * f))
    return f"{mix(r):02X}{mix(g):02X}{mix(b):02X}"


def _theme(dark: bool) -> Dict[str, str]:
    """تم روشن = پیش‌فرض جدید (الگوی مرجع 001؛ خوانا، کم‌مصرف جوهر، قابل چاپ).
    تم تیره همچنان با dark=True در دسترس است (همان حس داشبورد برنامه)."""
    if dark:
        return dict(dark=True, bg="070B13", panel="0E1620", panel_alt="132130",
                    border="1C3A4A", text="E6F4FB", dim="7E9AAC", grid="24425A",
                    header="16283A", header_text="E6F4FB", band=ALTITUDE.lstrip("#").upper(),
                    band_text="FFFFFF", table_style="TableStyleDark8", axis="8FB0C4",
                    good="35D07F", bad="EF5350", warn="F2C14E", zebra="101A26",
                    arc="1C3A4A", canvas2="0B111B", recolor={})
    return dict(dark=False, bg=REF["canvas"], panel=REF["card"], panel_alt=REF["zebra"],
                border=REF["grey"], text=REF["ink"], dim=REF["slate"], grid="E6EAF0",
                header=REF["navy"], header_text="FFFFFF", band=REF["orange"],
                band_text="FFFFFF", table_style="TableStyleLight8", axis="33475B",
                good=REF["green"], bad=REF["red"], warn=REF["amber"], zebra=REF["zebra"],
                arc=REF["arc"], canvas2=REF["canvas2"], recolor=LIGHT_RECOLOR)


def _paint(theme: Dict[str, str], color: str) -> str:
    """رنگ منحنی/بار را با تم هماهنگ می‌کند (روشن -> نسخهٔ پررنگ همان رنگ)."""
    if not color:
        return color
    return "#" + theme["recolor"].get(_hex(color), _hex(color))


# =========================================================================
# ابزارهای سطح‌پایین
# =========================================================================
def _fill(color: Optional[str]) -> PatternFill:
    return PatternFill("solid", fgColor=_hex(color) if color else "00000000")


def _font(theme: Dict[str, str], size: float = 10, bold: bool = False,
          color: Optional[str] = None, italic: bool = False) -> Font:
    # کفِ سراسری: «از ۱۱ کوچک‌تر نباشه» -- یک‌جا همین‌جا اعمال می‌شود تا هیچ
    # سازندهٔ شیتی (و هیچ فراخوانِ فراموش‌شده‌ای) نتواند زیرِ ۱۱ برود.
    return Font(name=FA_FONT_NAME, size=max(float(size), MIN_FONT_PT), bold=bold,
                italic=italic, color=_hex(color or theme["text"]))


def _side(color: str, style: str = "thin") -> Side:
    return Side(style=style, color=_hex(color))


def _box(theme: Dict[str, str], style: str = "thin") -> Border:
    s = _side(theme["border"], style)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(horiz: str = "right", vert: str = "center", wrap: bool = False,
           indent: int = 0, rtl: bool = True) -> Alignment:
    """چیدمان سلول. `rtl=True` یعنی جهت پاراگراف راست‌به‌چپ (`readingOrder=2`).

    بدون آن، اکسل پاراگراف را «Context» می‌گیرد و روی شیتِ RTL باز هم ترتیبِ
    کاراکترهای دوسویی (پرانتز، خطِ منفی، واحد لاتین مثل m/s) را به‌هم می‌ریزد؛
    راه‌حلِ دستیِ کاربر «راست‌چین کردن» بود، که اینجا خودکار اعمال می‌شود.
    """
    return Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap, indent=indent,
                     readingOrder=(2 if rtl else 1))


def _init_sheet(ws, theme: Dict[str, str], *, gridlines: bool = False, zoom: int = 100,
                landscape: bool = True, tab_color: Optional[str] = None,
                bg_rows: int = 60, bg_cols: int = 20):
    """چیدمان RTL + زمینه + تنظیمات چاپ. در تم روشن زمینه سفید و بدون خط‌کشی
    شبکه است و «ساختار» را کارت‌ها (_panel) می‌سازند، نه خطوط اکسل.

    توجه: freeze_panes در این ماژول عمداً فقط «سطری» (A<row>) ست می‌شود؛ هر
    xSplit (مثلاً "C4") یک خطِ عمودیِ تمام‌قد روی شیت می‌کشد که کاربر آن را «خطِ
    اضافهِٔ بی‌دلیل» می‌بیند و پاک کردنش را خواسته است.
    """
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = gridlines
    ws.sheet_view.zoomScale = zoom
    if tab_color:
        ws.sheet_properties.tabColor = _hex(_paint(theme, tab_color))
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    ws.sheet_format.defaultRowHeight = 16.5
    if theme["dark"]:
        for r in range(1, bg_rows + 1):
            for c in range(1, bg_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = _fill(theme["bg"])
    return ws


def _widths(ws, widths: Sequence[float], first_col: int = 1):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(first_col + i)].width = w


def _display_width(text: Any, decimals: int = 1, fa_digits: bool = True) -> float:
    """پهنای تقریبیِ متنِ *نمایش‌یافته* بر حسب کاراکتر.
    ارقام فارسیTahoma در اکسل حدود ۱٫۵ برابر لاتین جا می‌گیرند؛ همین باعث
    شده بود ستون‌های عددی «به هم بچسبند» -- پس عرض ستون از همین‌جا حساب می‌شود."""
    txt = str(text).replace(_LRM, "").replace("\u202A", "").replace("\u202C", "").replace("\u00A0", " ")
    if fa_digits and isinstance(text, (int, float)) and not isinstance(text, bool):
        txt = fa_num(text, decimals).replace(_LRM, "")
    wide = sum(1 for ch in txt if ch.isalpha() and ord(ch) > 0x0590)   # حروف فارسی
    digits = sum(1 for ch in txt if ch.isdigit())
    other = len(txt) - wide - digits
    # _W_SCALE: مدلِ بالا برایِ ~۹pt کوک شده؛ با کفِ ۱۱pt هر کاراکتر ~۱۰٪
    # جاِ بیشتر می‌خواهد (سربرگِ wrap‌شده، سطرهایِ بلند و راهنمایِ نمودار هم از همین
    # عدد پیروی می‌کنند، پس عرض و قد با هم درست می‌شوند).
    return (wide * 1.25 + digits * (1.62 if fa_digits else 1.0) + other * 0.95) * _W_SCALE


def _autosize(ws, first_col: int, headers: Sequence[str], rows: Sequence[Sequence[Any]],
              col_fmt: Optional[Dict[int, int]] = None, *, text_cols: Sequence[int] = (),
              min_w: float = 10.0, max_w: float = 21.0, text_max_w: float = 46.0,
              pad: float = 2.6, head_lines: int = 1) -> List[float]:
    """عرض هر ستون جدول را از محتوای واقعی‌اش می‌سازد (جلوگیری از چسبیدن اعداد).

    `head_lines`: اگر بیشتر از ۱ باشد، فرض می‌کنیم سربرگ فارسی می‌تواند در چند خط
    بشکند؛ پس عرضِ ستون فقط از *عدد* حساب می‌شود و سربرگِ بلند آن را پهن نمی‌کند
    (درخواستِ کاربر: «اعداد باز هم جمع‌تر بشن»). عرضهایِ ساخته‌شده برگردانده می‌شود
    تا سازندهٔ شیت بتواند قدِ سطرِ سربرگ را از همان‌ حساب کند.
    """
    col_fmt = col_fmt or {}
    n = len(headers)
    made: List[float] = []
    for j in range(n):
        col = first_col + j
        dec = col_fmt.get(j, 1)
        w = _display_width(headers[j], dec) / 1.15 / max(1, int(head_lines))
        for row in rows:
            if j >= len(row):
                continue
            v = row[j]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                w = max(w, _display_width(float(v), dec))
            elif v is not None and v != "":
                w = max(w, min(text_max_w, _display_width(str(v)[:34], 0)))
        # کف/سقف هم مقیاس می‌شوند: مدلِ w با _W_SCALE بزرگ شده، پس مثلاً
        # max_w=16 (جدولِ پروفایل) با فونتِ ۱۱ عددِ «−۲۶٫۰» را می‌بُرد و اکسل ###
        # نشان می‌دهد. تناسَبِ درخواستِ «جمع‌تر شدنِ اعداد» دست‌نخورده می‌ماند.
        lim = (text_max_w if j in text_cols else max_w) * _W_SCALE
        made.append(round(max(min_w * _W_SCALE, min(lim, w + pad)), 1))
        ws.column_dimensions[get_column_letter(col)].width = made[-1]
    return made


_PERSIAN_RE = re.compile(r"[\u0600-\u06FF]")
_NEG_RUN_RE = re.compile(r"(?<!\u200e)([-−]\s*[\u06f0-\u06f90-9][\u06f0-\u06f90-9٬.,]*)")


def _cm_across(ws, c_from: int, c_to: int) -> float:
    """عرضِ سانتی‌متریِ دقیقِ بازهٔ ستون‌های [c_from, c_to] (۱‌مبنا، inclusive).

    چرا لازم است: اگر عرضِ نمودار «چشمی» انتخاب شود (مثلاً ۲۲٫۵cm روی ستونِ B)،
    نمودار از نصفهٔ یک ستونِ پهن بیرون می‌زند و روی نمودارِ کناری می‌افتد --
    همان «نوشته‌ها داخل نمودار هستند» که کاربر می‌بیند. با این تابع هر نمودار
    دقیقاً تا آخرِ ستونِ خودش کش می‌آید و دو نمودار هرگز هم‌پوشان نمی‌شوند.
    """
    px = sum(round((ws.column_dimensions[get_column_letter(c)].width or 11.0) * 7) + 5
             for c in range(c_from, c_to + 1))
    # ۰٫۰۱cm (≈۰٫۴px) کوتاه‌تر: گردِ cm→px نباید نمودار را به ستونِ *بعدی*
    # بکشاند، وگرنه کادرِ نمودار یک ستون بیشتر ازِ خودِ نمودار حساب می‌شود و
    # نوشتهٔ کناری بی‌دلیل «داخلِ نمودار» گزارش/حذف می‌شود.
    return round(px / 96.0 * 2.54 - 0.01, 3)


def _chart_span(ws, r0: int, c0: int, w_cm: float, h_cm: float) -> Tuple[int, int]:
    """(سطر_پایان، ستون_پایان)ِ *inclusive* از کادری که `w_cm × h_cm` را می‌پوشاند.

    این عدد دیگر «تقریبِ اندازهٔ تصویر» نیست، بلکه **محدودهٔ سطر/ستونِ خودِ کادر**
    است: `_place_chart` همان بازه را به‌عنوان `twoCellAnchor` می‌نویسد، پس آنچه در
    اکسل رندر می‌شود با آنچه نگهبانِ چیدمان حساب می‌کند *یکی* است (نه دو مدلِ
    نزدیک‌به‌هم). سطر و ستونِ لنگر هم جزوِ کادرند (اولین چیزی که نمودار می‌پوشاند).
    """
    wpx, hpx = w_cm / 2.54 * 96.0, h_cm / 2.54 * 96.0
    acc, c = 0.0, c0
    while acc < wpx and c <= 16000:
        acc += round((ws.column_dimensions[get_column_letter(c)].width or 11.0) * 7) + 5
        c += 1
    acc, r = 0.0, r0
    while acc < hpx and r <= 1000:
        acc += (ws.row_dimensions[r].height or ws.sheet_format.defaultRowHeight or 16.5) / 72.0 * 96.0
        r += 1
    return max(r0, r - 1), max(c0, c - 1)


_CHART_BOXES: Dict[str, List[Tuple[int, int, int, int]]] = {}


def _cell_hidden(ws, cell) -> bool:
    dim = ws.column_dimensions.get(cell.column_letter)
    return bool(getattr(dim, "hidden", False)) or bool(getattr(ws.row_dimensions.get(cell.row), "hidden", False))


def _chart_bottom(ws, fallback_row: int) -> int:
    """آخرین سطری که *آخرین* نمودارِ این شیت واقعاً می‌پوشاند.

    کارت‌ها باید به این عدد قد بدهند، نه به تقریبِ «۱۱٫۴cm ≈ ۲۰ سطر»: سطرهایِ
    جدول‌ها ۳۰ تا ۳۴pt هستند و همان نمودار روی آن‌ها ۹ سطر جا می‌گیرد -- با
    عددِ دستی یا نمودار از کارت بیرون می‌زد (نوشتهٔ بعدی زیرش) یا فاصلهٔ مفت.
    """
    boxes = _CHART_BOXES.get(ws.title) or []
    return boxes[-1][2] if boxes else fallback_row


def _over_chart(ws, row: int, c1: int, c2: int) -> bool:
    """True اگر سطرِ درخواستی داخلِ کادرِ هر نمودارِ همین شیت باشد."""
    for r0, c0, r1, rc1 in _CHART_BOXES.get(ws.title, ()):
        if r0 <= row <= r1 and not (c2 < c0 or c1 > rc1):
            return True
    return False


def _chart_box_clash(boxes, r0: int, c0: int, r1: int, c1: int) -> bool:
    """برخوردِ کادرِ پیشنهادی با هر کادرِ ثبت‌شده (همه ۱‌مبنا و inclusive)."""
    for b0, b1, b2, b3 in boxes:
        if not (r1 < b0 or b2 < r0 or c1 < b1 or b3 < c0):
            return True
    return False


def _rows_after(ws, r0: int, cm: float) -> int:
    """چند سطر از سطرِ `r0` به بعد، `cm` سانتی‌متر جا می‌گیرند (قدِ *واقعیِ* سطور).

    جایگزینِ `_rows_for_cm` آن‌جاست که فاصله باید *دقیقاً* همان‌قدر باشد: سطرهایِ
    جدول ۳۰ تا ۴۰ پوینت‌اند و با عددِ ثابت، یا فاصله کم می‌آمد یا یک‌باره زیاد.
    """
    need = cm / 2.54 * 96.0
    acc, r = 0.0, r0
    while acc < need and r < 4000:
        acc += (ws.row_dimensions[r].height or ws.sheet_format.defaultRowHeight or 16.5) / 72.0 * 96.0
        r += 1
    return max(1, r - r0)


def _cells_anchor(r0: int, c0: int, r1: int, c1: int) -> TwoCellAnchor:
    """لنگری که کادرِ نمودار را به *سلول‌ها* قفل می‌کند (twoCellAnchor/editAs=oneCell).

    چرا این و نه اندازهٔ سانتی‌متری (oneCellAnchor + ext) که نسخه‌هایِ قبل داشتند:
    آن حالت قدِ ثابتی به تصویر می‌دهد و اکسل همان را روی سطرهایِ بلندِ جدول
    می‌نشاند؛ نتیجه این بود که نمودار ۲ تا ۴ سطر پایین‌تر از چیزی که سازندهٔ شیت
    «رزرو» کرده بود ادامه پیدا می‌کرد و نوشتۀِ زیرش زیرِ تصویر می‌رفت (شکایتِ
    مکررِ «نمودارها و نوشته‌ها توی هم است»). اینجا پایینِ نمودار *همان مرزِ سطرِ
    r1* است؛ پس هر چیزی که از سطرِ r1+1 شروع شود ساختاراً در امان است و لازم نیست
    قدِ نمودار را حدس بزنیم. `chart.width/height` فقط برایِ همین حسابِ سطر/ستون
    به‌کار می‌آید (پیشنهادِ اندازه، با تلورانسِ یک سطر).
    """
    return TwoCellAnchor(
        editAs="oneCell",
        _from=AnchorMarker(col=c0 - 1, colOff=0, row=r0 - 1, rowOff=0),
        to=AnchorMarker(col=c1, colOff=0, row=r1, rowOff=0))


def _place_chart(ws, chart, anchor: str, *, fit_to: Optional[Tuple[int, int]] = None) -> str:
    """نمودار را می‌گذارد و کادرِ سلولی‌اش را ثبت می‌کند (پایهٔ نگهبانِ چیدمان).

    `fit_to=(c_from, c_to)`: عرضِ نمودار *دقیقاً* پهنایِ آن ستون‌ها می‌شود. دلیلش
    باگِ واقعیِ نسخه‌هایِ قبل است: نمودار ۲۲٫۵cm روی شش ستونِ ~۱۷cm‌ای گذاشته
    می‌شد و نصفه‌اش زیرِ نمودارِ کناری می‌رفت؛ کاربر آن را «نوشته‌ها داخل نمودار»
    می‌دید (دو نمودار روی هم). حالا هر نمودار تا آخرِ ستونِ خودش کش می‌آید.

    نگهبانِ دوسویه: (۱) اگر نمودارِ تازه با نمودارِ دیگر برخورد کند، ستون‌به‌ستون
    جابه‌لا می‌شود و گزارش می‌گردد؛ (۲) متنی که *از قبل* زیرِ کادرش باشد گزارش
    می‌شود -- و `_merge` هم از این پس متنِ داخلِ کادرِ نمودار را نمی‌نویسد.

    قدِ نمودار اینجا «سانتی‌متر» نیست، بلکه *سطرهایِ پوشیده‌شده* است (`_cells_anchor`):
    تا اکسل همان کادر را بکشد و هیچ نوشتۀِ زیرِ تصویرِ نمودار نرود.
    """
    ws_anchor = str(anchor)
    m = re.match(r"([A-Z]+)(\d+)$", ws_anchor)
    if not m:
        ws.add_chart(chart, ws_anchor)
        return ws_anchor
    c_start, r_start = column_index_from_string(m.group(1)), int(m.group(2))
    if fit_to:
        chart.width = _cm_across(ws, fit_to[0], fit_to[1])
    boxes = _CHART_BOXES.setdefault(ws.title, [])
    w_cm, h_cm = float(chart.width), float(chart.height)
    c0, r0, moved = c_start, r_start, False
    for _ in range(90):
        r1, c1 = _chart_span(ws, r0, c0, w_cm, h_cm)
        if not _chart_box_clash(boxes, r0, c0, r1, c1):
            break
        moved = True
        if c0 - c_start < 10:
            c0 += 1
        else:                       # جا نبود: زیرِ آخرین نمودارِ همین شیت برو
            last = max((b[2] for b in boxes), default=r1)
            c0, r0 = c_start, last + 2
    if moved:
        _warn(f"چیدمانِ نمودار در {ws.title}",
              ValueError(f"نمودار با کادرِ نمودارِ دیگر برخورد داشت؛ از {ws_anchor} به "
                         f"{get_column_letter(c0)}{r0} منتقل شد"))
        ws_anchor = f"{get_column_letter(c0)}{r0}"
    r1, c1 = _chart_span(ws, r0, c0, w_cm, h_cm)
    boxes.append((r0, c0, r1, c1))
    for row in ws.iter_rows(min_row=r0, max_row=r1, min_col=c0, max_col=c1):
        for cell in row:
            txt = cell.value
            if isinstance(txt, str) and len(txt.strip()) > 2 and not _cell_hidden(ws, cell):
                _warn(f"تداخلِ متن با نمودار در {ws.title}!{cell.coordinate}",
                      ValueError(f"«{txt.strip()[:28]}…» زیرِ نمودار افتاده است"))
                break
    # لنگرِ *سلول‌پایه*: کادر = سطرهای r0..r1 و ستون‌های c0..c1 (نه یک قدِ ثابت که
    # اکسل رویِ سطرهایِ بلند می‌نشاند). add_chart با anchor=None مزاحمِ این نمی‌شود.
    chart.anchor = _cells_anchor(r0, c0, r1, c1)
    ws.add_chart(chart)
    return ws_anchor



def _bidi_safe(text: Any) -> Any:
    """پیچیدنِ هر «−عدد» داخلِ متنِ فارسی با علامتِ LRM.

    اکسل در پاراگرافِ راست‌به‌چپ، علامتِ منفیِ چسبیده به عدد را به آن‌سروی عدد
    می‌فرستد («۶٫۵۰-»)؛ کاربر همین را «بهم‌ریختگیِ پرانتز و اعداد» دیده بود.
    fa_num این کار را برای اعداد خودش می‌کند و اینجا همان محافظ روی متن‌هایِ
    آزادِ دست‌نویس (یادداشت‌ها، راهنما، لیبل‌ها) هم اعمال می‌شود.
    """
    if not isinstance(text, str):
        return text
    if _PERSIAN_RE.search(text):
        text = _NEG_RUN_RE.sub(lambda m: _LRM + m.group(1) + _LRM, text)
    return protect_latin_quantities(text)


# اگر روزی `f`ِ ابتدای یک رشته جا بیفتد (این همان باگی بود که «{_fnum(...)}» را
# داخلِ سلول نوشت)، همان لحظه در یادداشت‌ها دیده می‌شود -- نه در فایلِ کاربر.
_PLACEHOLDER_RE = re.compile(r"\{[^{}\n]*(?:_fnum|fa_num\(|f\.get\(|ctx\.|\.strip\(\)|or 0\)|\* 3\.6)[^{}\n]*\}")


def _merge_clash(ws, r1: int, c1: int, r2: int, c2: int) -> bool:
    """True اگر بازهٔ پیشنهادی با هر mergeِ موجود (یا با محدودهٔ Table) برخورد داشته باشد."""
    for rng in ws.merged_cells.ranges:
        if not (r2 < rng.min_row or rng.max_row < r1 or c2 < rng.min_col or rng.max_col < c1):
            return True
    for tab in list(ws.tables.values()):
        ref = getattr(tab, "ref", None) or ""
        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", str(ref))
        if not m:
            continue
        tc1, tr1 = column_index_from_string(m.group(1)), int(m.group(2))
        tc2, tr2 = column_index_from_string(m.group(3)), int(m.group(4))
        if not (r2 < tr1 or tr2 < r1 or c2 < tc1 or tc2 < c1):
            return True
    return False


def _merge(ws, r1: int, c1: int, r2: int, c2: int, value: Any = None):
    if value is not None and _over_chart(ws, r1, c1, c2):
        # نگهبانِ چیدمان: هیچ نوشته‌ای نباید داخل کادرِ نمودار بیفتد (شکایتِ
        # «نوشته‌ها داخل نمودارها هستند»). اگر سازندهٔ شیت سطرِ اشتباه بدهد،
        # اینجا می‌نویسیم *نه*، و در شیت راهنما گزارش می‌شود -- پس باگِ چیدمان
        # بی‌صدا نمی‌ماند و در هارنسِ تست (notes==[]) هم گرفته می‌شود.
        _warn(f"چیدمان در {ws.title}!{get_column_letter(c1)}{r1}",
              ValueError(f"متنِ «{str(value).strip()[:26]}…» داخل کادرِ نمودار است و نوشته نشد"))
        return ws.cell(row=r1, column=c1)
    if (r1, c1) != (r2, c2):
        # اکسل هر هم‌پوشانیِ merge را «خطا» می‌داند و کل mergeCellsِ آن شیت را
        # حذف می‌کند (پیام: Removed Records: Merge cells). اگر بازه‌ای از قبل
        # ادغام شده باشد، همان‌جا می‌نویسیم و merge تازه نمی‌سازیم.
        if not _merge_clash(ws, r1, c1, r2, c2):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        elif value is not None and ws.cell(row=r1, column=c1).value not in (None, ""):
            # merge نشد (برخورد با merge/جدولِ موجود) و سلول پر است -- پس
            # متنِ فعلی را پاک نمی‌کنیم؛ فراخوانِ دوم فقط آرایش می‌دهد.
            return ws.cell(row=r1, column=c1)
    cell = ws.cell(row=r1, column=c1)
    if value is not None:
        if isinstance(value, str) and _PLACEHOLDER_RE.search(value):
            _warn(f"متن در {ws.title}!{get_column_letter(c1)}{r1}",
                  ValueError("جای‌نگهدارِ f-string باز نشد: «" + value[:48] + "…»"))
        cell.value = _bidi_safe(value)      # «−۶٫۵» در متنِ فارسی وارونه می‌شود
    return cell


def _panel(ws, r1: int, c1: int, r2: int, c2: int, theme: Dict[str, str], *,
           title: str = "", accent: Optional[str] = None, fill: Optional[str] = None,
           title_row_h: float = 19.0) -> int:
    """کارت سفید با کادر طوسیِ مفت و تیترِ سرمه‌ای -- همان واحدِ بصریِ مرجع 001.
    نمودارها و جدول‌ها داخل این کارت‌ها قرار می‌گیرند تا صفحه «لکه‌های بی‌حد»
    به‌نظر نرسد. شمارهٔ ردیفِ بعد از کارت برگردانده می‌شود.
    (مرزها یک‌جا و از نو ساخته می‌شوند؛ انتساب cell.borderِ خوانده‌شده مجاز نیست
    چون StyleProxy در openpyxl قابل hash نیست.)"""
    body = fill if fill is not None else theme["panel"]
    edge = _side(theme["border"], "thin")
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill is None or cell.fill.fill_type is None:
                cell.fill = _fill(body)
            if r == r1 or r == r2 or c == c1 or c == c2:
                cell.border = Border(left=edge if c == c1 else None,
                                     right=edge if c == c2 else None,
                                     top=edge if r == r1 else None,
                                     bottom=edge if r == r2 else None)
    if title and not _merge_clash(ws, r1, c1 + 1, r1, c2):
        t = _merge(ws, r1, c1 + 1, r1, c2, title)
        t.font = _font(theme, 10.5, True, theme["header"] if not theme["dark"] else theme["text"])
        t.alignment = _align("right", indent=1)
        ws.row_dimensions[r1].height = title_row_h
        if accent:
            ws.cell(row=r1, column=c1).fill = _fill(accent)
    elif title:
        # سطرِ اولِ کارت خودش نوارِ تیترِ بخش است (`_section`)؛ اینجا نه merge
        # می‌سازیم (هم‌پوشانی = پیام Repair در اکسل) نه متنِ بخش را پاک می‌کنیم.
        pass
    return r2 + 1


def _rows_for_cm(cm: float, row_pt: float = 16.5) -> int:
    """چند سطرِ اکسل (با ارتفاع پیش‌فرض) به یک ارتفاعِ سانتی‌متری وارد می‌شود.

    همهٔ کارت‌هایی که نمودارِ «بصری» دارند قدشان از همین حساب می‌شود؛ اگر دستی
    عددِ سطر بدهیم، یا نمودار از کارت بیرون می‌زند و روی نوشتهٔ بعدی می‌نشیند
    (شکایتِ «نوشته‌ها داخل نمودارها رفتند») یا فاصلهٔ مفت می‌افتد.
    """
    return max(1, int(math.ceil(cm / 2.54 * 96 / (row_pt / 72.0 * 96.0))))


def _wrap_height(ws, c1: int, c2: int, text: str, line_pt: float = 15.2,
                 minimum: float = 15.0) -> float:
    """ارتفاعِ لازمِ یک سطرِ wrap‌شده در بازهٔ ستون‌های c1..c2 -- چون متن‌های
    فارسیِ بلند اگر فقط یک خط جا داشته باشند، بی‌صدا بریده می‌شوند."""
    avail = sum((ws.column_dimensions[get_column_letter(c)].width or 8.43)
                for c in range(c1, c2 + 1))
    lines = max(1, int(math.ceil(_display_width(text, 1) / max(20.0, avail * 1.06))))
    return max(minimum, line_pt * lines + 4.0)


def _section(ws, row: int, c1: int, c2: int, text: str, theme: Dict[str, str],
             *, sub: str = "") -> int:
    """نوار نارنجیِ تیتر بخش (متن سفید، وسط‌چین) -- امضای بصری مرجع 001."""
    cell = _merge(ws, row, c1, row, c2, text)
    cell.font = _font(theme, 11.5, True, theme["band_text"])
    cell.alignment = _align("center")
    ws.row_dimensions[row].height = 22.5
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = _fill(theme["band"])
    row += 1
    if sub:
        sc = _merge(ws, row, c1, row, c2, sub)
        sc.font = _font(theme, 9, False, theme["dim"])
        sc.alignment = _align("right", wrap=True, indent=1, vert="top")
        ws.row_dimensions[row].height = _wrap_height(ws, c1, c2, sub, 15.2, 16.0)
        row += 1
    return row


def _band_title(ws, row: int, c1: int, c2: int, text: str, theme: Dict[str, str],
                accent: str, subtitle: str = "") -> int:
    """سازگاری با فراخوان‌های قبلی: همان _section (رنگ accent فقط در نشانگر
    ابتدای خط استفاده می‌شود تا همه‌جا یک نوار نارنجی داشته باشیم)."""
    return _section(ws, row, c1, c2, text, theme, sub=subtitle)


def _note(ws, row: int, c1: int, c2: int, text: str, theme: Dict[str, str],
          *, height: int = 24, fill: bool = True, tone: Optional[str] = None,
          size: float = 9.5, bold: bool = False) -> int:
    """خط توضیح/پیشنهاد داخل کارت: متن با رنگ وضعیت (اختیاری) و جداکنندهٔ باز.

    `size`: برایِ متن‌هایی که کاربر باید *از دور* هم بخواند (مثلاً درصدِ سهمِ هر
    مرحله زیر نمودارِ دونات) درشت‌تر داده می‌شود.
    """
    cell = _merge(ws, row, c1, row, c2, text)
    cell.font = _font(theme, size, bold, tone or theme["text"])
    cell.alignment = _align("right", wrap=True, indent=1)
    if fill:
        for c in range(c1, c2 + 1):
            ws.cell(row=row, column=c).fill = _fill(theme["panel"])
            ws.cell(row=row, column=c).border = Border(bottom=_side(theme["grid"]))
    # `height` فقط کفِ ارتفاع است (و سقف ایمنی دارد: یک بار «۱۱٫۴cm» به‌جای
    # «تقریباً ۲ سطر» به این پارامتر داده شد و سطر ۲۰۵pt شد -- سطرِ خالیِ عظیم
    # کنارِ نمودار). ارتفاعِ واقعی را _wrap_height از پهنایِ ستون‌ها حساب می‌کند.
    height = min(max(12.0, float(height)), 90.0)
    line_pt = max(float(size), MIN_FONT_PT) * 1.34
    ws.row_dimensions[row].height = max(height, _wrap_height(ws, c1, c2, text, line_pt, height))
    return row + 1


def _kpi_card(ws, row: int, col: int, span: int, label: str, value: Any, unit: str,
              theme: Dict[str, str], accent: str, *, decimals: int = 1,
              hint: str = "", tone: Optional[str] = None) -> int:
    """کارت KPI سبکِ مرجع: برچسب خاکستری بالا، عدد درشت رنگی، و یک خط توضیح
    وضعیت (مثلاً «۵۰٪ از حد مجاز ▲»). بدون بلوکِ رنگیِ سنگین -- کادر باز است."""
    top, mid, bot = row, row + 1, row + 2
    accent = _paint(theme, accent)
    lab = _merge(ws, top, col, top, col + span - 1, label)
    lab.font = _font(theme, 9, True, theme["dim"])
    lab.alignment = _align("center")
    val = _merge(ws, mid, col, mid, col + span - 1)
    numeric = (isinstance(value, (int, float)) and not isinstance(value, bool)
               and np.isfinite(float(value)))
    if numeric:
        val.value = float(value)
        val.number_format = numfmt(decimals)
    else:
        val.value = "--" if value is None or (isinstance(value, float) and not np.isfinite(value)) else str(value)
    # اندازهٔ فونت عدد از پهنای واقعیِ کارت حساب می‌شود تا عدد به کارت کناری
    # نزند (مشکل «اعداد تو هم» در نسخهٔ قبلی).
    total_w = sum((ws.column_dimensions[get_column_letter(c)].width or 11.5)
                  for c in range(col, col + span))
    chars = float(total_w or (11.5 * span))
    est = _display_width(val.value if numeric else (val.value or ""), decimals)
    size = 20.0 if est <= chars * 0.42 else (16.0 if est <= chars * 0.55 else 13.0)
    val.font = _font(theme, size, True, tone or accent)
    val.alignment = _align("center")
    uni = _merge(ws, bot, col, bot, col + span - 1, hint or unit)
    uni.font = _font(theme, 8.5, False, tone or theme["dim"])
    uni.alignment = _align("center")
    hair = _side(theme["border"], "thin")
    for r in (top, mid, bot):
        for c in range(col, col + span):
            cell = ws.cell(row=r, column=c)
            cell.fill = _fill(theme["panel"])
            cell.border = Border(left=hair if c == col else None,
                                 right=hair if c == col + span - 1 else None,
                                 top=hair if r == top else None,
                                 bottom=hair if r == bot else None)
    ws.cell(row=top, column=col + span - 1).fill = _fill(accent)  # مربع گوشهٔ کارت
    ws.row_dimensions[top].height = 16
    ws.row_dimensions[mid].height = 30
    ws.row_dimensions[bot].height = 16
    return bot + 1


def _write_table(ws, *, top_row: int, first_col: int, headers: Sequence[str],
                 rows: Sequence[Sequence[Any]], theme: Dict[str, str], table_name: str,
                 col_fmt: Optional[Dict[int, int]] = None, wrap_cols: Sequence[int] = (),
                 autosize: bool = True, text_max_w: float = 46.0,
                 head_h: Optional[float] = 36.0, head_lines: int = 1,
                 min_w: float = 10.0, max_w: float = 21.0, pad: float = 2.6) -> int:
    """جدول با سربرگ سرمه‌ای/متن سفید + نوارهای متناوب باز + عرضِ محاسبه‌شده از
    محتوا (تا اعداد فارسی به هم نچسبند). col_fmt: ستون (۰مبنا) -> رقم اعشار."""
    col_fmt = col_fmt or {}
    hdr = top_row
    for j, h in enumerate(headers):
        cell = ws.cell(row=hdr, column=first_col + j, value=_bidi_safe(str(h)))
        cell.font = _font(theme, 9, True, theme["header_text"])
        cell.fill = _fill(theme["header"])
        cell.alignment = _align("center", wrap=True)
        cell.border = Border(top=_side(theme["header"], "medium"),
                             bottom=_side(theme["header"], "thin"),
                             left=_side(theme["border"]), right=_side(theme["border"]))
    ws.row_dimensions[hdr].height = head_h     # سربرگِ سه‌سطریِ فارسی جا می‌گیرد

    for i, vals in enumerate(rows):
        r = hdr + 1 + i
        for j, v in enumerate(vals):
            cell = ws.cell(row=r, column=first_col + j)
            numeric = (isinstance(v, (int, float)) and not isinstance(v, bool)
                       and np.isfinite(float(v)))
            if numeric:
                dec = col_fmt.get(j, 1)
                # نویز اعشار (۱۲٫۸۰۰۰۰۰۰۰۰۰۰۱۹) در نوار فرمول زشت است و عرض
                # ستون را هم بی‌دلیل می‌گیرد؛ تا قالب + ۳ رقم گرد می‌شود.
                cell.value = round(float(v), dec + 3)
                cell.number_format = numfmt(dec)
                # درخواستِ کاربر: «اعداد وسط‌چین باشه» -- وسط‌چین (بدونِ تورفتگی)
                # ستون‌هایِ عددی را به سربرگِ وسط‌چینِ همان ستون تراز می‌کند.
                cell.alignment = _align("center")
            else:
                cell.value = "--" if v in (None, "") else _bidi_safe(str(v))
                cell.alignment = _align("center" if j == 0 else "right",
                                        wrap=(j in wrap_cols), indent=0 if j == 0 else 1)
            cell.font = _font(theme, 9)
            cell.border = Border(bottom=_side(theme["grid"]))
            if i % 2 == 1:
                cell.fill = _fill(theme["zebra"])
        if wrap_cols:
            ws.row_dimensions[r].height = 34
    widths = None
    if autosize:
        widths = _autosize(ws, first_col, headers, rows, col_fmt, text_cols=wrap_cols,
                           text_max_w=text_max_w, min_w=min_w, max_w=max_w, pad=pad,
                           head_lines=head_lines)
    if head_h is None:
        # قدِ سطرِ سربرگ از *تعدادِ خطِ واقعیِ* هر برچسب (با پهنایِ نهاییِ ستون) --
        # نه یک عددِ دستی که همیشه یا کم است یا نصفِ سطر خالی.
        lines = 1
        for j, h in enumerate(headers):
            w = (widths[j] if widths and j < len(widths)
                 else (ws.column_dimensions[get_column_letter(first_col + j)].width or 10.0))
            # همان مدلِ عرضِ _autosize (تقسیمِ ۱٫۱۵) تا «تعدادِ خط» با عرضی که
            # ستون گرفته سازگار بماند و سربرگ در اکسل نصفه نماند.
            need = _display_width(str(h), 0) / 1.15
            lines = max(lines, int(-(-need // max(2.0, float(w) - 1.2))))
        head_h = min(96.0, max(28.0, lines * 15.2 + 6.0))
        ws.row_dimensions[hdr].height = head_h
    last_row = hdr + len(rows)
    if rows:
        ref = (f"{get_column_letter(first_col)}{hdr}:"
               f"{get_column_letter(first_col + len(headers) - 1)}{last_row}")
        try:
            tab = Table(displayName=table_name, ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name=theme["table_style"], showRowStripes=False,
                                                showFirstColumn=False, showLastColumn=False,
                                                showColumnStripes=False)
            ws.add_table(tab)
        except Exception:
            # استایل جدول شرطی است (نام تکراری/سربرگ تکراری...)؛ جدول ساده هم
            # خواناست، پس نباید کل گزارش به‌خاطر آن شکست بخورد.
            pass
    return last_row + 1


# =========================================================================
# اخطارِ «Number stored as text» (مثلثِ زردِ گوشهٔ سلول)
# =========================================================================
_IGNORED_ERRORS_XML = (
    '<ignoredErrors><ignoredError sqref="A1:XFD1048576" numberStoredAsText="1" '
    'twoDigitTextYear="1" evalError="1" unlockedFormula="1" emptyCellReference="1" '
    'listDataValidation="1" calculatedColumn="1"/></ignoredErrors>'
)
# در CT_Worksheet ترتیبِ اجزا الزامی است؛ ignoredErrors باید *قبل* از این‌ها بیاید
# (وگرنه اکسل دوباره پیام Repair می‌دهد).
_AFTER_IGNORED = ("</worksheet>", "<drawing ", "<drawing>", "<legacyDrawing", "<picture",
                  "<oleObjects", "<controls", "<webPublishItems", "<tableParts", "<extLst")


def _patch_theme_font(path: str, font_name: str) -> bool:
    """نامِ فونتِ تمِ ورک‌بوک (minorFont/majorFont) را به همان فونتِ فایل تغییر می‌دهد.

    چرا لازم است: هر چیزی که در فایل *استایلِ* صریح ندارد (متنِ پیش‌فرضِ نمودارها،
    سلول‌های تازه‌ای که کاربر خودش اضافه می‌کند، سرتیترهای تم) از تم خوانده می‌شود و
    آنجا «Calibri» نوشته شده بود. سبکِ Normal *عمداً* دست‌نخورده می‌ماند، چون اکسل
    پهنایِ ستون‌ها را با رقمِ همان سبک می‌شمارد و عوض‌کردنش همهٔ عرض‌هایِ محاسبه‌شده
    را ~۱۵٪ جابه‌جا می‌کرد (چیدمانِ تأییدشده خراب می‌شد).
    """
    from shutil import copyfile
    with zipfile.ZipFile(path) as zin:
        names = zin.namelist()
        parts = {n: zin.read(n) for n in names}
    tgt = "xl/theme/theme1.xml"
    if tgt not in parts:
        return False
    xml = parts[tgt].decode("utf-8")
    out = re.sub(r'(<a:latin typeface=")(?:Calibri|Cambria)(")', rf"\g<1>{font_name}\g<2>", xml)
    out = out.replace('<a:ea typeface=""/>', f'<a:ea typeface="{font_name}"/>')
    if out == xml:
        return False
    parts[tgt] = out.encode("utf-8")
    tmp = path + ".theming"
    copyfile(path, tmp)
    try:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                zout.writestr(name, parts[name])
        os.replace(tmp, path)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise
    return True


def _patch_ignored_errors(path: str) -> int:
    """<ignoredErrors> را به‌صورت XMLِ خام به همهٔ شیت‌ها اضافه می‌کند.

    چرا دستی: مقداری از اعداد عمداً *متنِ* فارسی‌رقم نوشته می‌شوند (برچسبِ راهنمای
    نمودار «… — ۳۲٪»، «۴۲ نقطه»، «۲۰٫۰ هرتز») تا روی هیچ ویندوزی به ارقامِ لاتین
    برنگردند؛ اکسل روی هر رشتهٔ عددنما مثلثِ زردِ «this number is stored as text»
    می‌گذارد. openpyxl خودِ این جزء را serialize نمی‌کند، پس بعد از save درجش
    می‌کنیم -- دقیقاً همان کاری که «Ignore Error» در خود اکسل می‌کند.
    """
    from shutil import copyfile

    with zipfile.ZipFile(path) as zin:
        names = zin.namelist()
        parts = {n: zin.read(n) for n in names}
    touched = 0
    for name in list(parts):
        if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
            continue
        xml = parts[name].decode("utf-8")
        if "<ignoredErrors" in xml:
            continue
        pos = min((i for i in (xml.find(t) for t in _AFTER_IGNORED) if i >= 0), default=-1)
        if pos < 0:
            continue
        parts[name] = (xml[:pos] + _IGNORED_ERRORS_XML + xml[pos:]).encode("utf-8")
        touched += 1
    if not touched:
        return 0
    tmp = path + ".ignoring"
    copyfile(path, tmp)          # اگر جایی خطا داد، فایل اصلی سالم می‌ماند
    try:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                zout.writestr(name, parts[name])
        os.replace(tmp, path)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise
    return touched


# =========================================================================
# نمودارها
# =========================================================================
_CHART_WARNINGS: List[str] = []


def _warn(where: str, exc: Exception) -> None:
    """خطاهای جزئیِ ظاهری نمودار گزارش را نمی‌شکنند، ولی «بی‌صدا» هم نمی‌مانند:
    در شیت راهنمای خواندن ثبت می‌شوند (الگوی مشابه try/exceptهای خاموش که در
    بازبینی کد به آن اشاره شده بود)."""
    msg = f"{where}: {type(exc).__name__}: {exc}"
    if msg not in _CHART_WARNINGS:
        _CHART_WARNINGS.append(msg)


def _char_props(color: str, size: int, bold: bool = False) -> CharacterProperties:
    """ویژگی‌های متن داخل نمودار. توجه: latin/cs باید شیء Fontِ لایهٔ DrawingML
    باشد (با فیلد typeface، نه name و نه نام فونت به‌صورت رشته) -- وگرنه
    openpyxl TypeError می‌دهد و کل قالب متن نمودار از دست می‌رود."""
    font = DrawingFont(typeface=FA_FONT_NAME)
    # نوشته‌هایِ داخلِ نمودار (محورها، راهنما، تیترِ محور) هم از همان کفِ ۱۱pt
    # پیروی می‌کنند: واحدِ size اینجا «سantisantimeter» است (۱۱۰۰ = ۱۱pt).
    return CharacterProperties(sz=max(int(size), _MIN_SZ), b=bool(bold),
                               solidFill=_hex(color), latin=font, cs=font)


def _chart_text(color: str, size: int, bold: bool = False):
    return RichText(bodyPr=RichTextProperties(), p=[Paragraph(
        pPr=ParagraphProperties(defRPr=_char_props(color, size, bold)),
        endParaRPr=_char_props(color, size, bold))])


def _chart_look(chart, theme: Dict[str, str], title: str, width: float, height: float,
                *, y_title: str = "", legend: bool = True,
                gridlines: bool = False, y2_title: str = "", chunky: bool = False,
                x_skip: int = 0, y_fmt: str = "[DBNum1]#,##0.##",
                legend_pos: str = "b", legend_items: Sequence[str] = ()) -> None:
    """ظاهر مشترک همهٔ نمودارها -- تم روشن: پلتِ سفید روی کارت، نوشتهٔ سرمه‌ای،
    **بدون خطوط شبکه**. (درخواست کاربر: شبکه‌ها نمودار را شلوغ و زشت می‌کردند؛
    مقدارخوانی با برچسب‌های محور + راهنمای زیر نمودار انجام می‌شود.)
    اگر جایی شبکه لازم شد، `gridlines=True` فقط شبکهٔ محور عمودی را برمی‌گرداند.
    """
    # .strip() عمداً است: فاصلهٔ ابتدا/انتها باعث می‌شود openpyxl روی <a:t>
    # پیِ attr بگذارد که طبق schema مجاز نیست و اکسل کل نمودار را حذف می‌کند.
    # title=None/"" ⇒ هیچِ <c:title> نمی‌سازیم: در دونات/پای، اکسل تیتر را *داخل*
    # کادر و روی حلقه می‌انداخت (شکایتِ «نوشته‌ها روی نمودار است»); آن متن را
    # سازندهٔ شیت به‌صورت سلولِ درشت بالای/پایینِ نمودار می‌نویسد.
    if title:
        chart.title = str(title).strip()
    else:
        chart.title = None
    chart.visible_cells_only = False     # دادهٔ ستون‌های پنهان هم رسم شود
    if title:
        try:   # رنگ/اندازهٔ عنوان نمودار
            chart.title.tx.rich.p[0].pPr = ParagraphProperties(
                defRPr=_char_props(theme["header"] if not theme["dark"] else theme["text"], 1100, True))
            chart.title.overlay = False     # «روی» پلت/برچسب‌ها نیفتد
        except Exception as exc:
            _warn("عنوان نمودار", exc)
    chart.width, chart.height = width, height
    chart.style = 2
    chart.display_blanks = "gap"
    axes = [a for a in (getattr(chart, "x_axis", None), getattr(chart, "y_axis", None)) if a is not None]
    # محورهای نمودارهای ترکیبی/محورِ دوم (زیرمجموعه‌ها) هم شبکه دارند؛ اگر آنها
    # پاک نشوند، چند مجموعهٔ خطوطِ روی‌هم «شلوغ‌ترین» حالت ممکن را می‌سازد
    # (همان چیزی که در نمودار بزرگِ داشبورد دیده شد).
    for sub in getattr(chart, "_charts", ()) or ():
        for name in ("x_axis", "y_axis"):
            a = getattr(sub, name, None)
            if a is not None and not any(a is b for b in axes):
                axes.append(a)
    # عنوانِ محورِ افقی *اصلاً* نمی‌گذاریم (پارامترِ x_title حذف شد): اکسل آن را در
    # همان نوارِ برچسب‌هایِ محور می‌نشاند و با برچسب‌هایِ بلندِ فارسی (مثلاً
    # «۹۰۰ تا ۱,۰۵۰») روی هم می‌افتاد -- همان چیزی که کاربر در اسکرین‌شات دید. نامِ
    # خودِ نمودار («... با ارتفاع»، «... نسبت به زمان») و توضیحِ زیرِ کارت، همان
    # اطلاعات را می‌دهند. عنوانِ محورِ عمودی می‌ماند: چرخیده است و با اعدادِ کوتاهِ
    # محورِ عمودی برخورد نمی‌کند.
    if axes and y_title:
        chart.y_axis.title = y_title
    for axis in axes:
        is_main_cat = axis is getattr(chart, "x_axis", None)
        try:
            # بدون شبکه (خواست کاربر). با gridlines=True فقط محور مقادیرِ اصلی
            # یک خطِ طوسیِ کم‌رنگ می‌گیرد؛ شبکهٔ محورِ زمان/دسته‌ها همیشه حذف است.
            axis.majorGridlines = (ChartLines(spPr=GraphicalProperties(
                ln=LineProperties(solidFill=theme["grid"], w=6350)))
                if (gridlines and is_main_cat is False and axis is getattr(chart, "y_axis", None))
                else None)
            axis.minorGridlines = None
            axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=theme["border"], w=6350))
            axis.majorTickMark = "none"
            axis.minorTickMark = "none"
            axis.delete = False
        except Exception as exc:
            _warn("حاشیهٔ محور", exc)
        try:   # نوشته‌های محور: رنگ هماهنگ با تم + ارقام فارسیِ کوتاه
            axis.txPr = _chart_text(theme["axis"], 950)
            # General یعنی «هرچه در سلول است» -- روی محورِ ۱۳نفره یعنی
            # برچسب‌های بلندِ هم‌افتاده؛ پس قالبِ عددیِ کوتاه می‌دهیم.
            axis.numFmt = "[DBNum1]#,##0" if is_main_cat else y_fmt
            if is_main_cat and x_skip:
                axis.tickLblSkip = x_skip
                axis.tickMarkSkip = x_skip
            if axis.title is not None:
                axis.title.tx.rich.p[0].pPr = ParagraphProperties(
                    defRPr=_char_props(theme["dim"], 900, True))
                # <c:overlay> اگر نوشته نشود، برخی نسخه‌های اکسل عنوانِ محور را
                # *روی* برچسب‌هایِ همان محور می‌اندازند؛ صریحاً خاموش می‌کنیم.
                axis.title.overlay = False
        except Exception as exc:
            _warn("عنوان محور", exc)
    if legend and chart.legend is not None:
        # اگر برچسب‌هایِ راهنما در عرضِ نمودار جا نشوند، اکسل آنها را دو/سه‌سطری
        # می‌کند و «روی» خودِ نمودار می‌نشینند (همان شکایتِ کاربر). در این حالت
        # راهنما به کناره می‌رود و جایِ خودش را از پلت می‌گیرد.
        w_px = width / 2.54 * 96.0
        # ۸٫۵pt -> ۱۱pt شد، پس هر کاراکترِ راهنما هم پهن‌تر جا می‌گیرد (ضریبِ ۶٫۸
        # برایِ ۸٫۵pt کوک شده بود). اگر کم برآورد شود، اکسل راهنما را به *رویِ* پلت
        # می‌فرستد و کاربر دوباره «نوشته داخل نمودار» می‌بیند.
        need = sum(26 + 8.4 * _display_width(str(t), 0) for t in legend_items) if legend_items else 0
        chart.legend.position = "r" if (need > w_px * 0.9 and legend_pos == "b") else legend_pos
        chart.legend.overlay = False
        try:
            chart.legend.txPr = _chart_text(theme["axis"], 850)
        except Exception as exc:
            _warn("راهنمای نمودار", exc)
    else:
        chart.legend = None
    try:
        # نمودار روی کارتِ سفید بی‌قاب می‌نشیند (قاب را _panel می‌کشد)
        chart.graphical_properties = GraphicalProperties(
            solidFill=theme["panel"], ln=LineProperties(noFill=True))
        chart.plot_area.graphicalProperties = GraphicalProperties(solidFill=theme["panel"])
    except Exception as exc:
        _warn("پنل نمودار", exc)
    if chunky or isinstance(chart, BarChart):
        try:
            chart.gapWidth = 42 if chunky else 60
        except Exception as exc:
            _warn("فاصلهٔ ستون‌ها", exc)
    if y2_title:   # عنوان/قالب محور دوم (برای نمودارهای ترکیبی)
        try:
            for sub in chart._charts[1:]:
                sub.y_axis.title = y2_title
                sub.y_axis.txPr = _chart_text(theme["axis"], 850)
                sub.y_axis.numFmt = "[DBNum1]General"
                sub.y_axis.majorGridlines = None
        except Exception as exc:
            _warn("محور دوم", exc)


def _gauge_chart(ws, value_row: int, value_col: int, theme: Dict[str, str],
                 color: str, *, width: float = 8.6, height: float = 4.4) -> DoughnutChart:
    """نیم‌دایرهٔ درصدی (گیج) با ترفند استاندارد اکسل: سه برش
    [پُر، خالی، نیمهٔ پنهان] -- نیمهٔ پایین بی‌رنگ است تا فقط کمان بالا دیده شود
    و متنِ وسطِ حلقه از زیر نمودارِ بی‌زمینه پیدا می‌ماند."""
    ch = DoughnutChart()
    labels = Reference(ws, min_col=value_col, min_row=value_row, max_row=value_row + 2)
    data = Reference(ws, min_col=value_col + 1, min_row=value_row - 1, max_row=value_row + 2)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(labels)
    ch.holeSize = 58
    ch.firstSliceAng = 270
    ch.title = None
    ch.legend = None
    fill_col = _hex(_paint(theme, color))
    pts = []
    for idx, (fill_c, nofill) in enumerate(((fill_col, False), (theme["arc"], False), ("FFFFFF", True))):
        gp = GraphicalProperties(noFill=True) if nofill else GraphicalProperties(
            solidFill=fill_c, ln=LineProperties(noFill=True))
        pts.append(DataPoint(idx=idx, spPr=gp))
    try:
        ch.series[0].data_points = pts
    except Exception as exc:
        _warn("گیج", exc)
    try:   # شفاف‌سازی: عددِ وسط حلقه باید از زیر نمودار دیده شود
        ch.graphical_properties = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
        ch.plot_area.graphicalProperties = GraphicalProperties(noFill=True)
    except Exception as exc:
        _warn("زمینهٔ گیج", exc)
    ch.visible_cells_only = False
    ch.width, ch.height = width, height
    return ch


def _short_legend(text: Any) -> str:
    """نامِ سری برایِ راهنمایِ نمودار: واحد و پرانتز حذف می‌شود.

    اکسل راهنما را در همان عرضِ نمودار جا می‌دهد؛ با برچسب‌هایِ بلندِ سربرگ
    («دمای ماژول فشار و دما (BME280)») راهنما یا به کناره می‌پرید و زیرِ
    نمودار دیده نمی‌شد، یا روی پلت می‌ریخت. نامِ کوتاه همان اطلاعات را می‌دهد.
    """
    t = str(text or "").split(" (")[0].strip()
    return t or str(text or "")


def _series(chart, data_ws, col: int, hdr_row: int, first_row: int, last_row: int,
            theme: Dict[str, str], color: str, *, width: float = 2.4, smooth: bool = False,
            marker: Optional[str] = None, marker_size: int = 7, area_fill: bool = False,
            no_line: bool = False, legend_name: Optional[str] = None):
    """یک ستون را به‌عنوان سری نمودار اضافه و رنگ‌آمیزی می‌کند.

    `legend_name`: نامی که در راهنمای نمودار نوشته می‌شود. اگر داده نشود، از
    خودِ سربرگ گرفته و کوتاه می‌شود (واحد و پرانتز حذف) تا راهنما زیر نمودار جا
    شود و به کناره یا روی پلت نیفتد.
    """
    chart.add_data(Reference(data_ws, min_col=col, min_row=hdr_row, max_row=last_row),
                   titles_from_data=True)
    color = _paint(theme, color)
    s = chart.series[-1]
    try:
        name = legend_name or _short_legend(data_ws.cell(row=hdr_row, column=col).value)
        if name:
            s.tx = SeriesLabel(v=str(name))     # متنِ مستقیم، نه ارجاعِ سلول
    except Exception as exc:
        _warn("نامِ راهنمای نمودار", exc)
    gp = GraphicalProperties()
    gp.line = (LineProperties(noFill=True) if no_line else
               LineProperties(solidFill=_hex(color), w=int(width * 12700)))
    if area_fill:
        # سطحِ زیر منحنی: در تم روشن بازِ همان رنگ (پُرِ کامل، نمودار را سنگین
        # و سیاه می‌کرد) و در تم تیره رنگِ کامل با کنتراست بالا
        gp.solidFill = _tint(color, 0.78) if not theme["dark"] else _hex(color)
        gp.line = LineProperties(solidFill=_hex(color), w=int(max(width, 1.6) * 12700))
    if isinstance(chart, BarChart):
        # ستون‌ها باید صریحاً رنگ بگیرند؛ اگر فقط خط ست شود اکسل رنگ پیش‌فرض
        # تم (آبی/نارنجیِ خودش) را انتخاب می‌کند و رنگ‌ها بی‌ربط به راهنما می‌شوند
        gp.solidFill = _hex(color)
        gp.line = LineProperties(noFill=True)
    s.graphicalProperties = gp
    s.marker = (Marker(symbol=marker, size=marker_size,
                      spPr=GraphicalProperties(solidFill=_hex(color),
                                               ln=LineProperties(solidFill=_hex(theme["panel"]), w=6350)))
                if marker else Marker(symbol="none"))
    try:
        s.smooth = bool(smooth)
    except Exception:
        pass
    return s


def _color_slices(series, colors: Sequence[str], theme: Dict[str, str]):
    """هر برش دونات/پای را با رنگ داده‌اش رنگ می‌کند؛ اگر این کار نشود اکسل
    رنگ‌های پیش‌فرض تم را به‌ترتیب می‌دهد و رنگِ «مرحلهٔ رانش» در جدول با رنگ
    همان برش در نمودار فرق می‌کند."""
    try:
        pts = []
        for i, col in enumerate(colors):
            # فقط solidFill؛ خطِ دورِ برش (ln) را اکسل در برخی نسخه‌ها «Drawing
            # shape» نامعتبر می‌شمارد و کل نقاشیِ شیت را حذف می‌کند.
            pts.append(DataPoint(idx=i, spPr=GraphicalProperties(
                solidFill=_hex(_paint(theme, col)))))
        series.data_points = pts
    except Exception as exc:
        _warn("رنگ برش‌ها", exc)


def _categories(chart, data_ws, col: int, first_row: int, last_row: int):
    chart.set_categories(Reference(data_ws, min_col=col, min_row=first_row, max_row=last_row))


def _secondary_axis(main: LineChart, second: LineChart, theme: Dict[str, str], y2_title: str = ""):
    """الگوی رسمی openpyxl برای محور دوم راست‌چین (مثلاً سرعت در برابر ارتفاع)."""
    second.y_axis.axId = 200
    second.y_axis.crosses = "max"
    if y2_title:
        second.y_axis.title = y2_title
        try:      # همان قاعدهٔ _chart_look: عنوانِ محور هیچ‌وقت «روی» برچسب‌ها نیفتد
            second.y_axis.title.overlay = False
        except Exception:
            pass
    try:
        second.y_axis.txPr = _chart_text(theme["axis"], 900)
        second.y_axis.numFmt = "[DBNum1]#,##0.##"
        # محور دوم هم شبکه دارد (پیش‌فرض openpyxl) و روی شبکهٔ محور اول می‌افتد؛
        # همان «کلی خطوط افقی» که نمودار را خراب می‌کرد.
        second.y_axis.majorGridlines = None
        second.y_axis.minorGridlines = None
    except Exception as exc:
        _warn("محور دوم", exc)
    main += second


# =========================================================================
# زمینهٔ مشترک: آماده‌سازی داده (یک‌بار؛ بدون تکرار تحلیل در هر شیت)
# =========================================================================
class _Context:
    def __init__(self, mission, motor, results: Dict[str, Any], df: Optional[pd.DataFrame],
                 events: Optional[Dict[str, Any]] = None, suggestions: Optional[List[str]] = None):
        self.mission = mission
        self.motor = motor
        self.results = results or {}
        self.df = df
        self.suggestions = list(suggestions or [])
        self.notes: List[str] = []
        self.events = events or self.results.get("events") or {}
        self.an = None
        self.idx: Dict[str, int] = {}
        self.series: Optional[pd.DataFrame] = None
        self.optional: List[Dict[str, str]] = []
        self.temps: Dict[str, Optional[np.ndarray]] = {}
        self.temp_sources = "--"
        _tl = _temp_labels(_sensor_selection())
        self.temp_mods: Dict[str, str] = _tl["mods"]
        self.t_hum, self.t_baro, self.t_gyro, self.t_air = (
            _tl["labels"]["hum"], _tl["labels"]["baro"], _tl["labels"]["gyro"], _tl["labels"]["air"])
        self.t_short: Dict[str, str] = _tl["short"]      # نامِ کوتاهِ راهنمایِ نمودار
        self.temp_keys: Tuple[str, ...] = (self.t_hum, self.t_baro, self.t_gyro, self.t_air)
        # سربرگِ شیتِ دادهٔ خام: نامِ ستونِ CSV -> فارسیِ ماژول‌محور
        self.raw_headers: Dict[str, str] = {
            **PERSIAN_HEADERS,
            "Temperature": self.t_baro, "Temperature_MPU": self.t_gyro,
            **{k: self.t_hum for k in (f"{self.temp_mods['hum']}_Temperature",
                                       f"Temperature_{self.temp_mods['hum']}",
                                       "Temperature_AHT",      # ستونِ شبیه‌سازِ حالت آموزشی
                                       "AHT21B_Temperature",
                                       "Temperature_AHT21B", "AHT_Temperature", "AHT_Temp",
                                       "Humidity_Temperature", "Temp_AHT")},
            f"{self.temp_mods['baro']}_Temperature": self.t_baro,
            f"Temperature_{self.temp_mods['baro']}": self.t_baro,
            f"{self.temp_mods['gyro']}_Temperature": self.t_gyro,
            f"Temperature_{self.temp_mods['gyro']}": self.t_gyro,
        }
        self.ok = False

        if df is None or len(df) == 0:
            self.notes.append("دادهٔ پروازی (CSV) بارگذاری نشده بود؛ به همین دلیل نمودارها و "
                              "جدول‌های تحلیلی خالی‌اند و فقط مشخصات مأموریت ثبت شده است.")
            return

        from core.analysis import FlightAnalyzer
        try:
            an = FlightAnalyzer(df, mission, motor)
            ev = an.detect_events()
            self.an = an
            self.idx = dict(getattr(an, "_idx", {}) or {})
            self.events = {**ev, **{k: v for k, v in (events or {}).items() if v is not None}}
        except Exception as exc:      # هر خرابی داده -> یادداشت، نه کرش گزارش
            self.notes.append(f"تحلیل خودکار داده میسر نشد ({type(exc).__name__}: {exc})؛ "
                              "شیت «دادهٔ خام» همچنان کامل است.")
            return
        if an.t is None or len(an.t) < 2:
            self.notes.append("ستون Time در فایل CSV یافت نشد یا کم‌تر از دو نمونه داشت؛ "
                              "به همین دلیل نمودارهای زمانی ساخته نشدند.")
            return
        self.ok = True
        self._temp_columns(an)
        self._build_series()

    # ------------------------------------------------------------------
    def _col(self, name: str) -> Optional[np.ndarray]:
        if self.df is None or name not in self.df.columns:
            return None
        return pd.to_numeric(self.df[name], errors="coerce").to_numpy(dtype=float)

    def _build_series(self):
        """سری نمودار = نمونه‌برداری مجدد + ستون‌های محاسبه‌شده (سرعت، شتاب کل،
        چگالی هوا از فشار/دمای واقعی، برچسب مرحله و نشانگر رویدادها)."""
        an, n = self.an, len(self.an.t)
        step = max(1, int(math.ceil(n / MAX_CHART_POINTS)))
        sel = np.arange(0, n, step)
        if sel[-1] != n - 1:
            sel = np.append(sel, n - 1)
        self.sel = sel

        def pick(arr):
            if arr is None:
                return np.full(len(sel), np.nan)
            a = np.asarray(arr, dtype=float)
            if len(a) >= sel[-1] + 1:
                return a[sel]
            return np.pad(a, (0, len(sel) - len(a)), constant_values=np.nan)[: len(sel)]

        frame = pd.DataFrame({
            "زمان (ثانیه)": pick(an.t),
            "ارتفاع (متر)": pick(an.alt),
            "سرعت عمودی (م/ث)": pick(an.vel),
            "شتاب کل (g)": (pick(an.a_total) / G0) if an.a_total is not None else np.full(len(sel), np.nan),
            "فشار (هکتوپاسکال)": pick(an.pressure),
        })
        # سه ماژولی که دما ثبت می‌کنند + برآیندِ دمای هوا (بدونِ ماژولِ ژیروسکوپ)
        for key in self.temp_keys:
            arr = self.temps.get(key)
            if arr is not None:
                frame[key] = pick(arr)
        for src, dst in (("AccelX", "شتاب X (m/s²)"), ("AccelY", "شتاب Y (m/s²)"),
                         ("AccelZ", "شتاب Z (m/s²)"), ("GyroX", "چرخش X (deg/s)"),
                         ("GyroY", "چرخش Y (deg/s)"), ("GyroZ", "چرخش Z (deg/s)"),
                         ("Voltage", "ولتاژ باتری (V)"), ("GPS_Altitude", "ارتفاع GPS (متر)")):
            arr = self._col(src)
            if arr is not None:
                frame[dst] = pick(arr)

        # چگالی هوا از دادهٔ واقعی همان لحظه (به‌جای ثابت ۱.۲۲۵ سطح دریا)
        p_hpa = frame["فشار (هکتوپاسکال)"].to_numpy(dtype=float)
        # دمایِ فرمولِ چگالی = برآیندِ ماژول‌های محیطی؛ اگر برآیند نبود، همان ماژولِ
        # فشار و دما. دمای ماژول ژیروسکوپ هرگز واردِ این محاسبه نمی‌شود.
        t_c = frame.get(self.t_air)
        if t_c is None:
            t_c = frame.get(self.t_baro)
        if t_c is None:
            t_c = frame.get(self.t_gyro)
        if t_c is None:
            t_c = np.full(len(sel), np.nan)
        t_c = np.asarray(t_c, dtype=float)
        with np.errstate(invalid="ignore", divide="ignore"):
            rho = (p_hpa * 100.0) / (AIR_GAS_CONSTANT * (t_c + 273.15))
        frame["چگالی هوا (kg/m³)"] = np.where(np.isfinite(rho) & (rho > 0.05) & (rho < 3.0), rho, np.nan)

        frame["مرحلهٔ پرواز"] = [self.phase_of_index(int(i)) for i in sel]

        # نشانگر رویدادها: سری تک‌نقطه‌ایِ «ارتفاع در لحظهٔ رویداد» -> المان لوزی
        for key in EVENT_ORDER:
            col_name = "نشانگر " + EVENT_NAMES[key]
            vals = np.full(len(sel), np.nan)
            i = self.idx.get(key)
            if i is not None and an.alt is not None and i < len(an.alt):
                j = int(np.argmin(np.abs(sel - i)))
                # باگِ واقعیِ نسخهٔ قبل: `an.alt[j]` قدِ سطرِ *جُم* دادهٔ خام را می‌داد،
                # درحالی‌که `j` جایگاهِ نمونهٔ j-aُم است (sel[j] = j*step) ⇒ لوزی‌ها
                # خیلی پایین‌تر/بالاتر از منحنی می‌نشستند. درست: همان سطرِ منحنی.
                vals[j] = float(an.alt[int(sel[j])])
            frame[col_name] = vals

        # ستون‌های اختیاری (رطوبت/اشعه) -- اگر در CSV باشند
        for label, names, unit, color in OPTIONAL_SERIES:
            for name in names:
                arr = self._col(name)
                if arr is None:
                    continue
                vals = pick(arr)
                if not np.isfinite(vals).any():
                    continue
                frame[label] = vals
                self.optional.append(dict(label=label, column=name, unit=unit, color=color))
                break
        if not self.optional:
            self.notes.append("ستون «رطوبت» (Humidity) یا «اشعه» (UV_Index / Irradiance) در فایل CSV پیدا "
                              "نشد؛ به‌محض این‌که فریمور یکی از این ستون‌ها را در لاگ بنویسد، نمودار و "
                              "ستون مربوطه خودکار به همین گزارش اضافه می‌شود. ستون‌های پذیرفته‌شده: "
                              + fa("، ").join(["Humidity، RH، AHT21B_Humidity",
                                               "UV_Index، UVI، UV",
                                               "Irradiance، Radiation، Solar_Wm2"]) + ".")
        self.series = frame

    # ------------------------------------------------------------------
    def _temp_columns(self, an) -> None:
        """سه ستونِ دما + «برآیندِ دمای هوا».

        قانونِ برآیند (درخواستِ کاربر): ماژولِ ژیروسکوپ (MPU6050) دمای *خودِ تراشه* را
        می‌دهد، نه دمای هوا؛ پس در هیچ میانگینی (برآیند، چگالی هوا، دلتایِ دما،
        میانگینِ هر مرحله) حساب نمی‌شود -- فقط در جدول *نمایش* داده می‌شود.
        """
        def arr(a):
            if a is None:
                return None
            v = np.asarray(a, dtype=float)
            return v if v.size > 1 and np.isfinite(v).sum() >= 2 else None

        n = int(len(np.asarray(an.t, dtype=float)))
        m = self.temp_mods
        # «temp_aht» را تحلیل‌گر خودش می‌خواند (با نامِ دقیقِ ستونِ شبیه‌سازِ حالت
        # آموزشی: Temperature_AHT)؛ اگر نبود، همین‌جا نام‌هایِ دیگر را امتحان می‌کنیم.
        hum = arr(getattr(an, "temp_aht", None))
        baro, gyro = arr(getattr(an, "temp", None)), arr(getattr(an, "temp_mpu", None))
        # نام‌هایِ ممکنِ ستونِ دمای ماژول دما و رطوبت: اسمِ انتخابیِ کاربر هم تست می‌شود
        for name in (f"{m['hum']}_Temperature", f"Temperature_{m['hum']}", "Temperature_AHT",
                     "AHT21B_Temperature", "Temperature_AHT21B", "AHT_Temperature", "AHT_Temp",
                     "Humidity_Temperature", "Temp_AHT"):
            if hum is not None:
                break
            hum = arr(self._col(name))
        if baro is None:      # لاگ ممکن است ستون را با نامِ ماژولِ انتخابی نوشته باشد
            for name in (f"{m['baro']}_Temperature", f"Temperature_{m['baro']}", "Temperature"):
                baro = arr(self._col(name))
                if baro is not None:
                    break
        if gyro is None:
            for name in (f"{m['gyro']}_Temperature", f"Temperature_{m['gyro']}",
                         "Temperature_MPU", "MPU_Temperature", "temp_mpu"):
                gyro = arr(self._col(name))
                if gyro is not None:
                    break
        parts = [a for a in (hum, baro) if a is not None]
        air = None
        if parts:
            # «m» نامِ دیکشنریِ ماژول‌هاست؛ این‌جا stack نام دارد تا سایه نیندازد
            stack = np.vstack([np.pad(a, (0, max(0, n - len(a))), constant_values=np.nan)[:n]
                               for a in parts]) if n else np.vstack(parts)
            with np.errstate(invalid="ignore"):
                air = np.nanmean(np.where(np.isfinite(stack), stack, np.nan), axis=0)
            if not np.isfinite(air).any():
                air = None
        elif gyro is not None:
            air = None          # فقط دمای تراشه هست -> برآیندی برایِ هوا نیست
        self.temps = {self.t_hum: hum, self.t_baro: baro, self.t_gyro: gyro, self.t_air: air}
        got = [m["hum"] if hum is not None else None, m["baro"] if baro is not None else None]
        self.temp_sources = " + ".join(g for g in got if g) or "--"
        # یادداشتِ شفاف: برآیند از کدام ماژول‌هاست و ژیروسکوپ چرا نیست
        self.temp_note = (f"برآیندِ دمای هوا = میانگینِ «{m['hum']}» و «{m['baro']}»؛ "
                          f"دمای «{m['gyro']}» (خودِ تراشه) در هیچ میانگینی حساب نمی‌شود."
                          if hum is not None and baro is not None else
                          f"برآیندِ دمای هوا فقط از «{m['baro']}» ساخته شده (ماژول دما و رطوبت در "
                          f"این لاگ ستونِ دما نداشت)؛ دمای «{m['gyro']}» هم عمداً حساب نمی‌شود.")

    # ------------------------------------------------------------------
    def phase_of_index(self, i: int) -> str:
        idx = self.idx
        if idx.get("launch") is not None and i < idx["launch"]:
            return "روی سکو"
        if idx.get("burnout") is not None and i <= idx["burnout"]:
            return "مرحلهٔ رانش"
        if idx.get("apogee") is not None and i <= idx["apogee"]:
            return "سیر صعودی آزاد"
        if idx.get("parachute") is not None and i <= idx["parachute"]:
            return "سقوط آزاد"
        if idx.get("landing") is not None and i < idx["landing"]:
            return "نزول با چتر"
        return "فرود"

    def metric(self, key: str) -> Optional[float]:
        try:
            v = float(self.results.get(key))
            return v if np.isfinite(v) else None
        except (TypeError, ValueError):
            return None

    def duration(self) -> Optional[float]:
        try:
            return float(self.events.get("landing")) - float(self.events.get("launch"))
        except (TypeError, ValueError):
            return None

    def time_of(self, key: str) -> Optional[float]:
        try:
            t0 = float(self.events.get("launch") or 0.0)
            return float(self.events.get(key)) - t0
        except (TypeError, ValueError):
            return None


# =========================================================================
# شیت ۱: داشبورد پرواز
# =========================================================================
# همان سقف‌های مرجعِ گزارش PDF (core/hud_report.py) -- تا گیج‌ها و حلقه‌های PDF
# یک معنا بدهند: (کلید results، سقف، برچسب، رقم اعشار)
GAUGE_LIMITS = (
    ("max_altitude", 3000.0, "اوج پرواز", "متر", 0),
    ("max_velocity", 200.0, "بیشینهٔ سرعت", "م/ث", 0),
    ("max_g", 20.0, "بیشینهٔ شتاب", "g", 1),
    ("landing_velocity", 25.0, "سرعت فرود", "م/ث", 1),
)

GOOD_LIMIT, WARN_LIMIT = 0.60, 0.90      # Like the PDF rings: سبز < ۶۰٪، کهربایی < ۹۰٪


def _tone_for(text: str, theme: Dict[str, str]) -> Optional[str]:
    """رنگِ نقطهٔ ابتدای یک پیشنهاد -- از روی کلیدواژهٔ متن (فقط برای خوانایی؛
    هیچ محاسبه‌ای به آن وابسته نیست)."""
    t = text or ""
    if any(k in t for k in ("بالاتر", "است", "خطر", "بررسی", "نامتعارف", "بیشتر از")) and \
       not any(k in t for k in ("مطلوب", "ایمن", "نرمال", "عادی")):
        return theme["bad"]
    if any(k in t for k in ("کافی نبود", "بررسی شود", "قابل بررسی", "توصیه")):
        return theme["warn"]
    if any(k in t for k in ("مطلوب", "ایمن", "عادی", "بهینه")):
        return theme["good"]
    return None


def _gauge_strip(ws, ctx: _Context, theme: Dict[str, str], row: int, c1: int, c2: int) -> Tuple[int, int]:
    """ردیف گیج‌های نیم‌دایره‌ای (سبک مرجع 001) -- دادهٔ هر گیج در ستون‌های
    پنهانِ همان شیت نوشته می‌شود تا نمودار بومی اکسل باشد، نه تصویر."""
    items = []
    for key, limit, label, unit, dec in GAUGE_LIMITS:
        v = ctx.metric(key)
        if v is None:
            continue
        items.append((label, v, limit, unit, dec))
    if not items:
        return row, 0
    hidden_c = c2 + 10                       # ستون‌های دادهٔ گیج (پنهان)
    n = len(items)
    span = max(3, int(math.floor((c2 - c1 + 1) / n)))
    r_data = row + 1                          # سرفصل داده‌ها
    for gi, (label, v, limit, unit, dec) in enumerate(items):
        frac = max(0.0, min(1.0, v / limit if limit else 0.0))
        pct = frac * 100.0
        col_lab = hidden_c + gi * 2          # برچسب‌ها این‌جا، عددها در ستون بعد
        pc = ws.cell(row=r_data, column=col_lab + 1, value="٪")
        pc.font = _font(theme, 9, False, theme["text"])   # تا فونتِ فایل یکی بماند
        for k, val in enumerate((pct, 100.0 - pct, 100.0)):
            dcell = ws.cell(row=r_data + 1 + k, column=col_lab + 1, value=float(val))
            dcell.number_format = numfmt(1)
            # فونت صریح: قالبِ سلولی، استایلِ ستون را بی‌اثر می‌کند و سلول به
            # فونتِ پیش‌فرض (Calibri) برمی‌گردد -- این ستون‌ها پنهان‌اند ولی اگر
            # کسی unhide کند، فونتِ فایل نباید فرق کند.
            dcell.font = _font(theme, 9, False, theme["text"])
            if k == 0:
                lc = ws.cell(row=r_data + 1 + k, column=col_lab, value=label)
                lc.font = _font(theme, 9, False, theme["text"])
        for c in (col_lab, col_lab + 1):
            ws.column_dimensions[get_column_letter(c)].hidden = True
            ws.column_dimensions[get_column_letter(c)].width = 8
    row = _section(ws, row, c1, c2, "درصد رسیدن به سقف‌های مرجع (سبز < ۶۰٪، کهربایی < ۹۰٪)", theme)
    panel_top = row
    card_row = panel_top + 1
    for gi, (label, v, limit, unit, dec) in enumerate(items):
        col = c1 + gi * span
        frac = max(0.0, min(1.0, v / limit if limit else 0.0))
        tone = theme["good"] if frac < GOOD_LIMIT else (theme["warn"] if frac < WARN_LIMIT else theme["bad"])
        lab = _merge(ws, card_row, col, card_row, col + span - 1, label)
        lab.font = _font(theme, 9, True, theme["dim"])
        lab.alignment = _align("center")
        ws.row_dimensions[card_row].height = 16
        num = _merge(ws, card_row + 1, col, card_row + 1, col + span - 1)
        num.value = float(v)
        num.number_format = numfmt(dec)
        num.font = _font(theme, 15, True, tone)
        num.alignment = _align("center")
        ws.row_dimensions[card_row + 1].height = 22
        gch = _gauge_chart(ws, r_data + 1, hidden_c + gi * 2, theme, tone)
        _place_chart(ws, gch, f"{get_column_letter(col)}{card_row + 2}", fit_to=(col, col + span - 1))
        foot = _merge(ws, card_row + 10, col, card_row + 10, col + span - 1,
                      f"سقف مرجع {fa_num(limit, 0)} {unit} — {fa_num(frac * 100, 0)}٪")
        foot.font = _font(theme, 8.5, False, tone)
        foot.alignment = _align("center")
        ws.row_dimensions[card_row + 10].height = 16
    panel_bottom = card_row + 11
    for gi in range(len(items)):          # مرجع: هر گیج داخل کادر خودش
        gcol = c1 + gi * span
        _panel(ws, card_row - 1, gcol, card_row + 11, gcol + span - 1, theme,
               fill=theme["panel"])
    _panel(ws, panel_top, c1 - 1, panel_bottom, c2, theme, fill=theme["canvas2"])
    return panel_bottom + 2, len(items)


def _build_dashboard(wb, ctx: _Context, theme: Dict[str, str], sws, cols):
    """شیت اول: داشبورد. چیدمان بر پایهٔ مرجع 001 -- کارت‌های سفید با کادر
    طوسی، تیتر بخش روی نوار نارنجی، اعداد درشت با رنگ وضعیت، و گیج."""
    ws = wb.create_sheet("داشبورد پرواز")
    _init_sheet(ws, theme, tab_color=ALTITUDE)
    C1, C2 = 2, 17
    _widths(ws, [1.7] + [12.6] * 16)
    m = ctx.mission
    r = 2

    # ---- سرصفحه: تیتر سرمه‌ای + نوار نارنجیِ اطلاعات پرواز ----
    for c in range(C1, C2 + 1):
        ws.cell(row=r - 1, column=c).fill = _fill(theme["header"])
    ws.row_dimensions[r - 1].height = 4
    title = _merge(ws, r, C1, r + 1, C1 + 8, "گزارش تحلیل پرواز راکت")
    title.font = _font(theme, 20, True, theme["header"])
    title.alignment = _align("right", indent=1)
    ws.row_dimensions[r].height = 24
    date_text = m.jalali_date
    if not date_text and m.date:
        try:
            from core.jalali import gregorian_date_to_jalali_str
            date_text = gregorian_date_to_jalali_str(datetime.date.fromisoformat(m.date))
        except Exception:
            date_text = m.date
    stamp = _merge(ws, r, C1 + 9, r + 1, C2,
                   f"تولید: {fa(datetime.datetime.now().strftime('%H:%M'))} — "
                   f"{fa(date_text or '--')}")
    stamp.font = _font(theme, 9, False, theme["dim"])
    stamp.alignment = _align("left", vert="bottom")
    r += 2
    meta = _merge(ws, r, C1, r, C2,
                  f"راکت: {m.rocket_name or '--'}     |     شمارهٔ پرواز: {fa(m.flight_number or '--')}     |     "
                  f"محل پرتاب: {m.launch_site or '--'}     |     تاریخ: {fa(date_text or '--')}  ساعت {fa(m.time or '--')}")
    meta.font = _font(theme, 10, True, theme["band_text"])
    meta.alignment = _align("center")
    for c in range(C1, C2 + 1):
        ws.cell(row=r, column=c).fill = _fill(theme["band"])
    ws.row_dimensions[r].height = 21
    r += 2

    # ---- ۱) کارت‌های شاخص ----
    landing_v = ctx.metric("landing_velocity")
    safe = landing_v is not None and 3.0 <= landing_v <= 8.0
    cards = [
        ("اوج پرواز", ctx.metric("max_altitude"), "متر", ALTITUDE, 1,
         f"{fa_num((ctx.metric('max_altitude') or 0) / 1000.0, 2)} کیلومتر"),
        ("بیشینهٔ سرعت", ctx.metric("max_velocity"), "متر بر ثانیه", VELOCITY, 1,
         f"≈ {fa_num((ctx.metric('max_velocity') or 0) * 3.6, 0)} کیلومتر بر ساعت"),
        ("بیشینهٔ شتاب", ctx.metric("max_g"), "جی (g)", ACCEL_TOTAL, 2,
         "شامل ضربهٔ باز شدن چتر"),
        ("زمان کل پرواز", ctx.duration(), "ثانیه", COLOR_INFO, 1,
         f"{fa_num((ctx.duration() or 0) / 60.0, 1)} دقیقه"),
        ("سرعت فرود", landing_v, "متر بر ثانیه", COLOR_OK if safe else COLOR_WARN, 1,
         "در بازهٔ ایمن ۳ تا ۸" if safe else "بازهٔ ایمن: ۳ تا ۸ m/s"),
        ("حداکثر فشار دینامیکی", ctx.metric("dynamic_pressure_max"), "پاسکال", PRESSURE, 0,
         "Max-Q — مبنای انتخاب جنس بدنه"),
        ("ارتفاع باز شدن چتر", ctx.metric("parachute_deploy_altitude"), "متر", COLOR_WARN, 1,
         "از رویداد «باز شدن چتر»"),
        ("نرخ افت دما با ارتفاع", ctx.metric("temperature_lapse_rate_c_per_km"), "°C/km", TEMPERATURE, 2,
         "مقدار استانداردِ جو ≈ " + fa_num(-6.5, 1)),
    ]
    top = r
    per_row, span = 4, 4
    r_cards_start = top
    for i, (label, value, unit, accent, dec, hint) in enumerate(cards):
        rr = r_cards_start + (i // per_row) * 3 + 1
        _kpi_card(ws, rr, C1 + (i % per_row) * span, span, label, value, unit, theme,
                  accent, decimals=dec, hint=hint)
    r = r_cards_start + math.ceil(len(cards) / per_row) * 3 + 2
    _panel(ws, top - 1, C1 - 1, r - 2, C2, theme, fill=theme["canvas2"],
           title="۱) شاخص‌های کلیدی پرواز", accent=theme["band"])
    r = _section(ws, r, C1, C2,
                 "اعداد از تحلیل خودکار فایل CSV آمده‌اند؛ خودِ سلول‌ها عدد واقعی‌اند "
                 "(قابل فرمول و Pivot) و فقط نمایششان فارسی است.", theme)

    # ---- ۲) گیج‌ها ----
    n_gauges = 0
    if ctx.ok:
        r, n_gauges = _gauge_strip(ws, ctx, theme, r, C1, C2)

    # ---- ۳) نمودار قهرمان + رویدادها ----
    if sws is not None:
        hdr, first, last = SER_HDR, SER_HDR + 1, SER_HDR + len(ctx.series)
        r = _section(ws, r, C1, C2, "ارتفاع، سرعت و لحظهٔ رویدادهای پرواز", theme,
                     sub="لوزی‌های رنگی روی منحنی ارتفاع = همان پنج رویدادِ جدول پایینی (محور افقی: ثانیه از شروع لاگ).")
        panel_top = r
        hero = LineChart()
        _series(hero, sws, cols["ارتفاع (متر)"], hdr, first, last, theme, ALTITUDE,
                width=2.6, area_fill=True, smooth=True)
        for key in EVENT_ORDER:
            _series(hero, sws, cols["نشانگر " + EVENT_NAMES[key]], hdr, first, last, theme,
                    EVENT_COLOR[key], no_line=True, marker="diamond", marker_size=11)
        vel_chart = LineChart()
        _series(vel_chart, sws, cols["سرعت عمودی (م/ث)"], hdr, first, last, theme, VELOCITY,
                width=1.7, smooth=True)
        _categories(hero, sws, cols["زمان (ثانیه)"], first, last)
        _categories(vel_chart, sws, cols["زمان (ثانیه)"], first, last)
        _chart_look(hero, theme, "ارتفاع (سطح زیر منحنی) و سرعت عمودی — نسبت به زمان",
                    38.6, 12.4, y_title="متر", legend=False,
                    y2_title="متر بر ثانیه", x_skip=12, y_fmt="[DBNum1]#,##0")
        _secondary_axis(hero, vel_chart, theme, "سرعت (م/ث)")
        _place_chart(ws, hero, f"{get_column_letter(C1)}{panel_top + 1}", fit_to=(C1, C2))
        # برچسب‌های رنگی رویدادها زیر نمودار (راهنمای دستی، به‌جای legend شلوغ)
        lab_r = panel_top + 26
        c = C1
        for key in EVENT_ORDER:
            cell = _merge(ws, lab_r, c, lab_r, c + 2, "◆ " + EVENT_NAMES[key])
            cell.font = _font(theme, 8.5, True, theme["panel"])
            cell.alignment = _align("center")
            for cc in range(c, c + 3):
                ws.cell(row=lab_r, column=cc).fill = _fill(_paint(theme, EVENT_COLOR[key]))
            c += 3
        ws.row_dimensions[lab_r].height = 18
        r = _panel(ws, panel_top, C1 - 1, lab_r + 1, C2, theme, fill=theme["panel"]) + 1

        # ---- جدول رویدادها ----
        r = _section(ws, r, C1, C2, "۲) خط زمانی رویدادهای پرواز", theme)
        rows = []
        for key in EVENT_ORDER:
            tv = ctx.events.get(key)
            i = ctx.idx.get(key)
            alt = None
            if i is not None and ctx.an.alt is not None and i < len(ctx.an.alt):
                alt = float(ctx.an.alt[i])
            vel = None
            if i is not None and ctx.an.vel is not None and i < len(ctx.an.vel):
                vel = float(abs(ctx.an.vel[i]))
            rows.append([EVENT_NAMES[key], ctx.time_of(key),
                         None if tv is None else float(tv), alt, vel,
                         "رویداد مرجع (شروع زمان)" if key == "launch" else EVENT_TIP[key]])
        # ستون آخر (معیار تشخیص) عمداً بیرون از Table نوشته می‌شود و روی
        # G..Q ادغام می‌شود: متن بلندِ داخل Table مجاز به ادغام نیست و
        # autosize هم شبکهٔ ۱۶ستونیِ یکنواختِ کارت‌ها را به‌هم می‌ریخت.
        ev_top = r
        r = _write_table(ws, top_row=ev_top, first_col=C1,
                         headers=["رویداد", "زمان نسبت به پرتاب (ثانیه)", "زمان در فایل (ثانیه)",
                                  "ارتفاع (متر)", "اندازهٔ سرعت (م/ث)"],
                         rows=[row[:5] for row in rows], theme=theme, table_name="TblEvents",
                         col_fmt={1: 2, 2: 2, 3: 1, 4: 1}, wrap_cols=(0,), autosize=False)
        ws.row_dimensions[ev_top].height = 48     # سربرگِ دوخطیِ فارسی (۱۱pt) در ۴۸ جا می‌گیرد
        tail_c1, tail_c2 = C1 + 5, C2
        hdr = _merge(ws, ev_top, tail_c1, ev_top, tail_c2, "معیار تشخیص")
        hdr.font = _font(theme, 9, True, theme["header_text"])
        hdr.alignment = _align("right", indent=1)
        for c in range(tail_c1, tail_c2 + 1):
            ws.cell(row=ev_top, column=c).fill = _fill(theme["header"])
            ws.cell(row=ev_top, column=c).border = Border(
                top=_side(theme["header"], "medium"), bottom=_side(theme["header"], "thin"),
                left=_side(theme["border"]) if c == tail_c1 else None,
                right=_side(theme["border"]) if c == tail_c2 else None)
        for i, row in enumerate(rows):
            rr = ev_top + 1 + i
            note = _merge(ws, rr, tail_c1, rr, tail_c2, str(row[5]))
            note.font = _font(theme, 8.5, False, theme["dim"])
            note.alignment = _align("right", indent=1)
            for c in range(tail_c1, tail_c2 + 1):
                cc = ws.cell(row=rr, column=c)
                if i % 2 == 1:
                    cc.fill = _fill(theme["zebra"])
                cc.border = Border(bottom=_side(theme["grid"]),
                                   left=_side(theme["border"]) if c == tail_c1 else None)
            ph = str(ws.cell(row=rr, column=2).value or "")
            for key, nm in EVENT_NAMES.items():
                if ph == nm:
                    cell = ws.cell(row=rr, column=2)
                    cell.font = _font(theme, 9, True, theme["panel"])
                    cell.fill = _fill(_paint(theme, EVENT_COLOR[key]))
                    break
        for c in range(C1, C2 + 1):
            for rr in range(ev_top, r):
                cell = ws.cell(row=rr, column=c)
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = _fill(theme["panel"])
        r += 1

    # ---- ۴) پیشنهادها ----
    r = _section(ws, r, C1, C2, "۳) ارزیابی خودکار و پیشنهادهای اصلاحی", theme,
                 sub="همین قواعد در گزارش PDF هم استفاده می‌شوند (core/advisor.py)؛ رنگ نقطه فقط برای خوانایی است. توضیحِ کاملِ هر عدد، کارنامهٔ پرواز و «قدم بعدی» در شیت «نتیجه‌گیری و آموزش» آمده است.")
    sug_top = r
    if ctx.suggestions:
        for text in ctx.suggestions:
            tone = _tone_for(text, theme)
            mark = ws.cell(row=r, column=C1, value="●")
            mark.font = _font(theme, 11, False, tone or theme["dim"])
            mark.alignment = _align("center")
            r = _note(ws, r, C1 + 1, C2, text, theme, height=30, tone=tone)
    else:
        r = _note(ws, r, C1, C2, "دادهٔ کافی برای ارزیابی خودکار موجود نیست.", theme, height=20)
    for rr in range(sug_top, r):
        for c in range(C1, C2 + 1):
            cell = ws.cell(row=rr, column=c)
            if cell.fill is None or cell.fill.fill_type is None:
                cell.fill = _fill(theme["panel"])
    r = _panel(ws, sug_top - 1, C1 - 1, r, C2, theme, fill=theme["canvas2"],
               title="خروجی کارشناس نرم‌افزار", accent=theme["band"]) + 1

    # ---- ۵) مشخصات مأموریت/موتور در دو کارت کنار هم ----
    r = _section(ws, r, C1, C2, "۴) مشخصات ثبت‌شدهٔ مأموریت، نازل و موتور", theme)
    spec = _spec_pairs(ctx)
    half = math.ceil(len(spec) / 2)
    left_c1, left_c2 = C1, C1 + 7
    right_c1, right_c2 = C1 + 8, C2
    start = r
    _merge(ws, start, left_c1, start, left_c2, "اطلاعات مأموریت").font = _font(theme, 10, True, theme["header_text"])
    _merge(ws, start, right_c1, start, right_c2, "موتور و نازل").font = _font(theme, 10, True, theme["header_text"])
    for c in range(left_c1, left_c2 + 1):
        ws.cell(row=start, column=c).fill = _fill(theme["header"])
    for c in range(right_c1, right_c2 + 1):
        ws.cell(row=start, column=c).fill = _fill(theme["header"])
    ws.row_dimensions[start].height = 20
    for i, (label, value) in enumerate(spec[:half]):
        _kv(ws, start + 1 + i, left_c1, left_c2, label, value, theme)
    for i, (label, value) in enumerate(spec[half:]):
        _kv(ws, start + 1 + i, right_c1, right_c2, label, value, theme)
    _panel(ws, start, left_c1, start + half, left_c2, theme)
    _panel(ws, start, right_c1, start + len(spec) - half, right_c2, theme)
    r = _panel(ws, start - 1, C1 - 1, start + max(half, len(spec) - half) + 1, C2, theme,
               fill=theme["canvas2"]) + 1
    ws.freeze_panes = "A6"
    # محدودهٔ چاپ = تا ستون آخرِ کارت (Q)؛ ستونِ خالیِ R فقط یک صفحهٔ اضافه می‌کرد
    ws.print_area = f"A1:{get_column_letter(C2)}{r}"
    return n_gauges + (1 if sws is not None else 0)


def _kv(ws, row: int, c1: int, c2: int, label: str, value: Any, theme: Dict[str, str]):
    """یک سطر «برچسب: مقدار» داخل کارت مشخصات -- فاصله‌گذاری با indent تا مقدار
    به برچسب نچسبد (نقطه‌ای که در نسخهٔ قبل شلوغ به‌نظر می‌رسید)."""
    lab = _merge(ws, row, c1, row, c1 + 2, label)
    lab.font = _font(theme, 9, False, theme["dim"])
    lab.alignment = _align("right", indent=1)
    val = _merge(ws, row, c1 + 3, row, c2, value)
    if (isinstance(value, (int, float)) and not isinstance(value, bool)
            and np.isfinite(float(value))):
        # عددِ واقعی می‌ماند (قابل Pivot/فرمول) و مثلثِ زردِ «number stored as text»
        # هم روی آن نمی‌آید؛ فقط قالبِ رقمِ فارسی می‌گیرد.
        val.number_format = numfmt(0 if float(value).is_integer() else 2)
    val.font = _font(theme, 9.5, True, theme["text"])
    val.alignment = _align("left", indent=1)
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(theme["zebra"] if row % 2 == 0 else theme["panel"])
        cell.border = Border(bottom=_side(theme["grid"]))
    ws.row_dimensions[row].height = 19


def _spec_pairs(ctx: _Context) -> List[Tuple[str, str]]:
    m, mo = ctx.mission, ctx.motor
    out = [
        ("شمارهٔ پروژه", fa(m.project_number or "--")),
        ("نام پروژه", m.project_name or "--"),
        ("نام راکت", m.rocket_name or "--"),
        ("شمارهٔ موتور", fa(m.motor_number or "--")),
        ("نسخهٔ Firmware", fa(m.firmware_version or "--")),
        ("زاویهٔ پرتاب", f"{fa_num(m.launch_angle, 1)} درجه"),
        ("وزن کل راکت", f"{fa_num(m.total_mass, 2)} کیلوگرم"),
        ("وزن موتور", f"{fa_num(m.motor_mass, 0)} گرم" if m.motor_mass else "--"),
        ("وزن سوخت", f"{fa_num(m.propellant_mass, 0)} گرم"),
        ("قطر بدنه", f"{fa_num((m.body_diameter or 0) * 1000, 0)} میلی‌متر"),
        ("مخروط سر", m.nose_cone or "اویو"),
        ("طول راکت", f"{fa_num((m.body_length or 0) * 100, 1)} سانتی‌متر"),
        ("ارتفاع محل پرتاب از سطح دریا", f"{fa_num(m.altitude_msl, 0)} متر"),
        ("نوع موتور", mo.motor_type or "--"),
        ("ضربهٔ کل موتور", f"{fa_num(mo.total_impulse, 0)} نیوتن‌ثانیه"),
        ("زمان سوخت", f"{fa_num(mo.burn_time, 2)} ثانیه"),
        ("رانش میانگین", f"{fa_num(mo.average_thrust, 1)} نیوتن"),
        ("قطر گلوگاه نازل", f"{fa_num(mo.throat_diameter, 1)} میلی‌متر"),
        ("قطر خروجی نازل", f"{fa_num(mo.exit_diameter, 1)} میلی‌متر"),
        ("زاویهٔ همگرا / واگرا", f"{fa_num(mo.convergent_angle, 0)}° / {fa_num(mo.divergent_angle, 0)}°"),
        ("طول نازل", f"{fa_num(mo.nozzle_length, 1)} سانتی‌متر"),
        ("فشار تخمینی محفظهٔ احتراق", f"{fa_num(mo.chamber_pressure_bar, 0)} بار"),
    ]
    try:
        from core.nozzle import optimal_expansion_ratio, classify_expansion
        if mo.throat_diameter and mo.exit_diameter:
            geo = (mo.exit_diameter / mo.throat_diameter) ** 2
            opt = optimal_expansion_ratio((mo.chamber_pressure_bar or 40) * 1e5, m.altitude_msl)
            state = {"under": "کم‌انبساط", "optimal": "بهینه", "over": "پرانبساط",
                     "unknown": "نامشخص"}.get(classify_expansion(geo, opt), "نامشخص")
            out += [("نسبت انبساط نازل (Ae/At)", f"{fa_num(geo, 2)}  —  {state}"),
                    ("نسبت انبساط بهینه در این ارتفاع", fa_num(opt, 2))]
    except Exception:
        pass
    if ctx.df is not None:
        out.append(("تعداد نمونه‌های ثبت‌شده", fa(f"{len(ctx.df):,}")))
        out.append(("نرخ نمونه‌برداری", fa_num(_sample_rate(ctx), 1) + " هرتز (تقریبی)"))
    return out


def _sample_rate(ctx: _Context) -> Optional[float]:
    try:
        t = np.asarray(ctx.an.t, dtype=float)
        d = np.diff(t)
        d = d[d > 0]
        return float(1.0 / np.median(d)) if len(d) else None
    except Exception:
        return None


# =========================================================================
# شیت ۲: نمودارهای زمانی همهٔ پارامترها
# =========================================================================
CHART_SPECS: List[Dict[str, Any]] = [
    dict(title="ارتفاع نسبت به زمان", cols=["ارتفاع (متر)"], colors=[ALTITUDE], y="متر", area=True,
         note="بیشینهٔ این نمودار همان «اوج پرواز» است."),
    dict(title="سرعت عمودی نسبت به زمان", cols=["سرعت عمودی (م/ث)"], colors=[VELOCITY], y="متر بر ثانیه",
         note="بالای صفر = در حال صعود، زیر صفر = در حال سقوط."),
    dict(title="شتاب کل نسبت به زمان", cols=["شتاب کل (g)"], colors=[ACCEL_TOTAL], y="g",
         note="۱g یعنی فقط گرانش؛ قله‌ها = رانش موتور و ضربهٔ باز شدن چتر."),
    dict(title="شتاب سه‌محور (X/Y/Z)", cols=["شتاب X (m/s²)", "شتاب Y (m/s²)", "شتاب Z (m/s²)"],
         colors=[AXIS_X, AXIS_Y, AXIS_Z], y="m/s²", note="قرمز=X، سبز=Y، آبی=Z (قرارداد استاندارد)."),
    dict(title="دمای سه ماژول و برآیندِ دمای هوا", cols=[T_HUM_MOD, T_BARO_MOD, T_GYRO_MOD, T_AIR],
         colors=[TEMPERATURE_AHT, TEMPERATURE, TEMPERATURE_MPU, ALTITUDE], y="°C",
         cols_any=True, temp_group=True,
         note="خطِ ماژول ژیروسکوپ فقط برایِ مقایسه است: آن عدد گرمای خودِ تراشه است، نه هوا؛ "
              "پس در «برآیندِ دمای هوا» و در چگالی هوا حساب نشده است."),
    dict(title="فشار هوا نسبت به زمان", cols=["فشار (هکتوپاسکال)"], colors=[PRESSURE], y="hPa", area=True,
         note="مبنای محاسبهٔ ارتفاع بارومتر در کامپیوتر پرواز."),
    dict(title="چگالی هوا نسبت به زمان", cols=["چگالی هوا (kg/m³)"], colors=[COLOR_INFO], y="kg/m³",
         note="از فشار و دمای همان لحظه: ρ = p / (R·T) — دقیق‌تر از عدد ثابت ۱٫۲۲۵."),
    dict(title="چرخش‌های راکت (ژیروسکوپ)",
         cols=["چرخش X (deg/s)", "چرخش Y (deg/s)", "چرخش Z (deg/s)"],
         colors=[AXIS_X, AXIS_Y, AXIS_Z], y="deg/s",
         note="نوسان زیاد حین رانش = ناپایداری هوایی (بالچه/مرکز ثقل را بررسی کنید)."),
    dict(title="مقایسهٔ ارتفاع بارومتر و GPS", cols=["ارتفاع (متر)", "ارتفاع GPS (متر)"],
         colors=[ALTITUDE, COLOR_INFO], y="متر",
         note="اختلاف دو خط = خطای ارتفاع بارومتریک (به‌دلیل دمای هوا و باد)."),
    dict(title="ولتاژ باتری نسبت به زمان", cols=["ولتاژ باتری (V)"], colors=[COLOR_OK], y="V",
         note="افت ولتاژ نزدیک انتهای پرواز = خطر خاموشی کامپیوتر پرواز."),
]


def _build_time_charts(wb, ctx: _Context, theme: Dict[str, str], sws, cols) -> int:
    ws = wb.create_sheet("نمودارهای زمانی")
    _init_sheet(ws, theme, tab_color=VELOCITY, bg_rows=8 * 17 + 4, bg_cols=17)
    _widths(ws, [1.7] + [14.8] * 16)
    hdr, first, last = SER_HDR, SER_HDR + 1, SER_HDR + len(ctx.series)

    def _keep(sp):
        have = [c for c in sp["cols"] if c in cols]
        return bool(have) if sp.get("cols_any") else len(have) == len(sp["cols"])

    # ستون‌هایِ دما نامِ ماژولِ انتخابی را دارند، پس نمودارِ دما همان‌جا بازساخته می‌شود
    specs = []
    for sp in CHART_SPECS:
        sp = dict(sp)
        if sp.get("temp_group"):
            sp["cols"] = [c for c in ctx.temp_keys if c in cols]
            sp["colors"] = [TEMPERATURE_AHT, TEMPERATURE, TEMPERATURE_MPU, ALTITUDE][:len(sp["cols"])]
        if _keep(sp):
            specs.append(sp)
    covered = {c for sp in specs for c in sp["cols"]}      # نمودارِ اختصاصی داریم؟
    for opt in ctx.optional:
        if opt["label"] in covered:
            continue          # وگرنه یک نمودارِ تکراریِ تک‌خطی کنارِ نمودارِ اصلی سبز می‌شود
        specs.append(dict(title=f"{opt['label'].split(' (')[0]} نسبت به زمان", cols=[opt["label"]],
                          colors=[opt["color"]], y=opt["unit"], area=False,
                          note=f"از ستون «{opt['column']}» فایل CSV خوانده شده است."))
    r = _section(ws, 2, 2, 17, "نمودار زمانیِ همهٔ داده‌های ثبت‌شده", theme,
                 sub="محور افقی همهٔ نمودارها «زمان از شروع لاگ (ثانیه)» است؛ دو نمودار در هر ردیف "
                     "و همه بومیِ اکسل‌اند، پس روی هر نقطه می‌توانید کلیک کنید.")
    r += 1
    built = 0
    CHART_ROWS = 23                     # ≈ ۱۳ سانتی‌متر (۱۶٫۵ پوینتی)
    for i, spec in enumerate(specs):
        col_pos = 2 + (i % 2) * 8
        if i and i % 2 == 0:
            r += CHART_ROWS + 3
        chart = AreaChart() if spec.get("area") else LineChart()
        if spec.get("area"):
            chart.grouping = "standard"
            chart.overlap = 100
        for cname, color in zip(spec["cols"], spec["colors"]):
            if cname not in cols:
                continue      # ستونِ اختیاریِ نبودن ⇒ فقط همان خط حذف می‌شود
            _series(chart, sws, cols[cname], hdr, first, last, theme, color,
                    legend_name=ctx.t_short.get(cname),
                    width=1.9 if len(spec["cols"]) > 1 else 2.6, area_fill=bool(spec.get("area")))
        _categories(chart, sws, cols["زمان (ثانیه)"], first, last)
        # ۱۵٫۸ سانتی‌متر = پهنایِ واقعیِ ۸ ستونِ ۱۰٫۸ی (کمتر از این، کارت
        # «خالی» به‌نظر می‌رسد؛ بیشتر، نمودار از کادرش بیرون می‌زند).
        _chart_look(chart, theme, fa(spec["title"]), 22.4, 11.6,
                    y_title=spec["y"], legend=len(spec["cols"]) > 1, x_skip=9)
        _place_chart(ws, chart, f"{get_column_letter(col_pos)}{r}", fit_to=(col_pos, col_pos + 7))
        # خط توضیح + کادر کارت: هر نمودار یک «کارت» مستقل است، نه جزیره‌ای بی‌حد
        # روی صفحه (الگوی مرجع 001)
        note = _merge(ws, r + CHART_ROWS, col_pos, r + CHART_ROWS, col_pos + 7,
                      "ℹ  " + spec["note"])
        note.font = _font(theme, 8.5, False, theme["dim"])
        note.alignment = _align("right", wrap=True, indent=1)
        ws.row_dimensions[r + CHART_ROWS].height = max(
            26.0, _wrap_height(ws, col_pos, col_pos + 7, "ℹ  " + spec["note"], 15.2, 26.0))
        _panel(ws, r - 1, col_pos - 1, r + CHART_ROWS, col_pos + 7, theme, fill=theme["panel"])
        built += 1
    r += CHART_ROWS + 4
    ws.print_area = f"A1:Q{r}"
    return built


# =========================================================================
# شیت ۳: پروفایل ارتفاع
# =========================================================================
def _altitude_bands(ctx: _Context, target_bands: int = 12) -> Optional[Tuple[pd.DataFrame, List[str]]]:
    """میانگین هر پارامتر در بازه‌های ارتفاعی -- برای این‌که اثر «ارتفاع» از اثر
    «زمان» جدا شود و کاربر مستقیم ببیند دما/رطوبت/فشار با بالا رفتن چه شد."""
    if not ctx.ok:
        return None
    an = ctx.an
    alt = np.asarray(an.alt, dtype=float) if an.alt is not None else None
    if alt is None or not np.isfinite(alt).any():
        return None
    top, ground = float(np.nanmax(alt)), float(np.nanmin(alt))
    span = max(top - ground, 1.0)
    step = max(50.0, round(span / target_bands / 50.0) * 50.0)
    edges = np.arange(ground, ground + step * (int(span / step) + 1), step)
    band = np.digitize(alt, edges) - 1
    valid = np.isfinite(alt)

    # تجمیع‌هر-بازه: دما/رطوبت/فشار میانگین گرفته می‌شوند، اما سرعت و شتاب
    # «بیشینه» -- چون راکت در یک بازهٔ ارتفاعی هم صعود می‌کند و هم فرود، و
    # میانگینِ سرعت عمودی تقریباً صفر می‌شود (علامت‌ها یکدیگر را خنثی می‌کنند).
    tp = ctx.temps or {}
    # ترتیبِ ستون‌ها = ترتیبِ درخواستیِ کاربر: دما و رطوبت، فشار و دما، ژیروسکوپ، برآیند
    sources: Dict[str, Tuple[Optional[np.ndarray], str]] = {
        ctx.t_hum: (tp.get(ctx.t_hum), "mean"),
        ctx.t_baro: (tp.get(ctx.t_baro), "mean"),
        ctx.t_gyro: (tp.get(ctx.t_gyro), "mean"),
        ctx.t_air: (tp.get(ctx.t_air), "mean"),
        "فشار (هکتوپاسکال)": (an.pressure, "mean"),
        "بیشینهٔ سرعت عمودی (م/ث)": (an.vel, "maxabs"),
        "بیشینهٔ شتاب (g)": ((an.a_total / G0) if an.a_total is not None else None, "max"),
    }
    sources["ولتاژ (V)"] = (ctx._col("Voltage"), "mean")
    for opt in ctx.optional:
        sources[opt["label"]] = (ctx._col(opt["column"]), "mean")
    # چگالی هوا از فشار و «برآیندِ دمای هوا» (نه دمای ماژول ژیروسکوپ)
    t_rho = tp.get(ctx.t_air)
    if t_rho is None:
        t_rho = tp.get(ctx.t_baro)
    if an.pressure is not None and t_rho is not None:
        with np.errstate(invalid="ignore", divide="ignore"):
            rho = (np.asarray(an.pressure, float) * 100.0) / (AIR_GAS_CONSTANT * (np.asarray(t_rho, float) + 273.15))
        sources["چگالی هوا (kg/m³)"] = (np.where(np.isfinite(rho) & (rho > 0.05) & (rho < 3.0),
                                                  rho, np.nan), "mean")

    def _has(v):
        return v is not None and np.isfinite(np.asarray(v, dtype=float)).sum() >= 2

    names = [k for k, (v, _agg) in sources.items()
             # سه ماژولِ دما حتی اگر داده نداشتند در لیست می‌آیند («--») تا خواننده بداند
             # کدام سنسور در این لاگ نبود؛ برآیند فقط وقتی که واقعاً میانگین گرفته شود.
             if _has(v) or k in (ctx.t_hum, ctx.t_baro, ctx.t_gyro)]
    if not names:
        return None
    rows = []
    for b in range(len(edges) - 1):
        mask = valid & (band == b)
        if mask.sum() < 2:
            continue
        row = {"lower": float(edges[b]), "upper": float(edges[b + 1]), "n": int(mask.sum())}
        for k in names:
            v, agg = sources[k]
            if v is None:
                row[k] = None
                continue
            v = np.asarray(v, dtype=float)[mask]
            v = v[np.isfinite(v)]
            if not len(v):
                row[k] = None
            elif agg == "maxabs":
                row[k] = float(np.max(np.abs(v)))
            elif agg == "max":
                row[k] = float(np.max(v))
            else:
                row[k] = float(np.mean(v))
        rows.append(row)
    if not rows:
        return None
    return pd.DataFrame(rows), names


def _profile_summary(ctx: _Context, bands: pd.DataFrame, names: List[str]) -> List[str]:
    out = []
    top = float(bands["upper"].iloc[-1])
    s = pd.Series(dtype=float)
    tkey = ctx.t_air if (ctx.t_air in names and bands[ctx.t_air].notna().sum() >= 2) else (
        ctx.t_baro if ctx.t_baro in names else None)
    if tkey:
        s = bands[tkey].dropna()
    if tkey and len(s) >= 2:
        out.append(f"برآیندِ دمای هوا ({ctx.temp_sources}) از {fa_num(s.iloc[0], 1)}°C نزدیک زمین به "
                   f"{fa_num(s.iloc[-1], 1)}°C در بازهٔ بالای {fa_num(top, 0)} متر رسیده است — یعنی "
                   f"{fa_num(abs(s.iloc[-1] - s.iloc[0]), 1)} درجه سردتر. دمای ماژول ژیروسکوپ در این "
                   "میانگین حساب نشده (آن عدد گرمای خودِ تراشه است)؛ پس نمودارِ خطِ نزولیِ دما باید "
                   "همین برآیند را نشان دهد، نه خطِ ژیروسکوپ را.")
    out.append("چطور فرقِ سه ماژول را بفهمیم؟ ماژول دما و رطوبت و ماژول فشار و دما بیرونِ محفظه "
               "هوا را می‌خوانند و با ارتفاع سردتر می‌شوند؛ ماژول ژیروسکوپ داخلِ جعبه و چسبیده به برد "
               "است، پس معمولاً گرم‌تر است و با سرد شدنِ هوا هم‌زمان گرم می‌شود. اگر خطِ آن هم "
               "نزولی بود، سنسور در معرضِ هوای بیرون است و می‌توان آن را هم در برآیند گذاشت.")

    for key, label in (("رطوبت (٪)", "رطوبت"), ("فشار (هکتوپاسکال)", "فشار هوا"),
                       ("چگالی هوا (kg/m³)", "چگالی هوا")):
        if key not in names:
            continue
        s = bands[key].dropna()
        if len(s) < 2 or not s.iloc[0]:
            continue
        pct = (s.iloc[-1] - s.iloc[0]) / abs(s.iloc[0]) * 100.0
        out.append(f"{label} با افزایش ارتفاع حدود {fa_num(abs(pct), 1)}٪ "
                   f"{'کاهش' if pct < 0 else 'افزایش'} یافته است ({fa_num(s.iloc[0], 1)} ← {fa_num(s.iloc[-1], 1)}).")
    lapse = ctx.metric("temperature_lapse_rate_c_per_km")
    if lapse is not None:
        out.append(f"میانگین نرخ افت دمای اندازه‌گیری‌شده: {fa_num(lapse, 2)} °C بر کیلومتر "
                   f"(مقدار مرجع جو استاندارد حدود {fa_num(-6.5, 1)} است؛ اگر عدد شما خیلی "
                   "کمتر/بیشتر باشد یعنی دمای سنسور تحت اثر تابش خورشید یا گرمای برد بوده است).")
    out.append("نکته: این جدول میانگینِ «هر بازهٔ ارتفاعی» است، نه هر بازهٔ زمانی؛ "
               "پس اگر راکت دو بار از یک ارتفاع عبور کند (صعود و فرود)، هر دو نمونه در یک "
               "میانگین حساب می‌شوند. برای دیدن ترتیب زمانی، شیت «نمودارهای زمانی» را ببینید.")
    return out


def _build_altitude_profile(wb, ctx: _Context, theme: Dict[str, str]) -> int:
    packed = _altitude_bands(ctx)
    if not packed:
        return False
    bands, names = packed
    headers = ["بازهٔ ارتفاع (متر)"] + names + ["اختلاف دمای هوا تا زمین"]
    if "رطوبت (٪)" in names:
        headers.append("اختلاف رطوبت تا زمین (٪)")
    headers.append("تعداد نمونه")
    # «برآیندِ دمای هوا» اگر نبود، اختلاف از همان ماژول فشار و دما حساب می‌شود
    tkey = ctx.t_air if ctx.t_air in names else (ctx.t_baro if ctx.t_baro in names else None)
    base_t = None
    if tkey and bands[tkey].notna().any():
        base_t = float(bands[tkey].dropna().iloc[0])
    base_h = bands["رطوبت (٪)"].dropna().iloc[0] if "رطوبت (٪)" in names else None
    rows = []
    for _, b in bands.iterrows():
        row = [f"{fa_num(b['lower'], 0)} تا {fa_num(b['upper'], 0)}"]
        for k in names:
            row.append(b.get(k))
        tv = b.get(tkey) if tkey else None
        row.append(None if (base_t is None or tv is None or not np.isfinite(float(tv)))
                   else float(tv) - float(base_t))
        if base_h is not None:
            hv = b.get("رطوبت (٪)")
            row.append(None if hv is None or not np.isfinite(float(hv)) else float(hv) - float(base_h))
        row.append(float(b["n"]))
        rows.append(row)

    LAST = 1 + len(headers)      # ستونِ آخرِ جدول: نوارِ نارنجی، کادر و نمودارها همه تا همین
    ws = wb.create_sheet("پروفایل ارتفاع")
    _init_sheet(ws, theme, tab_color=TEMPERATURE, bg_rows=6 + len(bands) + 24, bg_cols=LAST)
    # عرضِ ستون‌ها را _write_table/_autosize از *خودِ عددها* می‌سازد (سربرگ اجازه دارد
    # در سه خط بشکند)؛ به همین دلیل عدد‌ها به هم چسبیده‌اند و جایِ خالی نداریم.
    _widths(ws, [1.6] + [9.0] * (LAST - 1))
    r = _section(ws, 2, 2, LAST, "دما، رطوبت، فشار و چگالی هوا «نسبت به ارتفاع»", theme,
                 sub="برای هر بازهٔ ارتفاعی میانگین گرفته شده است (فقط سرعت عمودی و شتاب «بیشینه»، "
                     "چون راکت در هر بازه هم صعود می‌کند هم فرود)؛ ستون‌های «اختلاف تا زمین» مستقیم "
                     "می‌گویند با بالا رفتن راکت چقدر سردتر/کم‌رطوبت‌تر/کم‌فشارتر شده است. "
                     "«برآیندِ دمای هوا» میانگینِ ماژول‌های محیطی است"
                     + (f" (در این پرواز: {ctx.temp_sources})" if ctx.temp_sources != "--" else "")
                     + " و دمای ماژول ژیروسکوپ در آن حساب نمی‌شود، چون گرمای خودِ تراشه است. "
                       "واحدِ همهٔ ستون‌هایِ دما درجهٔ سلسیوس (°C) و واحدِ اختلافِ دما هم همین است "
                       "-- به همین دلیل این واحد در سربرگِ تک‌تکِ ستون‌ها تکرار نشده. "
                       "در نمودارهایِ پایینِ جدول، محورِ عمودی چند مقیاس را با هم نشان می‌دهد "
                       "(دما °C، رطوبت ٪، فشار hPa)؛ پس آن‌جا *شکلِ* تغییر را بخوانید، نه قدرتمطلقِ "
                       "خطوطِ مختلف را. برچسب‌هایِ محورِ افقیِ نمودارها هم *عددِ ابتدایِ هر بازه* است "
                       "(هر بازه ۵۰ متر است؛ جدولِ بالا ابتدا و انتهایِ کامل را نوشته).")
    r += 1

    dec = {}
    for j, h in enumerate(headers):
        if "تعداد" in h:
            dec[j] = 0
        elif "چگالی" in h:
            dec[j] = 3
        elif "بازهٔ" in h:
            dec[j] = 0
        else:
            dec[j] = 1
    table_top = r
    r = _write_table(ws, top_row=table_top, first_col=2, headers=headers, rows=rows,
                     theme=theme, table_name="TblAltBands", col_fmt=dec,
                     autosize=True, head_lines=3, min_w=7.2, max_w=16.0, pad=1.2, head_h=None)
    last_row = table_top + len(rows)
    # مقیاس رنگی: سبز ← زرد ← قرمز روی ستون‌های جوی و دeltas
    for j, h in enumerate(headers):
        if any(k in h for k in ("دمای", "رطوبت", "فشار", "چگالی", "اختلاف")):
            letter = get_column_letter(2 + j)
            ws.conditional_formatting.add(
                f"{letter}{table_top + 1}:{letter}{last_row}",
                ColorScaleRule(start_type="min", start_color="63BE7B",
                               mid_type="percentile", mid_value=50, mid_color="FFEB84",
                               end_type="max", end_color="F8696B"))

    # برچسبِ محورِ افقیِ نمودارها: ستونِ «بازهٔ ارتفاع (متر)» *متن* است («۱۰۰ تا ۱۵۰») و
    # در نمودار دو درد درست می‌کند: (۱) بیست‌وچند برچسبِ بلند در عرضِ ~۳۰cm جا نمی‌شوند و
    # روی هم می‌افتند، (۲) جهتِ دوسوییِ اکسل ترتیبشان را برعکس می‌کند («۱۵۰ تا ۱۰۰»).
    # پس یک ستونِ کمکیِ *پنهان* با عددِ ابتدایِ هر بازه نوشته می‌شود و نمودار همان را
    # می‌خواند؛ جدولِ بالا همچنان بازه‌های نوشتاریِ کامل را دارد (همهٔ نمودارها روی
    # «نمایش دادهٔ سلول‌هایِ پنهان» تنظیم‌اند، پس این ستونِ پنهان رسم نمودار را نمی‌شکند).
    bnd_col = LAST + 2
    hc = ws.cell(row=table_top, column=bnd_col, value="ابتدایِ بازه (متر)")
    hc.font = _font(theme, 9, True, theme["header_text"])
    for i, lo in enumerate(bands["lower"].tolist()):
        bcell = ws.cell(row=table_top + 1 + i, column=bnd_col, value=float(lo))
        bcell.number_format = numfmt(0)
        bcell.font = _font(theme, 9, False, theme["text"])
    ws.column_dimensions[get_column_letter(bnd_col)].width = 9
    ws.column_dimensions[get_column_letter(bnd_col)].hidden = True

    # نمودار ۱: پارامترهای جوی در برابر بازهٔ ارتفاع (برآیندِ دمای هوا + دو ماژولِ محیطی +
    # رطوبت). ماژول ژیروسکوپ در این نمودار *نیست*: دمای آن گرمای برد است و نمودارِ جوی را گمراه می‌کند.
    # +۰٫۸cm نسبت به قبل: اکسل جای راهنما را *از خودِ پلت* کم می‌کند؛ با کادرِ بلندتر
    # هم نمودار و هم راهنمایِ زیرش کامل دیده می‌شوند.
    CH_H = 13.2
    GAP_CM = 2.8
    prof = LineChart()
    # «دمای ماژول ژیروسکوپ» عمداً در این نمودارِ جوی نیست (شکایتِ کاربر: در برآیند حساب نشود)
    want = [c for c in (ctx.t_air, ctx.t_hum, ctx.t_baro, "رطوبت (٪)") if c in headers]
    color_of = {ctx.t_air: ALTITUDE, ctx.t_hum: TEMPERATURE_AHT, ctx.t_baro: TEMPERATURE,
                "رطوبت (٪)": HUMIDITY}
    for cname in want:
        _series(prof, ws, 2 + headers.index(cname), table_top, table_top + 1, last_row, theme,
                color_of.get(cname, COLOR_INFO), width=2.8, marker="circle", marker_size=7,
                smooth=True, legend_name=ctx.t_short.get(cname))
    if not want:
        cname = "فشار (هکتوپاسکال)"
        _series(prof, ws, 2 + headers.index(cname), table_top, table_top + 1, last_row, theme,
                PRESSURE, width=2.8, marker="circle", marker_size=7)
    _categories(prof, ws, bnd_col, table_top + 1, last_row)
    _chart_look(prof, theme, "برآیندِ دمای هوا و رطوبت با ارتفاع" if want else "تغییرِ فشار با ارتفاع",
                _cm_across(ws, 2, LAST), CH_H, y_title="",
                legend=bool(want), legend_pos="b",
                legend_items=[ctx.t_short.get(c, c.split(" (")[0]) for c in want])
    _place_chart(ws, prof, f"B{r + 1}", fit_to=(2, LAST))
    # بلوکِ دوم: *زیرِ* نمودارِ اول، با فاصله‌ای که اندازهٔ واقعیِ کادرِ اول را می‌بیند
    # (_chart_bottom) و یک نفسِ ۲٫۸ سانتی‌متری اضافه -- قبلاً فاصله ~۱cm بود و راهنمایِ
    # زیرِ نمودارِ بالا زیرِ نمودارِ پایین گم می‌شد.
    ch1_bot = _chart_bottom(ws, r + 1 + _rows_for_cm(CH_H))
    prof_r = ch1_bot + 1 + _rows_after(ws, ch1_bot + 1, GAP_CM)
    # نمودار ۲: فشار و چگالی هوا با ارتفاع
    others = [c for c in ("فشار (هکتوپاسکال)", "چگالی هوا (kg/m³)") if c in headers]
    if others:
        prof2 = LineChart()
        for cname in others:
            _series(prof2, ws, 2 + headers.index(cname), table_top, table_top + 1, last_row, theme,
                    PRESSURE if "فشار" in cname else COLOR_INFO, width=2.6,
                    marker="triangle", marker_size=6, smooth=True)
        _categories(prof2, ws, bnd_col, table_top + 1, last_row)
        _chart_look(prof2, theme, "فشار و چگالی هوا در هر بازهٔ ارتفاعی",
                    _cm_across(ws, 2, LAST), CH_H, y_title="",
                    legend=len(others) > 1, legend_pos="b",
                    legend_items=["فشار", "چگالی هوا"] if len(others) > 1 else [])
        _place_chart(ws, prof2, f"B{prof_r}", fit_to=(2, LAST))
    # هر نمودار کارتِ خودش را دارد؛ میانِ دو کارت همان فاصلهٔ ۲٫۸cm خالی می‌ماند
    card_bot = ch1_bot
    _panel(ws, r, 1, card_bot, LAST, theme, fill=theme["panel"])
    ch2_bot = _chart_bottom(ws, prof_r + _rows_for_cm(CH_H))
    _panel(ws, prof_r - 1, 1, ch2_bot, LAST, theme, fill=theme["panel"])
    rr = ch2_bot + 1
    rr = _section(ws, rr, 2, LAST, "این نمودار را چطور بخوانیم؟", theme)
    sum_top = rr
    for line in ([getattr(ctx, "temp_note", "")] + _profile_summary(ctx, bands, names)):
        if not line:
            continue
        rr = _note(ws, rr, 2, LAST, "•  " + line, theme, height=28)
    _panel(ws, sum_top - 1, 1, rr, LAST, theme, title="خواندنِ ساده", fill=theme["panel"])
    ws.freeze_panes = "A4"
    ws.print_area = f"A1:{get_column_letter(LAST)}{rr}"
    return 2 + (1 if others else 0)


# =========================================================================
# شیت ۴: تحلیل مرحله‌به‌مرحله
# =========================================================================
def _phase_stats(ctx: _Context) -> List[List[Any]]:
    if not ctx.ok:
        return []
    an = ctx.an
    n = len(an.t)
    labels = np.array([ctx.phase_of_index(int(i)) for i in range(n)], dtype=object)
    rows = []
    tp = ctx.temps or {}
    tsrc = tp.get(ctx.t_air)
    if tsrc is None:
        tsrc = tp.get(ctx.t_baro)          # فقط ماژولِ فشار و دما موجود است
    # دمای ماژول ژیروسکوپ عمداً در هیچ میانگینی نمی‌آید (گرمای خودِ برد)
    for phase in PHASE_ORDER:
        idxs = np.where(labels == phase)[0]
        if len(idxs) == 0:
            continue
        i0, i1 = int(idxs[0]), int(idxs[-1])
        sl = slice(i0, i1 + 1)
        v = np.abs(np.asarray(an.vel[sl], dtype=float)) if an.vel is not None else np.array([])
        a = (np.asarray(an.a_total[sl], dtype=float) / G0) if an.a_total is not None else np.array([])
        alt = np.asarray(an.alt[sl], dtype=float) if an.alt is not None else np.array([])
        temp = np.asarray(tsrc[sl], dtype=float) if tsrc is not None else np.array([])
        net_alt = float(alt[-1] - alt[0]) if len(alt) else None
        fin = lambda x: float(np.nanmax(x)) if len(x) and np.isfinite(x).any() else None
        mean = lambda x: float(np.nanmean(x)) if len(x) and np.isfinite(x).any() else None
        t0, t1 = float(an.t[i0]), float(an.t[i1])
        rows.append([phase, max(t1 - t0, 0.0), t0, t1, fin(v), mean(v), fin(a),
                     float(alt[0]) if len(alt) else None,
                     float(alt[-1]) if len(alt) else None,
                     net_alt, mean(temp), PHASE_TIP.get(phase, "")])
    return rows


def _build_phases(wb, ctx: _Context, theme: Dict[str, str]) -> int:
    rows = _phase_stats(ctx)
    if not rows:
        return False
    ws = wb.create_sheet("تحلیل مراحل پرواز")
    # bg_rows سخاوتمندانه: بلوک‌هایِ نمودار حالا قدِ خودِ نمودار را دارند و شیت
    # بلندتر شده؛ اگر زمینهٔ تم تیره کم پوشش باشد، پایینِ شیت سفید می‌ماند.
    _init_sheet(ws, theme, tab_color=ACCEL_TOTAL, bg_rows=40 + len(rows) * 2 + 90, bg_cols=14)
    # عرضِ ستون‌ها را _autosize از عددها می‌سازد (سربرگ در سه خط می‌شکند)؛ فقط
    # «توضیح مرحله» پهن می‌ماند. نمودارها به پهنایِ ستون‌هایِ خودشان تنظیم می‌شوند.
    _widths(ws, [1.6] + [9.0] * 13)
    r = _section(ws, 2, 2, 14, "سرعت، شتاب و ارتفاع «در هر مرحلهٔ پرواز»", theme,
                 sub="پرواز از روی همان رویدادها به بازه‌ها تقسیم شده است؛ ستون‌هایی که نوار داده "
                     "دارند، مقایسهٔ چشمیِ «کدام مرحله سریع‌تر بود» را می‌سازند. واحدِ ستونِ "
                     "«میانگین دمای هوا» درجهٔ سلسیوس (°C) است و میانگینِ ماژول‌هایِ محیطی است "
                     "(دمای ماژول ژیروسکوپ در آن حساب نمی‌شود).")
    r += 1
    headers = ["مرحله", "مدت (ثانیه)", "شروع (ثانیه)", "پایان (ثانیه)", "بیشینهٔ سرعت (م/ث)",
               "میانگین سرعت (م/ث)", "بیشینهٔ شتاب (g)", "ارتفاع ابتدا (متر)", "ارتفاع انتها (متر)",
               "تغییر ارتفاع در مرحله (متر)", "میانگین دمای هوا", "توضیح مرحله",
               "سهم از زمان کل (٪)"]
    # درصدِ هر مرحله به جدول اضافه می‌شود؛ این‌طور عددِ «سخت‌خواندنی» روی برشِ
    # نمودار لازم نیست -- در جدول دقیق، و در راهنمای نمودار خلاصه هست.
    _total_dur = sum(float(rw[1] or 0.0) for rw in rows) or 1.0
    for rw in rows:
        rw.append(round(float(rw[1] or 0.0) / _total_dur * 100.0, 1))
    table_top = r
    r = _write_table(ws, top_row=table_top, first_col=2, headers=headers, rows=rows,
                     theme=theme, table_name="TblPhases",
                     col_fmt={1: 2, 2: 2, 3: 2, 4: 1, 5: 1, 6: 2, 7: 1, 8: 1, 9: 1, 10: 1,
                              12: 1},
                     wrap_cols=(11,), autosize=True, head_lines=3, min_w=7.4, max_w=16.0,
                     pad=1.2, head_h=None)
    last_row = table_top + len(rows)
    for j, color in ((4, VELOCITY), (1, ACCEL_TOTAL), (5, ALTITUDE)):
        letter = get_column_letter(2 + j)
        ws.conditional_formatting.add(
            f"{letter}{table_top + 1}:{letter}{last_row}",
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color=_hex(color), showValue=True))
    for rr in range(table_top + 1, last_row + 1):
        ph = str(ws.cell(row=rr, column=2).value)
        accent = PHASE_COLORS.get(ph)
        if accent:
            cell = ws.cell(row=rr, column=2)
            cell.fill = _fill(accent)
            cell.font = _font(theme, 9.5, True, "FFFFFF")
        ws.cell(row=rr, column=13).alignment = _align("right", wrap=True)
        # قدِ سطر = متنِ همان سطر (سه/چهار خطِ «توضیح مرحله» با ۳۴ پوینت بریده می‌شد؛
        # اکسل سطرِ بلند را بی‌صدا نمی‌کند، پس متنِ نصفه از همان‌جا «گم» به نظر می‌رسید)
        ws.row_dimensions[rr].height = max(
            34.0, _wrap_height(ws, 13, 13, str(ws.cell(row=rr, column=13).value or ""), 15.2, 34.0))

    charts = 0
    # نامِ برش‌ها مستقیم از همان ستون «مرحله»ٔ جدول خوانده می‌شود. نسخه‌هایِ قبل
    # درصد را داخلِ نامِ برچسب می‌گذاشتند («مرحله — ۳۲٪») از یک ستونِ پنهان؛
    # برچسب‌ها آن‌قدر بلند بودند که در راهنما می‌شکستند و روی خودِ نمودار
    # می‌ریختند (همان «نوشته‌ها داخل نمودارها هستند»). درصد حالا ستونِ عددیِ
    # «سهم از زمان کل (٪)» در همان جدول است و نمودار فقط نام‌ها را نشان می‌دهد.
    PH_H = 12.6      # بدونِ تیتر/راهنمایِ داخلِ کادر، قدِ بیشتر = حلقه و ستون‌هایِ درشت‌تر
    ph_rows = _rows_for_cm(PH_H)

    bar = BarChart()
    bar.type = "col"
    _series(bar, ws, 2 + 4, table_top, table_top + 1, last_row, theme, VELOCITY, width=1.0)
    _categories(bar, ws, 2, table_top + 1, last_row)
    # عرضِ هر نمودار = پهنایِ *دقیق* ستون‌هایش (fit_to)؛ قبلاً نمودارِ ستونی
    # ۲۲٫۵cm روی ۶ ستونِ ~۱۷cm‌ای گذاشته می‌شد و نیمه‌اش زیرِ دوناتِ کناری
    # می‌رفت -- دو نمودارِ روی‌هم، یعنی «نوشتهٔ داخل نمودار».
    _chart_look(bar, theme, "بیشینهٔ سرعت در هر مرحله", _cm_across(ws, 2, 11), PH_H,
                y_title="متر بر ثانیه", legend=False, y_fmt="[DBNum1]#,##0.0")
    bar.gapWidth = 55
    _place_chart(ws, bar, f"B{r + 1}", fit_to=(2, 11))
    charts += 1
    _note(ws, r + 1, 12, 14, "ارتفاعِ هر ستون = بیشینهٔ سرعتِ عمودیِ همان مرحله؛ "
                             "مقیاسِ محورِ عمودی برایِ همهٔ مراحل یکی است.", theme)
    bar_bot = _chart_bottom(ws, r + 1 + ph_rows)
    _panel(ws, r, 1, bar_bot, 14, theme, fill=theme["panel"])
    r = bar_bot + 3

    donut = DoughnutChart()
    _series(donut, ws, 2 + 1, table_top, table_top + 1, last_row, theme, ACCEL_TOTAL, width=1.0)
    _categories(donut, ws, 2, table_top + 1, last_row)
    _color_slices(donut.series[0], [PHASE_COLORS.get(row[0], COLOR_INFO) for row in rows], theme)
    # هیچِ تیتر و هیچِ راهنمایی *داخل* کادر نمی‌گذاریم: اکسل تیتر را روی بالای حلقه
    # می‌انداخت و راهنما هم برشِ کوچک را می‌پوشاند («نوشته‌ها همچنان روی نمودار است»).
    # عنوان به سلولِ کنار نمودار و درصدِ هر مرحله به سطرهایِ *زیرِ* نمودار رفته است.
    _chart_look(donut, theme, None, _cm_across(ws, 2, 9), PH_H, legend=False)
    _place_chart(ws, donut, f"B{r + 1}", fit_to=(2, 9))
    charts += 1
    t = _note(ws, r + 1, 10, 14, "سهم زمانی هر مرحله (٪)", theme, size=12.5, bold=True)
    _note(ws, t, 10, 14, "اندازهٔ هر برش = مدتِ زمانِ همان مرحله؛ رنگش هم رنگِ همان سطرِ "
                         "جدولِ بالاست. درصدِ دقیق را در خط‌هایِ درشتِ زیرِ نمودار ببینید.",
          theme, size=10.5)
    donut_bot = _chart_bottom(ws, r + 1 + ph_rows)
    # درصدِ هر مرحله -- یک سطرِ درشتِ خوانا به‌جای نوشتهٔ ریزِ داخلِ حلقه
    rr = donut_bot + 1
    for row in rows:
        ph = str(row[0])
        dur = float(row[1] or 0.0)
        pct = float(row[-1])
        color = PHASE_COLORS.get(ph)
        rr = _note(ws, rr, 2, 14,
                   f"■  {ph} —— {fa_num(pct, 1)}٪ از زمانِ کل پرواز  "
                   f"({fa_num(dur, 1)} ثانیه؛ از {fa_num(float(row[2] or 0.0), 1)} تا "
                   f"{fa_num(float(row[3] or 0.0), 1)} ثانیه)",
                   theme, height=23, size=11.5, bold=True,
                   tone=(_hex(_paint(theme, color)) if color else None))
    _panel(ws, r, 1, rr - 1, 14, theme, fill=theme["panel"])
    r = rr + 1

    # تجزیهٔ «چطور به اوج رسید»
    alt, idx = ctx.an.alt, ctx.idx
    if alt is not None and idx.get("apogee") is not None:
        apogee_i, start_i = int(idx["apogee"]), int(idx.get("launch", 0))
        burn_i = int(idx.get("burnout", start_i))
        gained = float(alt[apogee_i] - alt[start_i])
        burn_alt = float(alt[burn_i] - alt[start_i])
        coast_alt = gained - burn_alt
        total = max(abs(gained), 1.0)
        r = _section(ws, r, 2, 14, "راکت چطور به اوج رسید؟ (تجزیهٔ مسیر صعود)", theme,
                     sub="هر چقدر سهم «پس از پایان رانش» بیشتر باشد، راکت سبک‌تر/"
                         "آیرودینامیک‌تر است.")
        # نام‌ها کوتاه‌اند تا راهنمایِ زیرِ نمودار یک‌سطری جا شود و رویِ برش‌ها نیفتد
        seg = [["صعود با رانشِ موتور", burn_alt, burn_alt / total * 100.0],
               ["صعود با اینرسی (پس از خاموشی)", coast_alt, coast_alt / total * 100.0]]
        r = _write_table(ws, top_row=r, first_col=2,
                         headers=["مؤلفهٔ رسیدن به اوج", "ارتفاع (متر)", "سهم از صعود (٪)"],
                         rows=seg, theme=theme, table_name="TblAscentSplit",
                         col_fmt={1: 0, 2: 1},   # ارتفاعِ صعود را با رقمِ اعشار نمی‌خواهیم؛
                         # ستونِ C عرضش برایِ همان «۱٬۲۰۱» کوک شده و رقمِ صفرم از ### نجات می‌دهد
                         wrap_cols=(0,), autosize=False)
        asc_top = r - 2                   # اولِ کارت (تیتر بخش + یکی از سطرهای توضیح)
        pie_top = r - len(seg) - 1        # نمودار از سطرِ سربرگِ جدول شروع می‌شود
        pie_h = 11.0
        pie = DoughnutChart()
        _series(pie, ws, 3, pie_top, pie_top + 1, r - 1, theme, ALTITUDE, width=1.0)
        _categories(pie, ws, 2, pie_top + 1, r - 1)
        _color_slices(pie.series[0], [COLOR_WARN, ALTITUDE], theme)
        # بیرون از کادر: تیترِ داخلِ حلقه، برش‌ها را می‌پوشاند
        _chart_look(pie, theme, None, _cm_across(ws, 6, 14), pie_h, legend=False)
        _place_chart(ws, pie, f"F{pie_top}", fit_to=(6, 14))
        charts += 1
        pie_bot = _chart_bottom(ws, pie_top + _rows_for_cm(pie_h))
        pr = pie_bot + 1
        for label, meters, pct in seg:
            pr = _note(ws, pr, 2, 5,
                       f"■  {label} —— {fa_num(pct, 1)}٪ از صعودِ کل  "
                       f"({fa_num(meters, 0)} متر)", theme, height=22, size=11, bold=True,
                       tone=_hex(_paint(theme, COLOR_WARN if "موتور" in label else ALTITUDE)))
        # کارت: جدول کوچک + دونات. قد کارت از ارتفاعِ خودِ دونات حساب می‌شود تا
        # نمودار از کادر بیرون نزند و سطرِ بعدی زیرِ آن نیفتد.
        card_bot = max(pie_bot, pr - 1, r + 1)
        _panel(ws, asc_top, 1, card_bot, 14, theme, fill=theme["panel"])
        r = card_bot + 1
    ws.freeze_panes = "A4"
    return charts


# =========================================================================
# شیت‌های داده
# =========================================================================
def _numeric_columns(frame: pd.DataFrame) -> Tuple[List[list], List[bool]]:
    """تبدیل یک DataFrame به ستون‌های آمادهٔ نوشتن (سریع و برداری).

    دلیل این تابع: نوشتن با df.iloc[i][col] برای ده‌ها هزار سلول، چند ثانیه
    وقت می‌گرفت (و چون تولید گزارش در ترد اصلی رابط کاربری انجام می‌شود،
    همان چند ثانیه یعنی انجماد پنجره). اینجا هر ستون یک‌بار به numpy رفته و
    مقادیر نامعتبر به None (سلول خالی) تبدیل می‌شوند.
    """
    values, numeric_flags = [], []
    for name in frame.columns:
        s = frame[name]
        if s.dtype == object:
            values.append([None if (v is None or (isinstance(v, float) and not np.isfinite(v))) else str(v)
                           for v in s.tolist()])
            numeric_flags.append(False)
            continue
        arr = pd.to_numeric(s, errors="coerce").to_numpy(dtype=float)
        valid = np.isfinite(arr)
        out = arr.astype(object)
        out[~valid] = None
        values.append(out.tolist())
        numeric_flags.append(True)
    return values, numeric_flags


def _build_series_sheet(wb, ctx: _Context, theme: Dict[str, str]):
    """شیتِ منبعِ دادهٔ نمودارها (نامِ قبلی: «سری‌های نمودار»).

    پنهان نیست تا کاربر بتواند همان اعدادی را که در نمودار رسم شده ببیند، ولی دو
    سطرِ اول توضیحِ کارکردش را می‌گوید. سربرگ در نسخه‌هایِ قبل متنِ هم‌رنگِ
    زمینه‌اش داشت (سرمه‌ای روی سرمه‌ای) و خوانده نمی‌شد -- کنتراست اصلاح شد.
    """
    ws = wb.create_sheet(SERIES_SHEET)
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _hex(COLOR_MISSING)
    ws.sheet_view.zoomScale = 80
    hdr = list(ctx.series.columns)
    _section(ws, 1, 1, max(2, min(len(hdr), 8)),
             "این شیت منبعِ دادهٔ همهٔ نمودارهای همین گزارش است", theme,
             sub="نمودارها این اعداد را می‌خوانند (نه شیت «دادهٔ خام پرواز» را)؛ اگر سطر/ستونی را "
                 "اینجا تغییر دهید یا این شیت را حذف کنید، نمودارها هم عوض یا خراب می‌شوند. "
                 f"ستون‌های دما به نامِ ماژول‌های انتخابی‌اند: {ctx.temp_mods['hum']}، "
                 f"{ctx.temp_mods['baro']} و {ctx.temp_mods['gyro']} (ماژولِ آخر فقط برایِ مقایسه است "
                 "و در برآیندِ دمای هوا حساب نمی‌شود). واحدِ همهٔ ستون‌هایِ دما درجهٔ سلسیوس (°C) است "
                 "و در سربرگِ ستون‌ها تکرار نشده تا شلوغ نشود.")
    head_font, body_font = _font(theme, 9, True, theme["header_text"]), _font(theme, 9, False, theme["text"])
    center = _align("center", wrap=True)
    for j, h in enumerate(hdr, start=1):
        c = ws.cell(row=SER_HDR, column=j, value=str(h))
        c.font = head_font
        c.fill = _fill(theme["header"])
        c.alignment = center
        c.border = Border(bottom=_side(theme["header"], "medium"))
        ws.column_dimensions[get_column_letter(j)].width = 18.5
    ws.row_dimensions[SER_HDR].height = 40

    values, numeric_flags = _numeric_columns(ctx.series)
    fmts = [numfmt(3 if "چگالی" in h else 2) for h in hdr]
    for i in range(len(ctx.series)):
        for j, col_vals in enumerate(values, start=1):
            v = col_vals[i]
            if v is None:
                continue
            cell = ws.cell(row=i + SER_HDR + 1, column=j, value=v)
            cell.font = body_font
            cell.alignment = _align("center")      # اعدادِ این شیت هم وسط‌چین
            if numeric_flags[j - 1]:
                cell.number_format = fmts[j - 1]
    tip = _merge(ws, len(ctx.series) + SER_HDR + 2, 1, len(ctx.series) + SER_HDR + 2,
                 min(len(hdr), 6),
                 "این شیت منبع دادهٔ نمودارهای گزارش است (نمونه‌برداری‌شده تا "
                 + fa(MAX_CHART_POINTS) + " نقطه تا فایل روان بماند). اگر مقداری را اینجا "
                 "تغییر دهید، همان تغییر در نمودارها دیده می‌شود.")
    tip.font = _font(theme, 9, False, theme["dim"])
    tip.alignment = _align("right")
    ws.freeze_panes = f"A{SER_HDR + 1}"
    return ws


def _build_raw_sheet(wb, ctx: _Context, theme: Dict[str, str]) -> int:
    """دادهٔ خام کامل. برای خوانایی و چاپ، این شیت روی زمینهٔ سفیدِ اکسل می‌ماند
    (متن تیره) ولی سربرگ‌ها و ارقام فارسی‌اند."""
    ws = wb.create_sheet("دادهٔ خام پرواز")
    t = _theme(False)
    t["table_style"] = theme["table_style"]
    _init_sheet(ws, t, gridlines=True, tab_color=COLOR_MISSING, zoom=85, bg_rows=0)
    df = ctx.df
    cols = list(df.columns)
    extra: Dict[str, Any] = {}
    if ctx.ok:
        an = ctx.an
        if an.vel is not None and len(an.vel) == len(df):
            extra["سرعت عمودی (م/ث)"] = an.vel
        if an.a_total is not None and len(an.a_total) == len(df):
            extra["شتاب کل (g)"] = an.a_total / G0
        extra["مرحلهٔ پرواز"] = [ctx.phase_of_index(int(i)) for i in range(len(df))]
    src: Dict[str, Any] = {}
    raw_map = getattr(ctx, "raw_headers", PERSIAN_HEADERS)
    for cname in cols:
        src[raw_map.get(cname, cname)] = df[cname].to_numpy()
    src.update(extra)
    out_df = pd.DataFrame(src)
    headers = ["ردیف"] + list(out_df.columns)
    r = _band_title(ws, 2, 1, len(headers), "دادهٔ خام کامل پرواز (بدون نمونه‌برداری)", t, COLOR_MISSING,
                    "همان چیزی که کامپیوتر پرواز روی کارت SD ثبت کرده است + سه ستون محاسبه‌شدهٔ "
                    "نرم‌افزار (سرعت عمودی، شتاب کل، مرحلهٔ پرواز). فیلتر سرستون‌ها روشن است.")
    r += 1
    hdr_row = r
    head_font, body_font = _font(t, 8.5, True, t["header_text"]), _font(t, 8.5)
    body_fmt = numfmt(2)          # قالبِ ارقامِ فارسیِ ستون‌هایِ عددی
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = head_font
        c.fill = _fill(t["header"])
        c.alignment = _align("center", wrap=True)
        c.border = _box(t)
        letter = get_column_letter(j)
        # ۱۴٫۵ = جا برای نمایشِ فارسیِ شش‌رقمیِ ماهواره (۵۳٫۳۹۶۹۹ در ۱۳ بریده می‌شد)
        ws.column_dimensions[letter].width = 14.5 if j > 1 else 6
        # فونت و قالبِ عددی هر دو *یک‌بار روی ستون* می‌نشینند. دلیلش یک باگِ واقعی
        # بود: `cell.number_format = ...` برای هر ۱۱۷هزار سلول، یک استایلِ سلولی با
        # فونتِ پیش‌فرض (Calibri) می‌ساخت و چون استایلِ سلول بر استایلِ ستون مقدم
        # است، کل شیتِ «دادهٔ خام» از فونتِ Shabnam بیرون می‌زد. اینجا سلول‌ها هیچ
        # استایلی ندارند تا استایلِ ستون به آن‌ها برسد (فایل هم کوچک‌تر می‌شود).
        ws.column_dimensions[letter].font = body_font
        ws.column_dimensions[letter].number_format = numfmt(0) if j == 1 else body_fmt
        # وسط‌چین هم روی *ستون*: سلول‌هایِ عددی استایلِ خودی ندارند، پس همین
        # استایل را می‌گیرند (بدونِ ۱۱۷هبار انتسابِ جداگانه)
        ws.column_dimensions[letter].alignment = _align("center")
    ws.row_dimensions[r].height = 46     # سربرگِ بلند (۳ خطِ ۱۱pt) در ۳۴ جا نمی‌گنجید

    values, numeric_flags = _numeric_columns(out_df)
    for i in range(len(out_df)):
        rr = hdr_row + 1 + i
        ws.cell(row=rr, column=1, value=i + 1)
        for j, col_vals in enumerate(values, start=2):
            v = col_vals[i]
            if v is None:
                continue
            cell = ws.cell(row=rr, column=j, value=v)
            if not numeric_flags[j - 2]:
                # ستونِ متنی (مثل «مرحلهٔ پرواز») وسط‌چین می‌شود؛ فونت هم صریح
                # داده می‌شود تا همین استایلِ دست‌ساز، استایلِ ستون را بی‌اثر نکند.
                cell.font = body_font
                cell.alignment = _align("center")
    last = hdr_row + len(out_df)
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(headers))}{last}"
    ws.print_title_rows = f"{hdr_row}:{hdr_row}"
    return len(df)


# =========================================================================
# شیتِ «نتیجه‌گیری و آموزش» -- تحلیلِ فارسیِ داده‌ها برایِ کاربری که
# مهندسِ هوافضا نیست (دانش‌آموز/دانشجو). هیچ نموداری در این شیت نیست؛
# فقط متن، جدول و قضاوت -- پس ریسکِ برخوردِ چیدمان هم ندارد.
# =========================================================================
def _grid_table(ws, row: int, *, first_col: int, cols: Sequence[Tuple[str, int]],
                rows: Sequence[Sequence[Any]], theme: Dict[str, str],
                tones: Optional[Sequence[Optional[str]]] = None, tone_col: int = -1,
                num_cols: Sequence[int] = (), decimals: Optional[Dict[int, int]] = None,
                head_h: float = 34.0, size: float = 9.5, min_h: float = 26.0) -> int:
    """جدولی که هر سلولش چند ستونِ ادغام‌شده است و متنش wrap می‌شود.

    چرا `_write_table` نه: آن تابع `Table` می‌سازد و «merge داخل محدودهٔ Table»
    در اکسل خطاست (پیام Removed Records: Merge cells / Repair). اینجا جدولِ
    دست‌سازِ استایل‌دار داریم و -- مهم‌تر از همه -- ارتفاعِ هر سطر از *بلندترین*
    متنِ همان سطر حساب می‌شود، تا توضیحِ فارسیِ مفصل بی‌صدا بریده نشود.
    `tones`: رنگِ قضاوت (good/warn/bad) فقط روی ستونِ `tone_col` می‌نشیند.
    """
    decimals = decimals or {}
    edges: List[Tuple[int, int]] = []
    c = first_col
    for _label, span in cols:
        edges.append((c, c + span - 1))
        c += span
    # ---- سربرگ ----
    for j, (label, (c1, c2)) in enumerate(zip([x[0] for x in cols], edges)):
        cell = _merge(ws, row, c1, row, c2, label)
        cell.font = _font(theme, size, True, theme["header_text"])
        cell.alignment = _align("center", wrap=True)
        for cc in range(c1, c2 + 1):
            ws.cell(row=row, column=cc).fill = _fill(theme["header"])
            ws.cell(row=row, column=cc).border = Border(
                top=_side(theme["header"], "medium"), bottom=_side(theme["header"], "thin"),
                left=_side(theme["border"]) if cc == c1 else None,
                right=_side(theme["border"]) if cc == c2 else None)
    # فونتِ سربرگ ۱۱pt است (کفِ سراسریِ _font)؛ head_hهایِ دست‌نویسِ ۲۸..۴۰ برایِ
    # دو خطِ ۹٫۵pt کوک شده بودند، پس کفِ ۳۴ اینجا امن‌تر است (بیشترش دست‌نخورده).
    ws.row_dimensions[row].height = max(head_h, 34.0)
    row += 1
    # ---- بدنه ----
    for i, vals in enumerate(rows):
        height = min_h
        for j, v in enumerate(vals):
            c1, c2 = edges[j] if j < len(edges) else edges[-1]
            is_num = j in num_cols and isinstance(v, (int, float)) \
                and not isinstance(v, bool) and np.isfinite(float(v))
            text = "" if v is None else ("" if is_num else str(v))
            cell = _merge(ws, row, c1, row, c2, None if is_num else (text or "--"))
            if is_num:
                cell.value = round(float(v), decimals.get(j, 1) + 3)
                cell.number_format = numfmt(decimals.get(j, 1))
            tone = (tones[i] if tones is not None and i < len(tones) else None) \
                if j == tone_col else None
            if tone:
                color = theme.get(tone, theme["text"])
                if theme["dark"]:
                    # در تمِ تیره «تِینت» یعنی روشن کردنِ رنگ؛ فیلِ روشن پشتِ
                    # متنِ روشن ناخوانا می‌شود، پس فیل تیره می‌ماند و فقط فونت رنگی است.
                    cell.font = _font(theme, size, True, color)
                else:
                    cell.font = _font(theme, size, True, color)
                    for cc in range(c1, c2 + 1):
                        ws.cell(row=row, column=cc).fill = _fill(_tint(color, 0.84))
            else:
                cell.font = _font(theme, size, j == 0, theme["dim"] if j == len(vals) - 1 else theme["text"])
                for cc in range(c1, c2 + 1):
                    ws.cell(row=row, column=cc).fill = _fill(theme["zebra"] if i % 2 else theme["panel"])
            cell.alignment = (_align("center", vert="center", wrap=True) if is_num
                              else _align("right", vert="top", wrap=True, indent=1))
            for cc in range(c1, c2 + 1):
                ws.cell(row=row, column=cc).border = Border(
                    bottom=_side(theme["grid"]), left=_side(theme["border"]) if cc == c1 else None,
                    right=_side(theme["border"]) if cc == c2 else None)
            if not is_num and text:
                height = max(height, _wrap_height(ws, c1, c2, text, 15.2, min_h))
        ws.row_dimensions[row].height = height
        row += 1
    return row + 1


def _fnum(value: Any, decimals: int = 1, suffix: str = "") -> str:
    """عدد -> رشتهٔ فارسیِ امن ('--' اگر نبود)؛ برایِ داخلِ متن‌هایِ توضیحی."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "--"
    if not np.isfinite(v):
        return "--"
    return fa_num(v, decimals) + (suffix or "")


def _conc_facts(ctx: "_Context") -> Dict[str, Any]:
    """هرچه برایِ نتیجه‌گیری لازم است، یک‌جا و *بی‌خطر* محاسبه می‌شود.

    همه‌چیز Optional است: اگر ستونی در CSV نباشد، مقدار None می‌ماند و متن‌ها
    '--' نشان می‌دهند (نه کرش، نه قضاوتِ ساختگی).
    """
    f: Dict[str, Any] = {}
    an = ctx.an

    def num(key: str) -> Optional[float]:
        return ctx.metric(key)

    for key in ("max_altitude", "max_velocity", "velocity_at_burnout", "landing_velocity",
                "max_g", "accel_at_landing", "dynamic_pressure_max", "max_q_time",
                "max_q_velocity", "parachute_deploy_altitude", "parachute_deploy_time",
                "velocity_before_chute", "velocity_after_chute", "descent_rate_reduction",
                "impact_energy_j", "ground_temperature_c", "apogee_temperature_c",
                "temperature_lapse_rate_c_per_km", "mpu_self_heating_offset_c"):
        f[key] = num(key)
    f["duration"] = ctx.duration()
    for key in EVENT_ORDER:
        f["t_" + key] = ctx.time_of(key)

    def arr_at(attr: str, i: Optional[int], scale: float = 1.0,
               absval: bool = False) -> Optional[float]:
        if an is None or i is None:
            return None
        a = getattr(an, attr, None)
        if a is None:
            return None
        try:
            if int(i) < 0 or int(i) >= len(a):
                return None
            v = float(a[int(i)]) * scale
        except (TypeError, ValueError, IndexError):
            return None
        if not np.isfinite(v):
            return None
        return abs(v) if absval else v

    # ارتفاع/سرعتِ لحظه‌ایِ هر رویداد (همان کاری که جدولِ رویدادهای داشبورد می‌کند)
    for key in EVENT_ORDER:
        f["alt_" + key] = arr_at("alt", ctx.idx.get(key))
        # علامتِ سرعتِ عمودی عمداً حفظ شده: «منفی» بخشی از درسِ همین جدول است (سقوط).
        f["vel_" + key] = arr_at("vel", ctx.idx.get(key))
    f["alt_pre"] = arr_at("alt", 0)
    f["vel_pre"] = arr_at("vel", 0)
    # لحظهٔ Max-Q: نزدیک‌ترین نمونهٔ زمانی به max_q_time
    if an is not None and f.get("max_q_time") is not None and an.t is not None:
        try:
            j = int(np.argmin(np.abs(np.asarray(an.t, dtype=float) - float(f["max_q_time"]))))
            f["alt_max_q"] = arr_at("alt", j)
            f["t_max_q_rel"] = float(an.t[j]) - float(ctx.events.get("launch") or 0.0)
        except (TypeError, ValueError, IndexError):
            pass
    # میانهٔ «سیر صعودی آزاد» (تا سطرِ پنجمِ جدولِ صحنه‌ها با سطرِ اوج یکی نشود)
    if (an is not None and an.t is not None and f.get("t_burnout") is not None
            and f.get("t_apogee") is not None):
        try:
            t0 = float(ctx.events.get("launch") or 0.0)
            mid = 0.5 * (float(f["t_burnout"]) + float(f["t_apogee"])) + t0
            j = int(np.argmin(np.abs(np.asarray(an.t, dtype=float) - mid)))
            f["t_coast_mid"] = float(an.t[j]) - t0
            f["alt_coast_mid"] = arr_at("alt", j)
            f["vel_coast_mid"] = arr_at("vel", j)
        except (TypeError, ValueError, IndexError):
            pass
    # ژرفایِ تحلیلِ خودکار: پایداری، خطای بارومتر، باتری، ماخ
    if an is not None:
        i0, i1 = ctx.idx.get("launch"), ctx.idx.get("burnout")
        osc = None
        for gname in ("GyroY", "GyroZ"):
            try:
                g = an.df.get(gname)
                if g is None or i0 is None or i1 is None or int(i1) <= int(i0):
                    continue
                w = np.asarray(g.to_numpy()[int(i0):int(i1) + 1], dtype=float)
                w = w[np.isfinite(w)]
                if w.size >= 3:
                    osc = (osc or 0.0) + float(np.std(w))
            except (TypeError, ValueError, IndexError, KeyError):
                continue
        f["gyro_osc"] = osc
    if ctx.series is not None:
        ser = ctx.series
        try:
            if "ارتفاع (متر)" in ser.columns and "ارتفاع GPS (متر)" in ser.columns:
                b = pd.to_numeric(ser["ارتفاع (متر)"], errors="coerce").to_numpy(dtype=float)
                g = pd.to_numeric(ser["ارتفاع GPS (متر)"], errors="coerce").to_numpy(dtype=float)
                good = np.isfinite(b) & np.isfinite(g) & (g > 1.0)
                f["baro_gps_valid"] = float(good.mean() * 100.0)
                if good.sum() > 8:
                    # بارومتر ارتفاع را از *سکوی پرتاب* می‌شمارد (AGL) و GPS از *سطح دریا* (MSL)؛
                    # پس تفریقِ مقادیرِ مطلق همیشه ~بلندیِ محلِ پرتاب خطا می‌دهد. درستِ کار:
                    # مقایسهٔ «تغییرات» نسبت به نخستین نمونۀِ معتبرِ هر دو.
                    bb, gg = b[good], g[good]
                    d = (bb - float(np.nanmean(bb[:4]))) - (gg - float(np.nanmean(gg[:4])))
                    d = d[np.isfinite(d)]
                    if d.size:
                        f["baro_gps_mean_abs"] = float(np.abs(d).mean())
                        f["baro_gps_max_abs"] = float(np.abs(d).max())
        except (TypeError, ValueError, KeyError):
            pass
        try:
            if "ولتاژ باتری (V)" in ser.columns:
                v = pd.to_numeric(ser["ولتاژ باتری (V)"], errors="coerce").to_numpy(dtype=float)
                v = v[np.isfinite(v)]
                if v.size > 2:
                    f["batt_start"], f["batt_min"] = float(v[0]), float(v.min())
                    f["batt_end"] = float(v[-1])
        except (TypeError, ValueError, KeyError):
            pass
        try:
            if "سرعت عمودی (م/ث)" in ser.columns:
                v = pd.to_numeric(ser["سرعت عمودی (م/ث)"], errors="coerce").to_numpy(dtype=float)
                v = v[np.isfinite(v)]
                if v.size:
                    f["climb_rate_max"] = float(v.max())
                    f["sink_rate_max"] = float(-v.min())
        except (TypeError, ValueError, KeyError):
            pass
    mach = ctx._col("Mach")
    if mach is not None:
        m = mach[np.isfinite(mach)]
        if m.size:
            f["mach_max"] = float(m.max())
    # مراحلِ صعود/سقوط (برایِ «چرا بعد از اوج هنوز بالا می‌رود؟»)
    if f.get("t_burnout") is not None and f.get("t_apogee") is not None:
        f["coast_time"] = float(f["t_apogee"]) - float(f["t_burnout"])
    if f.get("alt_burnout") is not None and f.get("max_altitude") is not None:
        f["coast_height"] = float(f["max_altitude"]) - float(f["alt_burnout"])
    if f.get("t_parachute") is not None and f.get("t_landing") is not None:
        f["chute_time"] = float(f["t_landing"]) - float(f["t_parachute"])
    if f.get("parachute_deploy_altitude") is not None and f.get("max_altitude"):
        f["chute_frac"] = 100.0 * float(f["parachute_deploy_altitude"]) / float(f["max_altitude"])
    if f.get("velocity_before_chute") and f.get("velocity_after_chute") is not None:
        try:
            v0 = abs(float(f["velocity_before_chute"]))
            v1 = abs(float(f["velocity_after_chute"]))
            if v0 > 1e-6:
                # descent_rate_reduction نسبت بی‌بعد است؛ بازدهی چتر =
                # سهم سرعتی که چتر کم کرده: 100·(1 − v_after/v_before)
                f["chute_eff_pct"] = max(0.0, 100.0 * (1.0 - v1 / v0))
        except (TypeError, ValueError, ZeroDivisionError):
            pass
    return f


def _build_conclusion(wb, ctx: "_Context", theme: Dict[str, str]) -> bool:
    """شیتِ نتیجه‌گیری: داستانِ پرواز + کارنامه + واژه‌نامه + آموزشِ خواندنِ نمودار."""
    ws = wb.create_sheet("نتیجه‌گیری و آموزش")
    _init_sheet(ws, theme, tab_color=COLOR_OK, bg_rows=210, bg_cols=13)
    _widths(ws, [1.6] + [12.0] * 12)
    C1, C2 = 2, 13
    r = _band_title(ws, 2, C1, C2, "نتیجه‌گیری و آموزش -- این پرواز چه گفت؟", theme, COLOR_OK,
                    "این شیت برایِ کسی نوشته شده که جدول و نمودارِ پرواز را تازه دیده است: "
                    "اولِ حرفِ حساب را ساده می‌زند، بعد هر عدد را تعریف می‌کند و در آخر یاد "
                    "می‌دهد نمودارها را چطور بخواند. هیچ‌چیز اینجا قضاوتِ نهاییِ مهندسی نیست؛ "
                    "ابزارِ کمکی است.")
    if not ctx.ok:
        _note(ws, r + 1, C1, C2,
              "دادهٔ پروازیِ قابلِ تحلیلی در این فایل نبود، پس نتیجه‌گیری خودکاری هم "
              "نوشته نشده است. جدولِ «پرسش‌هایِ پرتکرار» و «واژه‌نامه» همچنان خواندنی‌اند.",
              theme, height=26)
        return True
    f = _conc_facts(ctx)

    # ------------------------------------------------------------ ۱) حرفِ حساب
    r = _section(ws, r + 1, C1, C2, "۱) این پرواز در سه جمله", theme,
                 sub="هر عددِ داخلِ متن از همان فایل CSV محاسبه شده است؛ اگر جایی «--» دیدید، "
                     "یعنی آن ستون در لاگ نبود (نه این‌که صفر بوده باشد).")
    lv = f.get("landing_velocity")
    if lv is None:
        lv_txt = "دادهٔ سرعت فرود در دسترس نبود"
    elif 3.0 <= lv <= 8.0:
        lv_txt = "در محدودهٔ ایمن (۳ تا ۸ متر بر ثانیه) است؛ یعنی چتر اندازه بوده"
    elif lv < 3.0:
        lv_txt = "کمتر از ۳ متر بر ثانیه است؛ چتر احتمالاً بزرگ‌تر از نیاز است و باد راکت را زیاد می‌برد"
    else:
        lv_txt = "بیشتر از ۸ متر بر ثانیه است؛ ضربهٔ فرود خطرناک است و سطح چتر باید بیشتر شود"
    story = (
        f"راکت در ثانیهٔ {_fnum(f.get('t_launch'), 1)} از زمین جدا شد و موتور تا ثانیهٔ "
        f"{_fnum(f.get('t_burnout'), 1)} روشن ماند؛ در همان لحظهٔ خاموشی، سرعتِ راکت "
        f"{_fnum(f.get('velocity_at_burnout'), 0)} متر بر ثانیه (برابرِ "
        f"{_fnum((f.get('velocity_at_burnout') or 0) * 3.6, 0)} کیلومتر بر ساعت) بود. "
        f"پس از خاموشی موتور، راکت به‌خاطر لنگشِ خودش {_fnum(f.get('coast_time'), 1)} ثانیه و "
        f"{_fnum(f.get('coast_height'), 0)} متر دیگر بالا رفت تا این‌که در ثانیهٔ "
        f"{_fnum(f.get('t_apogee'), 1)} در ارتفاع {_fnum(f.get('max_altitude'), 0)} متری *ایستاد* "
        "(سرعت عمودی = صفر؛ اینجا «اوج» است).")
    story2 = (
        f"چتر در ثانیهٔ {_fnum(f.get('t_parachute'), 1)} و ارتفاع "
        f"{_fnum(f.get('parachute_deploy_altitude'), 0)} متری باز شد و سرعت سقوط را از "
        f"{_fnum(f.get('velocity_before_chute'), 1)} به {_fnum(f.get('velocity_after_chute'), 1)} "
        f"متر بر ثانیه رساند (کاهشِ {_fnum(f.get('chute_eff_pct'), 0)} درصدی). نزول با چتر "
        f"{_fnum(f.get('chute_time'), 0)} ثانیه طول کشید و فرود در ثانیهٔ "
        f"{_fnum(f.get('t_landing'), 1)} با سرعتِ {_fnum(lv, 1)} متر بر ثانیه انجام شد که "
        f"{lv_txt}. کلِ پرواز {_fnum(f.get('duration'), 0)} ثانیه بود.")
    r = _note(ws, r, C1, C2, "•  " + story, theme, height=28)
    r = _note(ws, r, C1, C2, "•  " + story2, theme, height=28)
    qmax = f.get("dynamic_pressure_max")
    extra = []
    if qmax is not None:
        extra.append(f"پرفشارترین لحظهٔ بدنهٔ راکت (Max-Q) در ثانیهٔ {_fnum(f.get('t_max_q_rel'), 1)} "
                     f"و در ارتفاع {_fnum(f.get('alt_max_q'), 0)} متری بود: فشار دینامیکی "
                     f"{_fnum(qmax, 0)} پاسکال (مثلِ این‌که با آن سرعت در بادِ ثابتِ شدیدی دست خودتان "
                     "را بیرونِ ماشین بگیرید).")
    if f.get("max_g") is not None:
        extra.append(f"بیشینهٔ شتابِ وارد بر راکت {_fnum(f.get('max_g'), 1)} برابرِ گرانشِ زمین بود؛ "
                     "اگر بار مفیدی (دوربین/سنسور) روی راکت است، تحملِ ضربه‌اش باید بیشتر از این عدد باشد.")
    if f.get("mach_max") is not None:
        extra.append(f"بیشینهٔ عددِ ماخ {_fnum(f.get('mach_max'), 2)} بود "
                     f"({_fnum(f.get('max_velocity'), 0)} متر بر ثانیه در برابرِ سرعتِ صوت ~۳۴۰ متر بر ثانیه)؛ "
                     "زیر ۰٫۹ یعنی پرواز کاملاً در محدودهٔ زیرصوت است که برایِ راکت‌های آماتورِ "
                     "آموزشیِ معمول است.")
    for line in extra:
        r = _note(ws, r, C1, C2, "•  " + line, theme, height=28)
    r = _panel(ws, r - len(extra) - 2, C1 - 1, r - 1, C2, theme,
               fill=theme["panel"], title="حرفِ حسابِ پرواز", accent=theme["band"]) + 1

    # ------------------------------------------------------------ ۲) داستانِ لحظه‌به‌لحظه
    r = _section(ws, r, C1, C2, "۲) پرواز لحظه‌به‌لحظه: هر پرده چه شد و چرا مهم است", theme,
                 sub="«زمان» از لحظهٔ پرتاب حساب می‌شود. برایِ دیدنِ همان لحظه در نمودارها به ستونِ "
                     "«کجا در نمودار ببینیم» نگاه کنید؛ اعدادِ ارتفاع و سرعتِ عمودی، مقدارِ واقعیِ "
                     "همان سطرِ لاگ‌اند. دو نکته: (الف) سرعتِ عمودی از *تفاضلِ* دو نمونهٔ متوالی حساب "
                     "می‌شود، پس در سطرهایِ اولِ لاگ می‌تواند چند متر بر ثانیه نویز داشته باشد و در "
                     "نزدیکیِ اوج هم به‌جای صفرِ مطلق، عددی مثلِ ۰٫۵ تا ۱ دیده می‌شود -- صفرِ واقعی "
                     "معمولاً میانِ دو نمونه می‌افتد؛ (ب) ستونِ منفی یعنی راکت دارد پایین می‌آید.")
    scene_rows = [
        ("۱. روی سکو", 0.0, f.get("alt_pre"), f.get("vel_pre"),
         "قبل از جدایی، همه‌چیز ساکت است و فقط حسگرها گرم می‌شوند. عددِ سرعتِ این سطر را جدی نگیرید: "
         "در چند سطرِ اول، تفاضلِ پشت‌سرهم بودنِ نمونه‌ها بزرگ به‌نظر می‌رسد. ولی اگر در همین بازه "
         "نوسانِ *زیاد و بی‌نظم* می‌بینید، یعنی سنسور یا پایه‌اش شل است.",
         "شیت «نمودارهای زمانی»: سطرهایِ اولِ ستونِ «زمان»"),
        ("۲. لحظهٔ رانش", f.get("t_launch"), f.get("alt_launch"), f.get("vel_launch"),
         "موتور روشن می‌شود؛ شتاب و سرعتِ عمودی به‌سرعت بالا می‌روند و فشارِ هوا روی بدنه زیاد "
         "می‌شود. مهم‌ترین پرسشِ این پرده: «آیا راکت صاف رفت؟»",
         "نمودار «شتاب کل» و «سرعت عمودی»؛ قلهٔ اولِ شتاب"),
        ("۳. Max-Q", f.get("t_max_q_rel"), f.get("alt_max_q"), None,
         "بیشینهٔ فشارِ دینامیکی = حساس‌ترین لحظهٔ سازه‌ای. اگر بالچه یا چسبِ بدنه ضعف داشته باشد، "
         "همین‌جا می‌شکند. بعد از Max-Q فشار کم می‌شود، چون هوا رقیق‌تر می‌شود.",
         "نمودار «فشار هوا» و «چگالی هوا» + جدولِ داشبورد"),
        ("۴. پایان رانش", f.get("t_burnout"), f.get("alt_burnout"), f.get("vel_burnout"),
         "سوخت تمام می‌شود. از این لحظه به بعد راکت «سُرخورد» می‌کند: فقط جاذبه و باد او را "
         "کنترل می‌کنند. سرعتِ این لحظه تقریباً بیشینهٔ سرعتِ کل پرواز است.",
         "نمودار «شتاب کل»: افتادنِ ناگهانی به زیرِ ۱g"),
        ("۵. سیر صعودی آزاد (میانهٔ بازه)", f.get("t_coast_mid"), f.get("alt_coast_mid"),
         f.get("vel_coast_mid"),
         f"موتور خاموش است اما راکت {_fnum(f.get('coast_time'), 1)} ثانیه و "
         f"{_fnum(f.get('coast_height'), 0)} متر دیگر بالا می‌رود؛ انرژیِ حرکتی به ارتفاع تبدیل "
         "می‌شود. شیبِ نمودارِ ارتفاع در این پرده کم‌کم صاف می‌شود.",
         "نمودار «ارتفاع»: قله؛ نمودار «سرعت عمودی»: رسیدن به صفر"),
        ("۶. اوج", f.get("t_apogee"), f.get("max_altitude"), f.get("vel_apogee"),
         "درست در قله، سرعتِ عمودی صفر است و راکت یک لحظه «معلق» می‌ماند. اگر چتر در این لحظه "
         "باز شود، کمترین آسیب را می‌بیند؛ به همین دلیل باز شدنِ چتر *بعد از* اوج طبیعی است.",
         "قلهٔ نمودارِ ارتفاع؛ نقطهٔ تلاقی سرعت عمودی با صفر"),
        ("۷. باز شدن چتر", f.get("t_parachute"), f.get("parachute_deploy_altitude"), f.get("vel_chute"),
         "چتر یک ضربهٔ ناگهانی (شکافِ تیزِ پایین‌رونده در نمودار شتاب) و سپس ترمیدِ شدید می‌سازد. "
         "اگر چتر زودتر از اوج باز شود، یعنی ارتفاعِ کافی برایِ آزاد شدنش بوده و راکت هنوز "
         "صعود می‌کرده -- در راکت‌های بالچه‌دار این خطای جدی است.",
         "نمودار «شتاب کل»: تیزترین دره؛ نمودار «سرعت عمودی»: پرش به منفی"),
        ("۸. نزول و فرود", f.get("t_landing"), f.get("alt_landing"), lv,
         "سرعتِ نزول بعد از چند ثانیه ثابت می‌شود («سرعت حدی») چون وزن و پسا برابر شده‌اند. "
         "عددِ مهمِ این پرده سرعتِ لحظهٔ تماس با زمین است.",
         "نمودار «ارتفاع»: فرود آمدنِ نرمال؛ انتهای نمودارِ سرعت عمودی"),
    ]
    body = []
    for name, t, alt, vel, why, where in scene_rows:
        body.append([name, None if t is None else float(t), None if alt is None else float(alt),
                     None if vel is None else float(vel), why, where])
    r = _grid_table(ws, r, first_col=C1,
                    cols=[("پرده", 2), ("زمان (ثانیه)", 1), ("ارتفاع (متر)", 1),
                          ("سرعت عمودی (م/ث) -- منفی = سقوط", 1),
                          ("چه اتفاقی افتاد و چرا مهم است", 4),
                          ("کجا در نمودار ببینیم", 3)],
                    rows=body, theme=theme, num_cols=(1, 2, 3), decimals={1: 2, 2: 0, 3: 1},
                    head_h=40)

    # ------------------------------------------------------------ ۳) کارنامه
    r = _section(ws, r, C1, C2, "۳) کارنامهٔ پرواز (قضاوتِ خودکار با حدِ معروفِ راکت‌های آموزشی)", theme,
                 sub="«حدِ خوب» از قواعدِ رایجِ راکتری آماتور گرفته شده، نه از یک استانداردِ "
                     "الزامی. اگر معیاری '--' شد، دادهٔ آن ستون در لاگ نبود.")
    grades = []
    tones: List[Optional[str]] = []

    def grade(label, value, limit, verdict, tone, plain):
        grades.append([label, value, limit, verdict, plain])
        tones.append(tone)

    if lv is not None:
        ie = f.get("impact_energy_j")
        if ie is None:
            ie_txt = ("انرژیِ ضربهٔ فرود حساب نشد، چون جرمِ کلِ مأموریت در «اطلاعات مأموریت» ثبت "
                      "نبود؛ با ثبتِ جرم، همین عدد هم در گزارش می‌آید. ")
        else:
            ie_txt = (f"انرژیِ ضربهٔ فرود ≈ {_fnum(ie, 0)} ژول؛ برایِ مقایسه، افتادنِ یک توپِ "
                      f"فوتبال (۴۵۰ گرم) از قدِ یک نفر ~ ۸ ژول است. ")
        grade("سرعتِ لحظهٔ فرود", _fnum(lv, 1) + " م/ث", "۳ تا ۸ م/ث",
              "عالی" if 3.0 <= lv <= 8.0 else ("قابل بررسی" if lv < 3.0 else "نیازمند اقدام"),
              "good" if 3.0 <= lv <= 8.0 else ("warn" if lv < 3.0 else "bad"),
              ie_txt + (str(ctx.results.get("chute_suggestion") or "").strip()
                       or "سرعتِ فرود را با تغییرِ سطحِ چتر تنظیم می‌کنند."))
    mg = f.get("max_g")
    if mg is not None:
        grade("بیشینهٔ شتاب", _fnum(mg, 1) + " g", "کمتر از ۱۲g",
              "عالی" if mg < 12 else ("قابل بررسی" if mg < 20 else "نیازمند اقدام"),
              "good" if mg < 12 else ("warn" if mg < 20 else "bad"),
              "۱g را خودِ زمین وارد می‌کند؛ مابقی از موتور و از ضربهٔ چتر است. "
              "سازه و چسب‌ها باید بیشتر از این عدد را تحمل کنند.")
    osc = f.get("gyro_osc")
    if osc is not None:
        grade("لرزشِ چرخشی حین رانش", _fnum(osc, 1) + " °/ث", "کمتر از ۳۰ °/ث",
              "عالی" if osc < 30 else ("قابل بررسی" if osc < 60 else "نیازمند اقدام"),
              "good" if osc < 30 else ("warn" if osc < 60 else "bad"),
              "مجموعِ انحرافِ معیارِ چرخش‌های Pitch و Yaw در بازهٔ روشن بودنِ موتور است؛ کم بودنش "
              "یعنی راکت صاف و با تعادل رفت و زیاد بودنش یعنی بالچه‌ها کوچک‌اند یا "
              "مرکز ثقل عقبِ مرکز فشار است.")
    bg = f.get("baro_gps_mean_abs")
    apog = f.get("max_altitude")
    valid = f.get("baro_gps_valid")
    if valid is not None and valid < 20.0:
        grade("ارتفاعِ GPS (برایِ مقایسه)", "دادهٔ کافی نبود", "بیش از ۲۰٪ سطرهایِ معتبر",
              "قابل بررسی", "warn",
              f"تنها {_fnum(valid, 0)}٪ از سطرهایِ این لاگ ارتفاعِ GPSِ معتبر داشت (قفلِ ماهواره در "
              "ابتدای پرواز و داخلِ محفظه ضعف می‌کند)؛ با دادهٔ کم مقایسه‌ای انجام نمی‌شود. ستونِ "
              "GPS_Altitude را در شیت «دادهٔ خام پرواز» ببینید.")
    elif bg is not None and apog:
        pct = 100.0 * bg / apog
        mx = f.get("baro_gps_max_abs")
        grade("اختلافِ تغییراتِ ارتفاعِ بارومتر و GPS", _fnum(bg, 1) + " متر",
              f"کمتر از ۱۰٪ اوج ({_fnum(0.10 * apog, 0)} متر)",
              "عالی" if pct < 10 else ("قابل بررسی" if pct < 20 else "نیازمند اقدام"),
              "good" if pct < 10 else ("warn" if pct < 20 else "bad"),
              "بارومتر فشار را می‌خواند و به ارتفاع ترجمه می‌کند (به دما و باد حساس است) و GPS از "
              "ماهواره می‌آید. اختلافِ *مقدارِ مطلقِ* این دو بی‌معنی است (یکی از سکوی پرتاب و "
              "دیگری از سطحِ دریا شماری می‌شود)، پس «تغییرات» مقایسه شده است: "
              + _fnum(bg, 1) + f" متر ({_fnum(pct, 1)}٪ اوج)"
              + ("" if mx is None else f"، و بیشترین خطایِ لحظه‌ای {_fnum(mx, 0)} متر")
              + (". اگر این عدد *دقیقاً* صفر شد، نشانهٔ خوبی نیست: یعنی هر دو ستون از یک منبع "
                 "ساخته شده‌اند (در لاگ‌هایِ شبیه‌سازی رایج است) و مقایسه‌ای انجام نشده؛ در دادهٔ "
                 "واقعی همیشه چند ده متر تفاوت هست." if bg is not None and bg < 0.5 else "."))
    if f.get("batt_start") is not None and f.get("batt_min") is not None:
        drop = f["batt_start"] - f["batt_min"]
        grade("افت ولتاژ باتری", _fnum(drop, 2) + " ولت", "کمتر از ۰٫۶ ولت",
              "عالی" if drop < 0.6 else "قابل بررسی", "good" if drop < 0.6 else "warn",
              f"باتری از {_fnum(f['batt_start'], 2)} به حداقلِ {_fnum(f['batt_min'], 2)} و در پایان "
              f"به {_fnum(f.get('batt_end'), 2)} ولت رسید. افتِ زیاد یعنی یا باتری خسته است یا "
              "مصرف (GPS/رادیو/گرم‌کن) بیش از ظرفیتش بوده؛ خطرِ خاموشیِ کامپیوتر پرواز در میانهٔ پرواز.")
    if f.get("parachute_deploy_altitude") is not None and f.get("chute_frac") is not None:
        ok = (f.get("t_parachute") or 0) >= (f.get("t_apogee") or 0)
        grade("باز شدنِ چتر", "در " + _fnum(f.get("chute_frac"), 0) + "٪ ارتفاعِ اوج",
              "بعد از اوج و در ارتفاعِ زیاد", "عالی" if ok else "نیازمند اقدام",
              "good" if ok else "bad",
              "چتر باید *بعد از* اوج و وقتی سرعت تقریباً صفر است باز شود تا پارچه پاره نشود. در این "
              "پرواز " + ("چتر درست *بعد از* اوج باز شد؛ ترتیبِ رویدادها سالم است." if ok else
                          "چتر *پیش از* اوج باز شده -- ارتفاعِ باز شدن یا تایمر را در تنظیماتِ "
                          "کامپیوتر پرواز اصلاح کنید."))
    cr, sr = f.get("climb_rate_max"), f.get("sink_rate_max")
    if cr is not None and sr is not None:
        grade("بیشینهٔ سرعتِ صعود در برابرِ سقوط", _fnum(cr, 1) + " به " + _fnum(sr, 1) + " م/ث",
              "— (سطرِ اطلاعاتی)", "—", None,
              "نمودارِ سرعتِ عمودی دو قله دارد: یکی در صعود و دیگری درست پیش از باز شدن چتر. "
              "فاصلهٔ زیادِ این دو یعنی چتر خوب کار کرده است؛ برایِ همین این سطر قضاوتِ «خوب/بد» "
              "ندارد و فقط به دردِ مقایسهٔ دو پرواز می‌خورد.")
    lapse = f.get("temperature_lapse_rate_c_per_km")
    if lapse is not None:
        grade("نرخِ کاهش دما با ارتفاع", _fnum(lapse, 2) + " °C بر کیلومتر", "نزدیکِ ۶٫۵- (منفی = سرد شدن)",
              "قابل بررسی" if abs(abs(lapse) - 6.5) > 3 else "عالی",
              "warn" if abs(abs(lapse) - 6.5) > 3 else "good",
              "در جوِ استاندارد، هوا تقریباً به ازای هر کیلومترِ ارتفاع ~۶٫۵ درجه سردتر می‌شود. عددِ خیلی "
              "نامعمول یعنی سنسورِ دما یا بازهٔ ارتفاع مشکوک است (مثلاً به‌جایِ دمای هوا، دمایِ خودِ "
              "ماژول ژیروسکوپ را میانگین گرفته‌اند). این گزارش برآیند را فقط از ماژول‌های محیطی "
              "حساب می‌کند.")
    if not grades:
        grades.append(["--", "داده‌ای برایِ قضاوت نبود", "—", "—",
                       "هیچ‌یک از ستون‌هایِ لازمِ این معیارها در فایل نبود."])
        tones.append(None)
    r = _grid_table(ws, r, first_col=C1,
                    cols=[("معیار", 3), ("مقدارِ این پرواز", 2), ("حدِ خوب", 2),
                          ("قضاوت", 1), ("چرا این عدد مهم است / چه کار کنیم", 4)],
                    rows=grades, theme=theme, tones=tones, tone_col=3, head_h=34)

    # ------------------------------------------------------------ ۴) واژه‌نامه
    r = _section(ws, r, C1, C2, "۴) هر عددِ این گزارش یعنی چه (واژه‌نامهٔ دانش‌آموزی)", theme,
                 sub="تعریف‌ها عمداً ساده شده‌اند؛ ستونِ آخر می‌گوید چه شکلی از این عدد «زنگِ خطر» است.")
    gloss = [
        ("ارتفاع (متر)", "فاصلهٔ عمودیِ راکت از نقطهٔ پرتاب. مهم‌ترین عددِ هر پرواز «بیشینهٔ» "
         "همین است که به آن «اوج» می‌گویند.",
         "متر", _fnum(f.get("max_altitude"), 0),
         "اگر نمودار ارتفاع در انتها به صفر نزدیک نشود، یعنی داده قطع شده یا راکت گم شده است."),
        ("سرعت عمودی (م/ث)", "فقط بخشِ بالا/پایینِ حرکت. مثبت یعنی بالا می‌رود، منفی یعنی "
         "پایین می‌آید، صفر یعنی در اوج (یا در یک لحظهٔ مکث) است.",
         "متر بر ثانیه",
         (_fnum(cr, 1) + " صعودی / " + _fnum(sr, 1) + " سقوطی") if cr else "--",
         "اگر در صعود عدد *هیچ‌وقت* به صفر نرسد و نمودار ناگهان منفی شود، یعنی راکت در حال "
         "بالا رفتن متوقف شده (موتور ضعیف یا بدنه سنگین)."),
        ("سرعت کل (م/ث)", "بلندیِ حرکت در هر جهتی (جذرِ مجموعِ مجذورِ مؤلفه‌ها). سرعتِ کل هیچ‌وقت "
         "از سرعتِ عمودی کمتر نیست.",
         "متر بر ثانیه", _fnum(f.get("max_velocity"), 0),
         f"برایِ درکِ بهتر: معادلِ {_fnum((f.get('max_velocity') or 0) * 3.6, 0)} کیلومتر بر ساعت "
         "(ضریبِ ۳٫۶ را همیشه می‌شود ذهنی زد: m/s × ۳٫۶ = km/h)."),
        ("شتاب کل (g)", "چند برابرِ گرانشِ زمین روی راکت نیرو آمده. ۱g را خودِ زمین "
                 "وارد می‌کند حتی وقتی راکت ساکن است؛ روی موتور و ضربهٔ چتر اضافه می‌شود.",
                 "g (۹٫۸۱ م/ث²)", _fnum(f.get("max_g"), 1),
                 "قله‌هایِ تیزِ چندg در لحظهٔ باز شدن چتر طبیعی است؛ قله‌هایِ پُرتکرارِ بی‌نظم "
                 "یعنی ضربه/لرزشِ غیرمنتظره."),
        ("فشار دینامیکی q (پاسکال)", "فشاری که «حرکت در هوا» ایجاد می‌کند: q = ½·ρ·v². "
         "بیشینهٔ آن Max-Q است.", "پاسکال", _fnum(f.get("dynamic_pressure_max"), 0),
         "هرچه Max-Q دیرتر/بالاتر بیفتد، بارِ سازه‌ای بیشتر است؛ برایِ بدنهٔ کاغذی/کارتونی "
         "مقادیرِ چند‌هزار پاسکالی خطرناک است."),
        ("عددِ ماخ (Mach)", "سرعتِ راکت تقسیم بر سرعتِ صوت. ۱٫۰ یعنی هم‌سرعتِ صوت.", "بی‌واحد",
         _fnum(f.get("mach_max"), 2),
         "بالای ~۰٫۹ برایِ أكثر راکت‌های آماتور خطرناک است (موجِ ضربه، گرمایش، ناپایداری)."),
        ("فشار هوا (hPa)", "فشارِ ستونِ هوا تا بالای سرِ راکت. با ارتفاع کم می‌شود؛ "
         "سنسورِ بارومتر همین را می‌خواند و به ارتفاع ترجمه می‌کند.",
         "هکتوپاسکال", "--", "اگر نمودار فشار «پله» بی‌دلیل داشته باشد، گیرنده فشار گرفته/مسدود شده است."),
        (f"دمای سه ماژول ({ctx.temp_mods['hum']} / {ctx.temp_mods['baro']} / {ctx.temp_mods['gyro']})",
         f"سه سنسور دما داریم: ماژول دما و رطوبت ({ctx.temp_mods['hum']}) و ماژول فشار و دما "
         f"({ctx.temp_mods['baro']}) هوای بیرون را می‌خوانند؛ ماژول ژیروسکوپ "
         f"({ctx.temp_mods['gyro']}) گرمای خودِ تراشه را. "
         "برآیندِ دمای هوا = میانگینِ دو ماژولِ اول (بدونِ ژیروسکوپ).",
         "درجه سلسیوس",
         f"{_fnum(f.get('ground_temperature_c'), 1)} در زمین → {_fnum(f.get('apogee_temperature_c'), 1)} در اوج",
         f"اگر ماژول ژیروسکوپ {_fnum(f.get('mpu_self_heating_offset_c'), 1)} درجه از هوا گرم‌تر بود "
         "طبیعی است؛ ولی اگر اختلاف مرتب زیاد می‌شود، جای بورد در محفظه و تهویه را بررسی کنید. "
         "هیچ محاسبهٔ جوّی با دمای ژیروسکوپ انجام نمی‌شود."),
        ("شتاب سه‌محور X/Y/Z (m/s²)", "نیرو در سه جهتِ بدنهٔ راکت. در حالت عادی روی یک محور "
         "(~۹٫۸) می‌ماند و بقیه نزدیکِ صفر است.", "متر بر ثانیه²", "--",
         "اگر هر سه محور مرتب می‌لرزند، راکت در هوا چرخشِ بی‌کنترل دارد (ناپایداری آیرودینامیکی)."),
        ("چرخش ژیروسکوپ (°/ث)", "سرعتِ چرخشِ راکت حولِ محورهایش (مثلِ پیچ‌وخارط شدنِ توپ).",
         "درجه بر ثانیه", (_fnum(f.get("gyro_osc"), 1) + " (میانگینِ حین رانش)")
         if f.get("gyro_osc") is not None else "--",
         "دامنهٔ بالا در ابتدای پرواز = بالچه‌ها کوچک‌اند یا مرکز ثقل بیش از حد عقب است."),
        ("ولتاژ باتری (V)", "انرژیِ برقیِ بورد. هرچه مصرف بیشتر و باتری خسته‌تر، افت بیشتر.",
         "ولت", (f"{_fnum(f.get('batt_start'), 2)} → {_fnum(f.get('batt_min'), 2)}"
                 if f.get("batt_start") is not None else "--"),
         "افتِ نزدیک به انتهای پرواز یعنی ریسکِ خاموشیِ ضبطِ داده؛ همیشه با ۲۰٪ حاشیه فرود بیایید."),
        ("مرحلهٔ پرواز", "برچسبی که نرم‌افزار به هر سطر می‌دهد: روی سکو، مرحلهٔ رانش، سیر صعودی "
         "آزاد، سقوط آزاد، نزول با چتر، فرود.", "متن", "--",
         "اگر برچسبِ مرحله با نمودارِ شما نمی‌خواند، آستانهٔ تشخیصِ رویدادها (مثلاً سرعتِ حدِ "
         "فرود) را در نرم‌افزار بازبینی کنید."),
    ]
    r = _grid_table(ws, r, first_col=C1,
                    cols=[("پارامتر", 2), ("به زبانِ ساده", 5), ("واحد", 1),
                          ("در این پرواز", 1), ("چه‌وقت نگران‌کننده است؟", 3)],
                    rows=gloss, theme=theme, head_h=40)

    # ------------------------------------------------------------ ۵) خواندن نمودار
    r = _section(ws, r, C1, C2, "۵) نمودارها را چطور بخوانیم (هفت قدمِ همیشه‌کار)", theme,
                 sub="این هفت قدم برایِ هر نمودارِ خطیِ زمانی در همین فایل جواب می‌دهد؛ "
                     "بعدش یک جدولِ «شکلِ نمودار ← معنی‌اش» آمده است.")
    steps = [
        "قدمِ ۱ -- اولِ محورها را بشناسید: محورِ افقیِ نمودارهایِ این فایل «زمان (ثانیه)» است و "
        "محورِ عمودی «مقدارِ پارامتر». چون شیت راست‌به‌چپ است، زمان از سمتِ *راست* شروع می‌شود و "
        "به سمتِ چپ زیاد می‌شود؛ خیلی از اشتباهاتِ خواندن فقط همین است.",
        "قدمِ ۲ -- مقیاسِ عمودی را بخوانید: ۱۰ متر با ۱۰۰ متر روی کاغذ یک شکل می‌شود. هر وقت "
        "«شیب» را قضاوت می‌کنید، اولِ عددِ دو انتها را از روی محورِ عمودی بردارید.",
        "قدمِ ۳ -- شکلِ کلی را در یک کلمه بگویید: بالا رونده؟ صاف؟ قله‌ای؟ پله‌ای؟ نوسانی؟ این "
        "کلمه، نیمِ تحلیل است.",
        "قدمِ ۴ -- شیب = سرعتِ تغییر. در نمودارِ ارتفاع، شیبِ همان لحظه یعنی سرعتِ عمودی؛ جایی "
        "که نمودارِ ارتفاع صاف (افقی) شد، سرعتِ عمودی صفر است؛ پس آنجا اوج را دارید.",
        "قدمِ ۵ -- قله‌ها و دره‌ها را رویدادها توضیح می‌دهند: هر شکافِ تیز یک «اتفاق» است -- "
        "روشن شدنِ موتور، Max-Q، پایانِ رانش، باز شدنِ چتر، فرود. جدولِ «زمانیِ رویدادها» در "
        "شیت داشبورد همین‌ها را ثانیه‌به‌ثانیه نوشته است.",
        "قدمِ ۶ -- برایِ عددِ دقیق، روی منحنی نشانگر را نگه دارید (Tooltip اکسل) یا در ستونِ "
        "همان سطرِ جدولِ «دادهٔ خام پرواز» بگردید؛ از رویِ شبکهٔ نمودار عددِ دقیق به دست نمی‌آید "
        "(و خط‌کشیِ افقی هم عمداً حذف شده تا نمودار چشم‌نواز بماند).",
        "قدمِ ۷ -- یک پیش‌بینی کنید و بعد داده را چک کنید: مثلاً «چون موتور در ثانیهٔ ۳ تمام "
        "می‌شود، نمودارِ سرعتِ عمودی باید بعدش نرمال نزولی شود». همین کارِ یک مهندسِ تست است.",
    ]
    for i, line in enumerate(steps):
        r = _note(ws, r, C1, C2, line, theme, height=26)
    r = _panel(ws, r - len(steps) - 1, C1 - 1, r - 1, C2, theme,
               fill=theme["panel"], title="هفت قدمِ خواندنِ هر نمودار", accent=theme["band"]) + 1
    shapes = [
        ("قلهٔ تیزِ نمودارِ ارتفاع", "اوجِ پرواز: لحظه‌ای که سرعتِ عمودی از مثبت به منفی می‌رسد.",
         f"در ثانیهٔ {_fnum(f.get('t_apogee'), 1)} و ارتفاعِ {_fnum(f.get('max_altitude'), 0)} متر"),
        ("صاف شدنِ تدریجیِ شیبِ صعود", "موتور خاموش شده و راکت فقط با لنگش بالا می‌رود (سیر صعودی آزاد).",
         f"از {_fnum(f.get('t_burnout'), 1)} تا {_fnum(f.get('t_apogee'), 1)} ثانیه"),
        ("عبورِ نمودار از خطِ صفر به منفی", "پایانِ صعود و شروعِ سقوط؛ اگر بی‌مقدمه و ناگهانی باشد، "
         "یعنی چیزی جدا شده یا شکسته است (بالچه، بدنه، چترِ اولیه).", "نمودارِ «سرعت عمودی»"),
        ("درهٔ تیزِ رو به پایین در شتاب", "ضربهٔ باز شدن چتر؛ یک لحظهٔ کوتاه با g زیاد.",
         f"نزدیکِ ثانیهٔ {_fnum(f.get('t_parachute'), 1)}"),
        ("خطِ صافِ افقی در سرعتِ عمودی", "سرعتِ حدی: وزن و پسا برابر شده‌اند؛ نزول با چتر باید این شکلی باشد.",
         f"بعد از {_fnum(f.get('t_parachute'), 1)} ثانیه؛ مقدارش ~{_fnum(f.get('velocity_after_chute'), 1)} م/ث"),
        ("دو خطِ موازیِ نزدیک‌به‌هم (بارومتر و GPS)", "هر دو ارتفاع را می‌گویند؛ اختلافِ ثابتِ آن‌ها "
         "خطایِ سیستماتیکِ بارومتر است (دما/باد)، نه خرابیِ حسگر.",
         f"میانگینِ اختلافِ قدرمطلق: {_fnum(f.get('baro_gps_mean_abs'), 1)} متر"),
        ("نوسانِ ریزِ بی‌نظم در ژیروسکوپ", "لرزش/تکانهٔ راکت. کم و منظم = پایدار؛ زیاد = ناپایدار.",
         f"مجموعِ انحرافِ معیار حین رانش: {_fnum(f.get('gyro_osc'), 1)} °/ث"),
        ("پله در انتهای نمودارِ ولتاژ", "یک بارِ برقی اضافه شده یا باتری به ته رسیده.",
         f"افتِ کل: {_fnum((f.get('batt_start') or 0) - (f.get('batt_min') or 0), 2)} ولت"),
        ("برشِ بزرگِ نمودار دونات", "مرحله‌ای که بیشترین *زمان* را برده، نه لزوماً بیشترین ارتفاع را؛ "
         "معمولاً نزول با چتر.", "شیت «تحلیل مراحل پرواز» + ستونِ «سهم از زمان کل (٪)»"),
        ("ستون‌هایِ بلند در نمودارِ میله‌ای", "بیشینهٔ سرعتِ عمودیِ هر مرحله -- مقیاسِ محورش برایِ "
         "همهٔ مراحل یکی است، پس ارتفاعِ ستون‌ها قابلِ مقایسه است.", "شیت «تحلیل مراحل پرواز»"),
        ("نیم‌دایره‌هایِ رنگیِ داشبورد", "گیج: درصدِ رسیدن به «سقفِ مرجع»ِ همان پارامتر. "
         "زیرِ ۶۰٪ سبز، ۶۰ تا ۹۰٪ کهربایی و بالایِ ۹۰٪ قرمز است؛ قرمز *به‌خودی‌خود* بد نیست، "
         "فقط یعنی به حدِ طراحی نزدیک شده‌اید.",
         "شیت «داشبورد پرواز» (چهار نیم‌دایرهٔ زیرِ کارت‌ها)"),
        ("خطِ چگالی هوا که آهسته پایین می‌آید", "هوا با ارتفاع رقیق‌تر می‌شود؛ این خط را خودِ "
         "نرم‌افزار از فشار و دمای همان لحظه حساب کرده (ρ = p/(R·T)).",
         "نمودارِ «چگالی هوا نسبت به زمان»"),
    ]
    r = _grid_table(ws, r, first_col=C1,
                    cols=[("اگر این شکل را دیدید…", 3), ("معنی‌اش این است", 5),
                          ("در این پرواز", 4)],
                    rows=shapes, theme=theme, head_h=32)

    # ------------------------------------------------------------ ۶) تمرین
    r = _section(ws, r, C1, C2, "۶) تمرین‌هایِ دستی رویِ همین فایل (خودتان حساب کنید)", theme,
                 sub="سلول‌هایی که با «=» شروع شوند در اکسل فرمول‌اند؛ متن‌هایِ زیر را در یک "
                     "سلولِ *خالیِ جدید* تایپ کنید (در همین شیت یا یک شیت تازه). فرمول‌ها هیچ "
                     "چیزی در فایلِ اصلی عوض نمی‌کنند.")
    ex = [
        ("۱. اوج را خودتان پیدا کنید",
         "ببینید آیا عددی که نمودار و کارتِ داشبورد نشان می‌دهند با خودِ داده‌هایِ خام یکی است؛ "
         "این مهم‌ترین عادتِ یک مهندسِ تست است («دو منبعِ مستقل، یک جواب»).",
         "به شیت «دادهٔ خام پرواز» بروید (سربرگ در سطرِ ۵ است: A «ردیف»، B «زمان (ثانیه)»، "
         "C «ارتفاع (متر)»). در یک سلولِ خالی -- در همین شیت یا یک شیتِ تازه -- تایپ کنید: "
         "  =MAX(C:C)  و باید همان «اوج پرواز»ِ داشبورد را بدهد "
         f"(این پرواز: {_fnum(f.get('max_altitude'), 0)} متر). اگر فایلِ شما ستون‌هایِ دیگری دارد، "
         "همان قانون را با حرفِ ستونِ ارتفاع تکرار کنید."),
        ("۲. سرعتِ میانگینِ صعود را حساب کنید",
         "سرعتِ میانگین = جابه‌جایی تقسیم بر زمان. این عدد از «بیشینهٔ سرعت» کمتر است و نشان می‌دهد "
         "راکت به‌طورِ میانگین چقدر طول کشیده تا به اوج برسد.",
         "در شیت «دادهٔ نمودارها (فنی)» سربرگ در سطرِ ۳ است و B ستونِ «زمان» و C ستونِ «ارتفاع»؛ "
         "در یک سلولِ خالی تایپ کنید:  =(MAX(C:C)-MIN(C:C))/(MAX(B:B)-MIN(B:B))  و بعد با «بیشینهٔ سرعت» "
         f"({_fnum(f.get('max_velocity'), 0)} م/ث) مقایسه کنید."),
        ("۳. هر رویداد را به یک سطرِ لاگ وصل کنید",
         "تا وقتی ندانید «اوج» سطرِ چندمِ فایل است، تحلیلِ دستی معنی ندارد. این کار را که یاد بگیرید، "
         "می‌توانید هر نمودار را با همان سطرِ جدول چک کنید.",
         "در شیت «دادهٔ خام پرواز» (سربرگ در سطرِ ۵: A ردیف، B زمان، C ارتفاع) به‌جای Sort کردن -- "
         "که ترتیبِ داده را به‌هم می‌ریزد -- شمارهٔ سطرِ اوج را با فرمول بگیرید: "
         "  =MATCH(MAX(C:C),C:C,0)  (چون محدوده از سطرِ ۱ شروع می‌شود، همان شمارهٔ سطرِ "
         "واقعی را می‌دهد) و بعد زمانِ همان سطر را با  =INDEX(B:B,n)  (به‌جای n، همان "
         f"شماره) بخوانید. زمانِ اوج این پرواز {_fnum(f.get('t_apogee'), 2)} ثانیه است؛ اگر عددتان "
         "نزدیکِ همین شد، تحلیلِ دستی‌تان درست کار می‌کند. MATCH و INDEX تقریباً همهٔ کارِ لازم را "
         "در اکسل راه می‌اندازند."),
        ("۴. تأثیر چتر را خودتان اندازه بگیرید",
         "عددِ «کاهشِ سرعتِ سقوط» که نرم‌افزار گزارش می‌کند، فقط تفریقِ دو اندازهٔ سرعت است؛ "
         "وقتی خودتان تفریق کنید، به نتیجه اعتماد می‌کنید.",
         f"در جدولِ «پرواز لحظه‌به‌لحظهٔ» همین شیت، دو سطرِ «باز شدن چتر» و «نزول و فرود» را "
         "ببینید (سرعتِ عمودیِ "
         f"{_fnum(f.get('velocity_before_chute'), 1)} و {_fnum(f.get('velocity_after_chute'), 1)} "
         "متر بر ثانیه). حالا در یک سلولِ خالی تایپ کنید:  =ROUND(X-Y,2)  و به‌جای X و Y آن دو عدد "
         "را بگذارید. نکتهٔ مهم: داخلِ فرمول حتماً ارقامِ *لاتین* بنویسید -- اکسل ارقامِ فارسی را در "
         "فرمول نمی‌شناسد؛ ارقامِ فارسیِ این فایل فقط «نمایشِ» سلول‌اند."),
        ("۵. واحدِ ذهنی بسازید",
         "«۱۵ متر بر ثانیه» برایِ هیچ کسی ملموس نیست؛ تا آن را به کیلومتر بر ساعت یا «زمانِ رسیدن "
         "به انتهایِ یک خیابان» تبدیل نکنید، عدد فقط حفظ می‌شود.",
         "قانونِ ساده: m/s × ۳٫۶ = km/h. پس بیشینهٔ سرعت این پرواز "
         f"{_fnum(f.get('max_velocity'), 0)} × ۳٫۶ ≈ {_fnum((f.get('max_velocity') or 0) * 3.6, 0)} "
         "کیلومتر بر ساعت است و سرعت فرود "
         f"{_fnum(lv, 1)} × ۳٫۶ ≈ {_fnum((lv or 0) * 3.6, 1)} کیلومتر بر ساعت."),
    ]
    r = _grid_table(ws, r, first_col=C1,
                    cols=[("تمرین", 3), ("چرا مفید است", 5), ("در اکسل این‌طور انجامش دهید", 4)],
                    rows=ex, theme=theme, head_h=32)

    # ------------------------------------------------------------ ۷) پرسش و پاسخ
    r = _section(ws, r, C1, C2, "۷) پرسش‌هایِ پرتکرارِ کسانی که تازه این گزارش را می‌بینند", theme)
    qa = [
        ("چرا راکت بعد از خاموش شدنِ موتور هنوز بالا می‌رود؟",
         "چون «لنگش» (اینرسی) دارد: تا وقتی هوا و جاذبه او را نکَنند، همان سرعتِ اولش را نگه می‌دارد و "
         "کم‌کم جاذبه او را کند می‌کند. به همین دلیل بینِ پایانِ رانش تا اوج یک بازهٔ زمانی هست "
         f"(در این پرواز {_fnum(f.get('coast_time'), 1)} ثانیه و {_fnum(f.get('coast_height'), 0)} متر)."),
        ("چرا «اوج» با «بیشینهٔ سرعت» فرق دارد؟",
         "بیشینهٔ سرعت در لحظهٔ خاموشیِ موتور است؛ اوج جایی است که سرعتِ عمودی صفر شده. پس اوج "
         "همیشه *بعد* از بیشینهٔ سرعت می‌آید. در نمودارِ ارتفاع، اوج = قله؛ در نمودارِ سرعت، اوج = "
         "عبور از صفر."),
        ("چرا سرعتِ عمودی منفی می‌شود؟",
         "علامتِ منفی یعنی جهتِ حرکت پایین است. سرعت هرگز نمی‌تواند «کمتر از صفر» باشد چون "
         "سرعتِ *کل* (Mach/speed) قدرمطلق است؛ آنچه منفی می‌شود مؤلفهٔ عمودیِ آن است."),
        ("چرا عددِ m/s برایم ملموس نیست؛ چطور به کیلومتر بر ساعت تبدیلش کنم؟",
         f"ضریبِ تبدیل ۳٫۶ است: {_fnum(f.get('max_velocity'), 0)} متر بر ثانیه × ۳٫۶ = "
         f"{_fnum((f.get('max_velocity') or 0) * 3.6, 0)} کیلومتر بر ساعت. برایِ سرعتِ فرود هم "
         f"همین کار را بکنید: {_fnum(lv, 1)} م/ث = {_fnum((lv or 0) * 3.6, 1)} کیلومتر بر ساعت."),
        ("چرا ارتفاعِ بارومتر و ارتفاعِ GPS با هم فرق دارند؟",
         f"بارومتر فشار را می‌خواند و از رویِ «جوِ استاندارد» به ارتفاع ترجمه می‌کند؛ هوا و دمای واقعی "
         f"با استاندارد فرق دارند. GPS موقعیتِ ماهواره‌ای می‌دهد و خطایش مستقل از هواست، ولی در لحظه‌هایِ "
         f"سریع و در محفظهٔ بسته بد کار می‌کند. اختلافِ میانگینِ این پرواز {_fnum(f.get('baro_gps_mean_abs'), 1)} "
         "متر است که برایِ یک پروازِ کوتاه طبیعی به‌شمار می‌آید."),
        ("چرا بعضی نمودارها «۱ واحد روی نمودار = …» دارند؟",
         "چون در شیتِ فنی، داده برایِ روان شدنِ فایل نمونه‌برداری/میان‌یابی شده است. مقیاسِ محور "
         "همین را جبران می‌کند؛ یعنی شکلِ نمودار واقعی است ولی هر «واحدِ شبکه» ممکن است چند ثانیه یا "
         "چند متر باشد. برایِ عددِ دقیق به جدولِ همان شیت نگاه کنید."),
        ("منظور از «سهم از زمان کل (٪)» چیست؟",
         "هر مرحلهٔ پرواز چقدر از زمانِ کل را گرفته است (مثلاً نزول با چتر معمولاً بیشترین سهم را دارد "
         "چون کندترین بخش است). این درصد ربطی به ارتفاع ندارد و فقط «وقتش» را نشان می‌دهد."),
        ("چرا روی برش‌ها و نقاطِ نمودار عدد نوشته نشده؟",
         "تجربه نشان داده نوشتهٔ داخلِ نمودار خواندن را سخت (و در برخی نسخه‌هایِ اکسل فایل را "
         "ناخواند) می‌کند. همهٔ اعدادِ دقیق در جدول‌هایِ همان شیت‌اند؛ نمودار فقط «شکل» را نشان می‌دهد."),
        ("«شیت دادهٔ نمودارها (فنی)» را پاک کنم؟",
         "نه؛ نمودارها به سلول‌هایِ همان شیت ارجاع می‌دهند و اگر پاکشان کنید نمودارها خالی می‌شوند. "
         "برایِ خواندنِ فایل، همان شیت‌هایِ اول کافی است."),
        ("عددِ '--' یعنی صفر؟",
         "نه؛ یعنی داده نبود. مثلاً اگر در لاگِ شما ستونِ رطوبت یا ولتاژ نباشد، نمودارش هم ساخته "
         "نمی‌شود و جایِ آن '--' می‌بینید. صفرِ واقعی دقیقاً نوشته می‌شود."),
        ("از کجا بفهمم داده‌ها کامل است؟",
         "تعدادِ سطرها را در شیتِ «دادهٔ خام پرواز» بشمارید و با «زمانِ کل × نرخِ نمونه‌برداری» بسنجید؛ "
         "اختلافِ زیاد یعنی فریمور در میانهٔ پرواز قطعیِ داده داشته است."),
        ("این نتیجه‌گیری را می‌توانم در گزارشِ کلاسی بنویسم؟",
         f"بله -- ولی همیشه عددِ اصلی را هم ذکر کنید و منبعش را (نام فایل CSV). همین شیت از دادهٔ "
         f"«{_fnum(len(ctx.df) if ctx.df is not None else None, 0)} سطر» محاسبه شده و هیچ عددی در آن "
         "دستی وارد نشده است."),
    ]
    r = _grid_table(ws, r, first_col=C1, cols=[("پرسش", 4), ("پاسخ", 8)], rows=qa, theme=theme,
                    head_h=28, size=9.5)

    # ------------------------------------------------------------ ۸) قدمِ بعد
    r = _section(ws, r, C1, C2, "۸) حالا چه کار کنیم؟ (قدم‌هایِ بعدیِ این تیم)", theme,
                 sub="این فهرست از همان قضاوت‌هایِ «کارنامهٔ پرواز» بالا ساخته شده است؛ هر موردِ "
                     "مطلوب هم نوشته می‌شود تا تغییراتِ بی‌مورد ندهید.")
    actions: List[str] = []
    for label, tone in zip([g[0] for g in grades], tones):
        if tone == "bad":
            actions.append(f"«{label}» خارجِ محدودهٔ مطمئن است: اولِ همین را در پروازِ بعدی درست کنید "
                           "و بعد سراغِ بهبودِ عملکردِ کلِ راکت بروید.")
        elif tone == "warn":
            actions.append(f"«{label}» قابلِ بررسی است: یک بارِ دیگر با داده‌هایِ جدید چک کنید؛ "
                           "لازم نیست حتماً چیزی را تغییر دهید.")
    good_n = sum(1 for t in tones if t == "good")
    if good_n:
        actions.append(f"{fa_num(good_n, 0)} معیار از {fa_num(len(grades), 0)} معیار در محدودهٔ مطلوب بود؛ "
                       "آن‌ها را *دست نزنید* -- در راکت‌آموزشی، تغییر دادنِ چیزِ سالم، رایج‌ترین علتِ "
                       "خراب شدنِ پروازِ بعدی است.")
    actions.append("قبل از پروازِ بعدی: ۱) بالچه‌ها و چسب‌ها را چشمی بررسی کنید، ۲) آستانهٔ باز شدن "
                   "چتر را با ارتفاعِ واقعیِ این پرواز تنظیم کنید، ۳) ستون‌هایِ Humidity / UV_Index / "
                   "Voltage را در فریمور فعال کنید تا این گزارش کامل‌تر شود، ۴) یک ویدئو از پرتاب بگیرید "
                   "تا بتوانید نمودارِ چرخش را با تصویرِ واقعی مقایسه کنید.")
    actions.append("اگر می‌خواهید این پرواز را با پروازِ دیگر مقایسه کنید: دو فایلِ Excel را در دو "
                   "پنجره باز کنید و فقط «اوج»، «بیشینهٔ شتاب»، «سرعت فرود» و «نوسانِ چرخشی» را با هم "
                   "قیاس کنید -- همین چهار عدد بیشترِ داستان را می‌گویند.")
    for line in actions:
        r = _note(ws, r, C1, C2, "•  " + line, theme, height=26)
    r = _panel(ws, r - len(actions) - 1, C1 - 1, r - 1, C2, theme,
               fill=theme["canvas2"], title="قدمِ بعدی", accent=theme["band"]) + 1

    # ------------------------------------------------------------ پاورقی
    foot = _merge(ws, r + 1, C1, r + 1, C2,
                  "این شیت خودکار از رویِ داده‌هایِ همین فایل ساخته شده و جایِ ارزیابیِ ایمنیِ انسانی و "
                  "قوانینِ محلی را نمی‌گیرد. «حدِ خوب»ها تجربه‌هایِ رایجِ راکت‌هایِ آموزشیِ آماتورند؛ "
                  "برایِ بدنهٔ غیرمعمول، سوختِ پرفشار یا مأموریتِ با بارِ حساس، این عددها کافی نیستند.")
    foot.font = _font(theme, 9, False, theme["dim"])
    foot.alignment = _align("right", wrap=True, indent=1)
    ws.row_dimensions[r + 1].height = 30
    ws.print_area = f"A1:{get_column_letter(C2)}{r + 1}"
    return True

# =========================================================================
# شیت راهنما
# =========================================================================
def _build_guide(wb, ctx: _Context, theme: Dict[str, str], charts: int):
    ws = wb.create_sheet("راهنمای خواندن")
    _init_sheet(ws, theme, tab_color=COLOR_INFO, bg_rows=8 + 12 * 2, bg_cols=6)
    _widths(ws, [1.6, 32, 25, 25, 25, 25])
    r = _band_title(ws, 2, 2, 6, "راهنمای خواندن این فایل", theme, COLOR_INFO,
                    "اگر اولین بار است با این گزارش کار می‌کنید، از همین صفحه شروع کنید.")
    r += 1
    landing = ctx.metric("landing_velocity")
    items = [
        ("هر شیت برای چه است؟",
         "«داشبورد پرواز» خلاصهٔ یک‌نگاه (کارت‌ها، گیج‌ها، خطِ زمانیِ رویدادها، ارزیابیِ خودکار). "
         "«نتیجه‌گیری و آموزش» همان حرفِ حساب به زبانِ ساده + کارنامه + واژه‌نامه + آموزشِ خواندنِ "
         "نمودار (اگر فقط یک شیت را می‌خواهید بخوانید، این باشد). «نمودارهای زمانی» همهٔ پارامترها "
         "نسبت به زمان؛ «پروفایل ارتفاع» تغییرات دما/فشار نسبت به *ارتفاع* (نه زمان) که برایِ "
         "مقایسهٔ دو پرواز با ارتفاعِ مختلف مناسب‌تر است؛ «تحلیل مراحل پرواز» مدت، سرعت و شتابِ هر "
         "مرحله؛ «دادهٔ نمودارها (فنی)» همان اعدادی است که نمودارها می‌خوانند (پاکش نکنید؛ اگر "
         "حذف شود همهٔ نمودارها خالی می‌شوند)؛ و «دادهٔ خام پرواز» آخرین شیت، کامل و دست‌نخورده است.",
         ),
        ("با چه ترتیبی این فایل را بخوانم؟",
         "تقریباً سه دقیقه: ۱) چهار کارتِ بالای داشبورد (اوج، بیشینهٔ سرعت، بیشینهٔ شتاب، سرعت فرود) "
         "۲) شیت «نتیجه‌گیری و آموزش» -- بخشِ «این پرواز در سه جمله» و «کارنامهٔ پرواز» ۳) اگر عددی "
         "مشکوک بود، همان لحظه را در نمودارِ زمانیِ مربوطه پیدا کنید ۴) برایِ حسابِ دستی، سطرِ همان "
         "زمان را در «دادهٔ خام پرواز» باز کنید. بقیهٔ شیت‌ها (راهنما و شیتِ فنی) مرجع‌اند و لازم نیست "
         "خط‌به‌خط خوانده شوند."),
        ("«سقفِ مرجع» در گیج‌ها یعنی چه؟",
         "هر نیم‌دایره یک سقفِ طراحیِ دلخواه دارد (اوج ۳۰۰۰ متر، سرعت ۲۰۰ م/ث، شتاب ۲۰g، سرعت فرود "
         "۲۵ م/ث) که برایِ راکت‌های آموزشیِ معمول انتخاب شده است. درصدِ گیج یعنی «چقدر به آن سقف "
         "نزدیک شده‌اید»، نه «چقدر بد بوده»: گیجِ ۹۵٪ برایِ اوج یعنی پروازِ بلندی داشته‌اید و گیجِ "
         "۱۵٪ یعنی پرواز کوتاه بوده. برایِ قضاوتِ ایمنی، رنگِ پایینِ همان گیج (سبز/کهربایی/قرمز) را "
         "ببینید، نه درصدِ آن را."),
        ("چطور یک نقطهٔ خاص را در جدول پیدا کنم؟",
         "در شیت «دادهٔ خام پرواز» کلید Ctrl+F را بزنید و مقدارِ زمان را جست‌وجو کنید (مثلاً 18.4). "
         "اگر ستونِ موردنظر را انتخاب کنید (Ctrl+Space)، نوارِ وضعیتِ پایینِ اکسل MIN/MAX/میانگینِ همان "
         "ستون را نشان می‌دهد. سرستون‌هایِ این شیت فیلتر دارند، پس می‌توانید مثلاً فقط سطرهایِ "
         "«مرحلهٔ رانش» را نگه دارید و آمارِ همان بازه را ببینید."),
        ("تبدیلِ واحدِ ذهنی (چون m/s ملموس نیست)",
         "قاعده: m/s × ۳٫۶ = km/h. چند مرجعِ مفید: ۱۰ م/ث ≈ ۳۶ کیلومتر بر ساعت (دوچرخهٔ سریع)، "
         "۳۰ م/ث ≈ ۱۰۸ کیلومتر بر ساعت (ماشین در بزرگراه)، سرعتِ صوت ≈ ۳۴۰ م/ث. برایِ فرود: ۳ م/ث مثلِ "
         "افتادن از ارتفاعِ کمر روی چمن است و ۱۰ م/ث یک ضربهٔ جدی؛ به همین دلیل ۳ تا ۸ م/ث محدودهٔ "
         "ایمنِ معمول شناخته می‌شود."),
        ("میان‌یابیِ نمودارها چه وقت گول‌زننده است؟",
         # هر دو عدد از خودِ ثابت ساخته می‌شود؛ نسخهٔ قبل «~۱۴۰ نقطه» دست‌نویس بود و
         # با MAX_CHART_POINTS = 1400 فرق داشت (راهنمایِ فایل باید راست بگوید).
         "نمودارها برایِ روان ماندنِ فایل حداکثر ~" + fa(MAX_CHART_POINTS) + " نقطه می‌گیرند؛ "
         "اگر پروازی ۵۰هزار سطر داشته باشد، هر نقطهٔ نمودار میانگینِ ~"
         + fa(int(round(50000.0 / MAX_CHART_POINTS))) + " سطر است. یعنی یک نوسانِ یک‌لحظه‌ایِ خیلی کوتاه (مثلاً شکستنِ "
         "بالچه) ممکن است در نمودار دیده نشود ولی در «دادهٔ خام پرواز» باشد. قضاوتِ نهایی همیشه با "
         "شیتِ خام است، نه با نمودار."),
        ("سطرهایِ توضیحِ بالایِ هر شیت را جدی بگیرید",
         "رویِ همهٔ شیت‌ها، یک یا دو سطرِ اول «راهنمایِ همان شیت» است (مثلاً در شیتِ فنی نوشته شده که "
         "منبعِ دادهٔ نمودارهاست). اگر فایل را برایِ کسی می‌فرستید، همان سطرهایِ کمکی هم با فایل "
         "می‌روند و نیازی به توضیحِ جداگانه نیست."),
        ("اعداد فارسی",
         "قالب همهٔ سلول‌های عددی [DBNum1] است: مقدارِ داخل سلول عدد واقعی است و فقط «نمایش» فارسی "
         "است؛ پس فرمول، Pivot و مرتب‌سازی درست کار می‌کند. اگر ارقام را لاتین دیدید، ربطی به این "
         "فایل ندارد: در ویندوز «Regional Formats → Additional settings → Numerals» را روی Hindi "
         "بگذارید تا نمایشِ فارسیِ ارقام در همه‌جا فعال شود."),
        ("فونتِ فایل",
         "متنِ همهٔ سلول‌ها و همهٔ نوشته‌هایِ داخلِ نمودارها «Shabnam» است -- همان فونتی که "
         "خودِ نرم‌افزار و گزارش PDF دارند (فایل‌هایش در پوشهٔ assets برنامه است: Shabnam.ttf و "
         "Shabnam-Bold.ttf). اکسل برخلاف برنامه نمی‌تواند فونت را داخلِ فایل بگذارد، پس اگر "
         "Shabnam روی آن رایانه نصب نباشد، اکسل جایش یک فونتِ فارسیِ دیگر می‌گذارد: فقط *ظاهر* "
         "فرق می‌کند و هیچ عدد یا نموداری خراب نمی‌شود. برایِ همان شکلی که در برنامه می‌بینید، "
         "رویِ همان دو فایل ttf راست‌کلیک → Install را بزنید."),
        ("خطوطِ شبکهٔ نمودار",
         "نمودارها عمداً بدون خط‌کشیِ افقی (Gridlines) رسم شده‌اند تا چشم‌نواز و قابل‌چاپ باشند. برای "
         "دیدنِ مقدارِ دقیقِ هر نقطه نشانگر را روی منحنی بگذارید (Tooltip اکسل) یا از همان سطرِ جدول "
         "زیرِ نمودار استفاده کنید."),
        ("چیدمان راست‌به‌چپ",
         "گزینهٔ Right-to-left در همهٔ شیت‌ها روشن است؛ سربرگ جدول‌ها از راست شروع می‌شود و اکسل "
         "محور افقی نمودارها را هم راست‌به‌چپ می‌چیند."),
        ("محدودهٔ ایمن سرعت فرود",
         f"بین {fa_num(3, 0)} تا {fa_num(8, 0)} متر بر ثانیه ایمن است؛ کمتر از ۳ یعنی چتر بزرگ است "
         f"(رانش باد بیشتر) و بیشتر از ۸ یعنی ضربهٔ فرود خطرناک. سرعت فرود این پرواز: "
         + (fa_num(landing, 1) + " m/s" if landing is not None else "--") + "."),
        ("چرا چگالی هوا جدا محاسبه شده؟",
         "برای فشار دینامیکی دقیق‌تر، چگالی از فشار و دمایِ همان لحظه با رابطهٔ "
         "ρ = p / (R·T) محاسبه و در ستون «چگالی هوا» آورده شده است، نه عدد ثابت ۱٫۲۲ سطح دریا."),
        ("رطوبت و اشعه را کی می‌خوانم؟",
         "به‌محض این‌که فریمور ستون‌های Humidity یا UV_Index/Irradiance را در CSV بنویسد، نمودار و "
         "ستون آن‌ها خودکار به همین گزارش اضافه می‌شود (بدون هیچ تغییر در این فایل)."),
        ("نمودارها",
         f"تعداد نمودارهای این فایل: {fa(charts)}. همه بومیِ اکسل‌اند (تصویر نیستند)؛ برای تغییر رنگ یا "
         "اندازه روی نمودار راست‌کلیک کنید. اگر فایل را با LibreOffice باز کنید ممکن است فونت محورها "
         "متفاوت دیده شود."),
        ("چاپ",
         "همهٔ شیت‌ها روی افقی و «یک صفحه عرض» تنظیم شده‌اند. تم روشن برای چاپ کم‌مصرف جوهر ساخته "
         "شده؛ برای نسخهٔ تیره (هم‌رنگ داشبورد برنامه) برنامه را با export_excel(path, dark=True) "
         "بگیرید."),
        ("گیج‌ها و ستون‌های پنهان",
         "نیم‌دایره‌های «درصد رسیدن به سقف مرجع» در شیت داشبورد، برچسبِ درصدیِ نمودارهای شیت "
         "«تحلیل مراحل پرواز» و ستونِ «ابتدایِ بازه» در شیت «پروفایل ارتفاع» (که فقط برایِ محورِ "
         "افقیِ نمودارها عدد می‌دهد) داده‌شان را در ستون‌های پنهان (سمت راستِ محدودهٔ چاپ، و ستون Pِ "
         "همان شیت) نگه می‌دارند. همهٔ نمودارها روی «نمایش دادهٔ سلول‌های پنهان» تنظیم شده‌اند؛ پس حذف "
         "یا Unhideِ آن ستون‌ها نمودارها را خراب یا شلوغ می‌کند — دست نزنید."),
        ("چرا عدد روی برش‌های نمودار نوشته نشده؟",
         "برچسبِ دادهٔ داخل نمودار (dLbls) در شیت «تحلیل مراحل پرواز» هم خواندن را سخت می‌کرد و هم "
         "در برخی نسخه‌های اکسل باعث پیام «Repair» می‌شد؛ حالا درصدِ هر مرحله در «راهنمای زیرِ "
         "نمودار» نوشته می‌شود (مثلاً «سیر صعودی آزاد — ۳۲٪») و عددِ دقیق در همان جدولِ بالاست."),
        ("نقطهٔ رنگیِ پیشنهادها",
         "در کارت «ارزیابی خودکار»، رنگ نقطهٔ ابتدای هر پیشنهاد فقط خوانایی را زیاد می‌کند "
         "(قرمز = نیازمند اقدام، کهربایی = قابل بررسی، سبز = مطلوب) و هیچ محاسبه‌ای به آن وابسته نیست."),
    ]
    for title, text in items:
        lab = _merge(ws, r, 2, r + 1, 2, title)
        lab.font = _font(theme, 10.5, True, _hex(_paint(theme, ALTITUDE)))
        lab.alignment = _align("right", vert="top", wrap=True)
        body = _merge(ws, r, 3, r + 1, 6, text)
        body.font = _font(theme, 10)
        body.alignment = _align("right", wrap=True, vert="top")
        for rr in (r, r + 1):
            for c in range(2, 7):
                ws.cell(row=rr, column=c).fill = _fill(theme["panel"])
                ws.cell(row=rr, column=c).border = _box(theme)
        ws.row_dimensions[r].height = 17
        avail = sum((ws.column_dimensions[get_column_letter(c)].width or 8.43) for c in range(3, 7))
        lines = max(2, int(math.ceil(_display_width(text, 1) / max(60.0, avail * 0.97))))
        # ۱۵٫۲ = قدِ خطِ فونتِ ۱۱ (قبلاً ۱۳٫۵ برایِ ۹٫۵pt بود: دو خطِ آخرِ هر
        # توضیح بی‌صدا بریده می‌شد)
        ws.row_dimensions[r + 1].height = 15.2 * lines + 5
        r += 3
    if ctx.notes:
        r = _band_title(ws, r, 2, 6, "نکته‌های تولید این فایل", theme, COLOR_WARN)
        for note in ctx.notes:
            r = _note(ws, r, 2, 6, "•  " + note, theme, height=32)
    r += 1
    stamp = _merge(ws, r, 2, r, 6,
                   "تولیدشده توسط «نرم‌افزار مدیریت و تحلیل کامپیوتر پرواز راکت» — "
                   f"زمان تولید: {fa(datetime.datetime.now().strftime('%Y/%m/%d %H:%M'))}")
    stamp.font = _font(theme, 9, False, theme["dim"])
    stamp.alignment = _align("right")


# =========================================================================
# نقطهٔ ورود
# =========================================================================
def _build_prediction_sheet(wb, ctx: "_Context", theme: Dict[str, str], prediction) -> bool:
    """شیت «پیش‌بینی در برابر واقعیت» -- کامل‌ترین نسخهٔ مقایسه با توضیحات.

    جدول کمیت‌ها (پیش‌بینی/بازهٔ محتمل/واقعیت/اختلاف/ارزیابی) + علت‌یابی حدودی
    با متن کامل + روش محاسبه و محدودیت‌ها. اگر اسنپ‌شات لحظهٔ پرتاب موجود
    نباشد، شیت اصلاً ساخته نمی‌شود (به‌جای شیتِ خالیِ گمراه‌کننده)."""
    from core.prediction_compare import compare_snapshot  # noqa: F401  (بدون Qt)

    if not prediction:
        return False
    comp = compare_snapshot(prediction, ctx.results)
    if not comp.get("available"):
        return False

    ws = wb.create_sheet("پیش‌بینی در برابر واقعیت")
    _init_sheet(ws, theme, tab_color=COLOR_INFO, bg_rows=90, bg_cols=9)
    _widths(ws, [1.6, 24, 13, 20, 13, 15, 22, 8])
    C1, C2 = 2, 8
    r = _band_title(ws, 2, C1, C2, "پیش‌بینی در برابر واقعیت -- نرم‌افزار چه گفت، پرواز چه شد؟",
                    theme, COLOR_INFO,
                    "پیش‌بینی‌ها از شبیه‌سازی فیزیکی با پارامترهای فرمِ «لحظهٔ پرتاب» (اسنپ‌شات) "
                    "هستند و واقعیت از تله‌متری همین فایل؛ ستون «بازهٔ محتمل» تلرانس‌های واقعی "
                    "موتور/وزن/درگ را حساب می‌کند. علت‌یابی «حدودی» است -- بدون بادسنج و "
                    "تله‌متری موتور، قطعیت ممکن نیست.")
    r = _note(ws, r + 1, C1, C2,
              f"زمان ثبت اسنپ‌شات پیش‌بینی: {fa(comp.get('meta', {}).get('created_at', '--'))}",
              theme, height=20)

    # ---------------- ۱) جدول مقایسه ----------------
    r = _section(ws, r + 1, C1, C2, "۱) جدول مقایسه -- کمیت به کمیت", theme,
                 sub="«اختلاف» نسبت به عدد پیش‌بینی است (منفی = کمتر از پیش‌بینی). "
                     "اگر واقعیت داخل بازهٔ محتمل باشد، ارزیابی سبز است حتی اگر با عدد "
                     "مرکزی فاصله داشته باشد -- پیش‌بینی یک عدد نیست، یک بازه است.")
    headers = ["کمیت", "پیش‌بینی", "بازهٔ محتمل (۱۰-۹۰٪)", "واقعیت", "اختلاف", "ارزیابی"]
    hrow = r + 1
    for i, h in enumerate(headers):
        cell = ws.cell(row=hrow, column=C1 + i, value=fa(h))
        cell.font = _font(theme, 10, True, theme["band_text"])
        cell.alignment = _align("center", wrap=True)
        cell.fill = _fill(theme["band"])
        cell.border = _box(theme)
    ws.row_dimensions[hrow].height = 24
    rr = hrow
    tone_of = {"ok": COLOR_OK, "minor": COLOR_WARN, "major": COLOR_ERROR,
               "nodata": theme["dim"]}
    for row in comp["rows"]:
        rr += 1
        pred = row["pred"]; actual = row["actual"]; band = row["band"]
        dev = row["dev_pct"]
        if dev is None:
            dev_txt = "--"
        else:
            dev_txt = "\u202a{}{}\u202c".format("+" if dev > 0 else "\u2212", fa_num(abs(dev), 0) + "٪")
        cells = [
            fa(row["label"]),
            ("\u202a{} {}\u202c".format(fa_num(pred, 1), row["unit"]) if pred is not None else "--"),
            ("\u202a{}\u202c تا \u202a{} {}\u202c".format(fa_num(band[0], 0), fa_num(band[2], 0), row["unit"]) if band else "--"),
            ("\u202a{} {}\u202c".format(fa_num(actual, 1), row["unit"]) if actual is not None else "--"),
            dev_txt,
            fa(row["verdict"]),
        ]
        for i, v in enumerate(cells):
            cell = ws.cell(row=rr, column=C1 + i, value=v)
            cell.font = _font(theme, 10, i in (0, 5))
            cell.alignment = _align("right" if i == 0 else "center", wrap=True, indent=1 if i == 0 else 0)
            cell.border = _box(theme)
            if i == 5:
                cell.font = _font(theme, 10, True, tone_of.get(row["kind"], theme["text"]))
        ws.row_dimensions[rr].height = 21
    # ردیف‌های مرجع (فقط پیش‌بینی -- معادل واقعی برای مقایسه ندارند)
    rr += 1
    for ref in comp["ref_rows"]:
        note = "مرجع -- معادل واقعی ندارد"
        cells = [fa(ref["label"]),
                 ("\u202a{} {}\u202c".format(fa_num(ref["pred"], 1), ref["unit"]) if ref["pred"] is not None else "--"),
                 "--", "--", "--", note]
        for i, v in enumerate(cells):
            cell = ws.cell(row=rr, column=C1 + i, value=fa(v) if i == 5 else v)
            cell.font = _font(theme, 9.5, i == 0)
            cell.alignment = _align("right" if i == 0 else "center", wrap=True)
            cell.border = _box(theme)
            if i == 5:
                cell.font = _font(theme, 9.5, False, theme["dim"])
        ws.row_dimensions[rr].height = 19
        rr += 1
    r = rr + 1

    # ---------------- ۲) علت‌یابی ----------------
    r = _section(ws, r, C1, C2, "۲) علت‌یابی حدودی -- محتمل‌ترین توضیح‌ها", theme,
                 sub="قاعده‌محور و از روی الگوی انحراف‌ها؛ لحن «احتمالاً» عمدی است "
                     "چون چند علتِ متفاوت می‌توانند همان الگو را بسازند.")
    sev_tone = {"ok": COLOR_OK, "warn": COLOR_WARN, "danger": COLOR_ERROR,
                "info": theme["text"]}
    sev_mark = {"ok": "✅", "warn": "⚠️", "danger": "🛑", "info": "ℹ️"}
    for cause in comp["causes"]:
        r = _note(ws, r, C1, C2,
                  f"{sev_mark.get(cause['severity'], '•')}  {cause['title']}",
                  theme, height=20, bold=True,
                  tone=sev_tone.get(cause["severity"]))
        r = _note(ws, r, C1, C2, cause["text"], theme, height=34)

    # ---------------- ۳) روش و محدودیت‌ها ----------------
    r = _section(ws, r + 1, C1, C2, "۳) روش محاسبه و محدودیت‌ها (صادقانه)", theme)
    r = _note(ws, r, C1, C2, comp["method"], theme, height=44)
    r = _note(ws, r, C1, C2, comp["limits"], theme, height=40)
    return True

def export_excel(path: str, *, mission=None, motor=None, results=None, df=None,
                 events=None, suggestions=None, dark: bool = False,
                 prediction=None) -> Dict[str, Any]:
    """ساخت فایل اکسل داشبوردیِ گزارش پرواز.

    همهٔ پارامترها اختیاری‌اند و در صورت نبودن از data_manager خوانده می‌شوند
    (پس فراخوانی قدیمی export_excel(path) بدون تغییر کار می‌کند). اگر همه را
    صریح بدهید، این تابع هیچ‌وقت data_manager را import نمی‌کند و در نتیجه
    بدون PySide6/Qt هم کار می‌کند -- همان چیزی که برای استفادهٔ جداگانهٔ
    موتور گزارش در برنامه‌های دیگر لازم است.
    تم پیش‌فرض روشن است (سبک داشبورد چاپی: کارت سفید، کادر طوسی، نوار نارنجی)؛
    dark=True همان ظاهر تیرهٔ داشبورد برنامه را می‌سازد. مقدار بازگشتی خلاصهٔ کار است
    (شیت‌های ساخته‌شده، تعداد نمودارها و یادداشت‌ها) که برای لاگ/دیباگ مفید است.
    """
    # core.data_manager یک QObject است و import‌کردنش PySide6 را با خودش
    # می‌آورد. اگر فراخوان‌کننده همهٔ ورودی‌ها را خودش بدهد (کاری که
    # اسکریپت‌ها، تست‌ها و برنامه‌های دیگر هنگام استفادهٔ جداگانه از این
    # موتور می‌کنند) هیچ نیازی به data_manager و در نتیجه به Qt نیست؛ پس
    # import فقط وقتی انجام می‌شود که واقعاً ورودی‌ای جا افتاده باشد.
    def _dm(name: str, default=None):
        from core.data_manager import data_manager
        return getattr(data_manager, name, default)

    mission = _dm("mission") if mission is None else mission
    motor = _dm("motor") if motor is None else motor
    results = _dm("analysis_results") if results is None else results
    df = _dm("flight_df") if df is None else df
    events = _dm("events") if events is None else events
    if suggestions is None:
        try:
            from core.advisor import generate_suggestions
            suggestions = generate_suggestions(df, results, mission, motor)
        except Exception:
            suggestions = []
    prediction = _dm("prediction_snapshot") if prediction is None else prediction

    _CHART_WARNINGS.clear()
    _CHART_BOXES.clear()          # کادرِ نمودارهایِ این فایل (از فراخوانیِ قبلی نماند)
    theme = _theme(dark)
    ctx = _Context(mission, motor, results, df, events, suggestions)
    summary: Dict[str, Any] = dict(sheets=[], charts=0, rows_raw=0, notes=[], theme="dark" if dark else "light")

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "گزارش تحلیل پرواز راکت"
    wb.properties.creator = "نرم‌افزار مدیریت و تحلیل کامپیوتر پرواز راکت"
    wb.properties.description = "داشبورد Excel -- اعداد فارسی، چیدمان راست‌به‌چپ و نمودارهای بومی"
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    def _safe(name, fn):
        """ساخت یک شیت با محافظ خطا. سازنده‌ای که چیزی نسازد False برمی‌گرداند
        تا نامش هم در خلاصه نیاید (وگرنه «شیت ساخته شد»-but-empty گمراه‌کننده است)."""
        try:
            out = fn()
            if out is not False:
                summary["sheets"].append(name)
            return out
        except Exception as exc:
            ctx.notes.append(f"ساخت شیت «{name}» ناموفق بود ({type(exc).__name__}: {exc}).")
            return None

    # اگر تحلیلی در کار نباشد، شیت‌های تحلیلی «خطا» نمی‌دهند بلکه کلاً ساخته
    # نمی‌شوند (شرط‌های ctx.ok داخل سازنده‌ها) -- یادداشتش در راهنما می‌آید.
    sws = _safe(SERIES_SHEET, lambda: _build_series_sheet(wb, ctx, theme)) if ctx.series is not None else None
    cols = {name: i + 1 for i, name in enumerate(ctx.series.columns)} if ctx.series is not None else {}

    summary["charts"] += _safe("داشبورد پرواز", lambda: _build_dashboard(wb, ctx, theme, sws, cols)) or 0
    if sws is not None:
        summary["charts"] += _safe("نمودارهای زمانی",
                                   lambda: _build_time_charts(wb, ctx, theme, sws, cols)) or 0
        summary["charts"] += _safe("پروفایل ارتفاع", lambda: _build_altitude_profile(wb, ctx, theme)) or 0
    summary["charts"] += _safe("تحلیل مراحل پرواز", lambda: _build_phases(wb, ctx, theme)) or 0
    # شیتِ «نتیجه‌گیری و آموزش» عمداً بعد از داشبورد و *قبل* از نمودارها می‌آید:
    # کاربرِ تازه‌کار اول باید بفهمد چه اتفاقی افتاده، بعد سراغِ نمودار برود.
    _safe("نتیجه‌گیری و آموزش", lambda: _build_conclusion(wb, ctx, theme))
    _safe("پیش‌بینی در برابر واقعیت",
          lambda: _build_prediction_sheet(wb, ctx, theme, prediction))
    if ctx.df is not None:
        summary["rows_raw"] = _safe("دادهٔ خام پرواز", lambda: _build_raw_sheet(wb, ctx, theme)) or 0
    _safe("راهنمای خواندن", lambda: _build_guide(wb, ctx, theme, summary["charts"] + 3))

    if not wb.worksheets:
        ws = wb.create_sheet("خلاصه")
        _init_sheet(ws, theme)
        _widths(ws, [2.2, 95])
        cell = ws.cell(row=2, column=2, value="گزارش اکسل تولید شد ولی جدول‌های تحلیلی خالی ماند. "
                                             "لطفاً فایل CSV پرواز را بارگذاری کنید.")
        cell.font = _font(theme, 11)
    # داشبورد همیشه اولین شیت و انتخاب‌شده باشد (شیت فنیِ نمودارها یکی مانده به آخر)
    # ترتیب زبانه‌ها: اول شیت‌های خواندنی، بعد دادهٔ خام و در آخر شیت فنیِ
    # «دادهٔ نمودارها (فنی)» منبع دادهٔ نمودارهاست و «دادهٔ خام پرواز» آخرین شیت است
    # ترتیبِ شیت‌ها: اول خواندنی‌ها، بعد راهنما و شیتِ فنی، و در آخر «دادهٔ خام»
    # (درخواستِ کاربر: شیتِ دادهٔ خام آخرین شیت باشد).
    order = {"داشبورد پرواز": 0, "نتیجه‌گیری و آموزش": 1, "پیش‌بینی در برابر واقعیت": 2,
             "نمودارهای زمانی": 3, "پروفایل ارتفاع": 4, "تحلیل مراحل پرواز": 5,
             "راهنمای خواندن": 6, SERIES_SHEET: 7, "دادهٔ خام پرواز": 8}
    wb._sheets.sort(key=lambda s: order.get(s.title, 7))
    wb.active = 0
    for s in wb.worksheets:
        s.sheet_view.tabSelected = (s.title == "داشبورد پرواز")
    if _CHART_WARNINGS:
        ctx.notes.append("برخی تنظیمات ظاهری نمودار اعمال نشد (داده‌ها و نمودارها "
                         "سالم‌اند، فقط رنگ/قالب نوشته‌ها): " + " | ".join(_CHART_WARNINGS[:4]))
    # درخواست کاربر: همهٔ سلول‌های عددیِ کل کارنامه وسط‌چین باشند (بقیهٔ
    # ویژگی‌های ترازِ موجود مثل شکستن خط و تراز عمودی حفظ می‌شود).
    for ws_num in wb.worksheets:
        for row_cells in ws_num.iter_rows():
            for cell in row_cells:
                v = cell.value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    a = cell.alignment
                    cell.alignment = Alignment(horizontal="center", vertical=a.vertical,
                                               wrap_text=a.wrap_text, indent=a.indent)

    wb.save(path)
    # مثلث‌هایِ زردِ «number stored as text» (بلافاصله بعد از save؛ فقط XMLِ
    # شیت‌ها دست‌کاری می‌شود و اگر نشد، گزارش همان فایلِ سالم می‌ماند).
    try:
        summary["theme_font"] = _patch_theme_font(path, FA_FONT_NAME)
    except Exception as exc:          # فونتِ تمِ فایل اختیاری است، شرطِ سلامتِ فایل نیست
        ctx.notes.append(f"فونتِ تم اعمال نشد ({type(exc).__name__}: {exc}).")
    try:
        summary["ignored_errors"] = _patch_ignored_errors(path)
    except Exception as exc:
        ctx.notes.append(f"حذف اخطارِ «عدد به‌صورت متن» اعمال نشد "
                         f"({type(exc).__name__}: {exc}).")
    summary["notes"] = list(ctx.notes)
    return summary
